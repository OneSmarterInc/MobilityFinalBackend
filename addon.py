
import json
import os
import zipfile
import pandas as pd
import shutil

def parse_until_dict(data):
    while isinstance(data, str):
        try:
            data = json.loads(data)
        except json.JSONDecodeError:
            break
    return data

def str_to_bool(value):
    print(value)
    return str(value).strip().lower() in ['true','yes']

import re

def remove_digits_and_symbols(text):
    
    return re.sub(r'[^a-zA-Z\s]', '', text)
def string_similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    min_len = min(len(a), len(b))
    matches = sum(1 for i in range(min_len) if a[i] == b[i])
    return matches / max(len(a), len(b))


def get_close_match_key(target_key: str, key_list: list[str]) -> str | None:
    closest = ""
    max_similarity = 0.0

    for candidate in key_list:
        target_key_enhanced = remove_digits_and_symbols(target_key)
        candidate_enhanced = remove_digits_and_symbols(candidate)
        sim = string_similarity(target_key_enhanced, candidate_enhanced)
        if sim > 0.8 and sim > max_similarity:
            max_similarity = sim
            closest = candidate

    return closest if closest else None

def compare(str1, str2):
    str1lst = str1.split()
    str2lst = str2.split()
    points = 0
    for i in str2lst:
       if i in str1lst: points+=1 
    percent = (points/len(str2lst)) * 100
    return int(percent) > 80
    


def extract_data_from_zip(path):
    extract_dir = 'Bills/media/extracted_files'
    os.makedirs(extract_dir, exist_ok=True)
    dataframes = []

    try:
        with zipfile.ZipFile(path, 'r') as zip_ref:
            zip_ref.extractall(extract_dir)
    except zipfile.BadZipFile:   # correct exception name in py3
        return {'message': 'Invalid zip file', 'error': -1}
    except Exception as e:
        print("Error extracting zip data=", str(e))
        return {'message': str(e), 'error': -1}

    for root, _, files in os.walk(extract_dir):
        for file in files:
            file_path = os.path.join(root, file)

            try:
                if file.endswith((".xlsx", ".xls")):
                    df = pd.read_excel(file_path)
                    dataframes.append(df)

                elif file.endswith(".csv"):
                    df = pd.read_csv(file_path)
                    dataframes.append(df)

                elif file.endswith(".txt"):
                    df = pd.read_csv(file_path, delimiter="\t", engine="python")
                    dataframes.append(df)

            except Exception as e:
                print(f"Skipping {file}: {e}")
    
    shutil.rmtree(extract_dir)
    
    return [df for df in dataframes if not df.empty]

def get_cat_obj_total(obj):
    obj = parse_until_dict(obj)
    total = 0
    for key, value in obj.items():
        for c,v in value.items(): 
            total += float(v['amount']) if isinstance(v, dict) else float(v)
    return round(total,2)

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def auto_width_excel(file_path, output_path=None):
    """
    Adjust column width automatically for all sheets in an Excel file.
    """

    # If no output_path given, overwrite the same file
    if output_path is None:
        output_path = file_path

    wb = load_workbook(file_path)

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    value = str(cell.value) if cell.value is not None else ""
                    if len(value) > max_length:
                        max_length = len(value)
                except:
                    pass

            sheet.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_path)
    return output_path


# cleaned = auto_width_excel(
#     "Bills/media/ViewUploadedBills/836086478_January_2024_mArYUpU.xlsx",
# )