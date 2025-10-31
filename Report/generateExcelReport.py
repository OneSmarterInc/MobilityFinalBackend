import io
import pandas as pd
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class GenerateExcelReport:
    def __init__(self, report_name, data):
        self.report_name = report_name
        self.data_df = pd.DataFrame(data)

        # ✅ Remove 'id' column if it exists
        if 'id' in self.data_df.columns:
            self.data_df.drop(columns=['id'], inplace=True)
        for col in self.data_df.select_dtypes(include=['datetimetz']).columns:
            self.data_df[col] = self.data_df[col].dt.tz_convert(None)

    def _generate(self, note_text="Note: This is a system-generated report."):
        output = io.BytesIO()
        file_name = f"{self.report_name}.xlsx"

        # ✅ Write data to in-memory Excel file
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            self.data_df.to_excel(writer, index=False, startrow=3)
            wb = writer.book
            ws = writer.sheets['Sheet1']

            # ✅ Add title row (merged + styled)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.data_df.columns))
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = self.report_name
            title_cell.font = Font(size=16, bold=True, color="FFFFFF")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            ws.row_dimensions[1].height = 30

            # ✅ Style header row
            header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            header_font = Font(bold=True)
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            header_row = 4
            for cell in ws[header_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # ✅ Auto adjust column widths
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = max_length + 4
                ws.column_dimensions[col_letter].width = adjusted_width

            # ✅ Add note at end
            note_row = ws.max_row + 2
            ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(self.data_df.columns))
            note_cell = ws.cell(row=note_row, column=1)
            note_cell.value = note_text
            note_cell.font = Font(italic=True, color="808080")
            note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.row_dimensions[note_row].height = 25

        # ✅ Set response headers for download
        output.seek(0)
        return output, file_name

