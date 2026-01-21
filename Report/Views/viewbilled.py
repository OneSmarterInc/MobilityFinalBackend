# from rest_framework.views import APIView
# from rest_framework import status
# from rest_framework.response import Response
# from rest_framework import permissions
# from authenticate.views import saveuserlog
# from OnBoard.Organization.models import Organizations
# from OnBoard.Company.models import Company
# from Dashboard.ModelsByPage.DashAdmin import Vendors
# from OnBoard.Ban.models import BaseDataTable


# from django.db.models.functions import Cast
# from django.db.models import FloatField

# import matplotlib.pyplot as plt
# from openpyxl import load_workbook
# from openpyxl.drawing.image import Image
# from openpyxl.styles import PatternFill
# from io import BytesIO
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import Font, Alignment, PatternFill
# from openpyxl.styles import Font, Alignment, PatternFill
# from openpyxl.styles import Font, Alignment, PatternFill
# from openpyxl.styles import Font, Alignment, PatternFill
# from openpyxl.styles import Font, Alignment, PatternFill
# from django.db.models import Q
# from django.core.files.base import ContentFile
# from ..models import Report_Billed_Data
# import pandas as pd
# from ..ser import showBanSerializer, showBilledReport, OrganizationShowSerializer, VendorShowSerializer

# from openpyxl import Workbook
# from datetime import datetime
# from ..models import Downloaded_reports
# from collections import defaultdict
# class ViewBilledReportView(APIView):
#     permission_classes = [permissions.IsAuthenticated]

#     def __init__(self, **kwargs):
#         self.report_type = "Billed_Data_Usage"
#         self.current_year = datetime.now().year
#         self.data = None
#         self.unique_dates = None
#         super().__init__(**kwargs)

#     def get(self, request,sub_type=None, *args, **kwargs):
#         if request.user.designation.name == 'Admin':
#             orgs = OrganizationShowSerializer(Organizations.objects.filter(company=request.user.company), many=True)
#             vendors = VendorShowSerializer(Vendors.objects.all(), many=True)
#             bans = showBanSerializer(BaseDataTable.objects.filter(company=request.user.company, viewuploaded=None, viewpapered=None), many=True)
#         else:
#             orgs = OrganizationShowSerializer(Organizations.objects.all(), many=True)
#             vendors = VendorShowSerializer(Vendors.objects.all(), many=True)
#             bans = showBanSerializer(BaseDataTable.objects.filter(viewuploaded=None, viewpapered=None), many=True)

#         if not sub_type:
#             self.data = showBilledReport(Report_Billed_Data.objects.all(), many=True)
#         else:
#             company = request.GET.get('company')
#             sub_company = request.GET.get('sub_company')
#             report_type = request.GET.get('report_type')
#             month = request.GET.get('month')
#             vendor = request.GET.get('vendor')
#             year = request.GET.get('year')
#             if not (company and sub_company and report_type and month and vendor and year):
#                 return Response({"message":"Missing required fields."},status=status.HTTP_400_BAD_REQUEST)
#             filter_kwargs = {}
#             if company:
#                 filter_kwargs['company'] = Company.objects.filter(Company_name=company)[0]
#             if sub_company:
#                 filter_kwargs['organization'] = Organizations.objects.filter(Organization_name=sub_company)[0]
#             if report_type:
#                 filter_kwargs['Report_Type'] = report_type
#             if month:
#                 filter_kwargs['Month'] = month
#             if year:
#                 filter_kwargs['Year'] = year
#             if vendor:
#                 filter_kwargs['vendor'] = Vendors.objects.filter(name=vendor)[0]
#             filtered_data = Report_Billed_Data.objects.filter(company=filter_kwargs['company'])
#             if not filtered_data.exists():
#                 return Response({"message":f"no billed report found for {filter_kwargs['company'].Company_name}"},status=status.HTTP_400_BAD_REQUEST)
            
#             filtered_data = filtered_data.filter(organization=filter_kwargs['organization'])
#             if not filtered_data.exists():
#                 return Response({"message":f"no billed report found for {filter_kwargs['organization'].Organization_name}"},status=status.HTTP_400_BAD_REQUEST)
            
#             filtered_data = filtered_data.filter(Report_Type=filter_kwargs['Report_Type'])
#             if not filtered_data.exists():
#                 return Response({"message":f"no billed report found for {filter_kwargs['Report_Type']}"},status=status.HTTP_400_BAD_REQUEST)
            
#             filtered_data = filtered_data.filter(Month=filter_kwargs['Month'])
#             if not filtered_data.exists():
#                 return Response({"message":f"no billed report found for {filter_kwargs['Month']}"},status=status.HTTP_400_BAD_REQUEST)
            
#             filtered_data = filtered_data.filter(Year=filter_kwargs['Year'])
#             if not filtered_data.exists():
#                 return Response({"message":f"no billed report found for {filter_kwargs['Year']}"},status=status.HTTP_400_BAD_REQUEST)
            
#             filtered_data = filtered_data.filter(vendor=filter_kwargs['vendor'])
#             if not filtered_data.exists():
#                 return Response({"message":f"no billed report found for {filter_kwargs['vendor'].name}"},status=status.HTTP_400_BAD_REQUEST)
        
#             if sub_type == "Top10BilledUsers":
#                 try:
#                     filtered_data = filtered_data.annotate(
#                         usage_float=Cast('Data_Usage_GB', FloatField())
#                     ).order_by('-usage_float')[:10]
#                     self.data = showBilledReport(filtered_data, many=True)
#                 except Exception as e:
#                     print(e)
#                     return Response({"message": "Unable to view report."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)   
#             elif sub_type == "getZeroUsageReportbilledData":
#                 try:
#                     filtered_data = Report_Billed_Data.objects.filter(
#                         Q(Data_Usage_GB="0") | Q(Data_Usage_GB="0.0"),
#                         **filter_kwargs
#                     )
#                     if not filtered_data.exists():
#                         return Response({"message": "No data found for the given filters."}, status=status.HTTP_404_NOT_FOUND)
#                     self.data = showBilledReport(filtered_data, many=True)
#                 except Exception as e:
#                     print(e)
#                     return Response({"message": "Unable to view report."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)  

#             else:
#                 return Response({"message": "Invalid sub_type provided."}, status=status.HTTP_400_BAD_REQUEST)
#             saveuserlog(request.user, f"Billed data report excel file of {month}-{year} downloaded.")
#         return Response(
#             {"orgs": orgs.data, "vendors": vendors.data, "bans":bans.data, "data":self.data.data if self.data else None, "unique_dates":self.unique_dates if self.report_type == "Unbilled_Data_Usage" else None},
#             status=status.HTTP_200_OK,
#         )
        
#     def change_kwargs(self,kwargs):
#         if 'organization' in kwargs and str(kwargs['organization']) != str:
#             kwargs['organization'] = kwargs['organization'].Organization_name
#         if 'company' in kwargs and str(kwargs['company']) != str:
#             kwargs['company'] = kwargs['company'].Company_name
#         if 'vendor' in kwargs and str(kwargs['vendor']) != str:
#             kwargs['vendor'] = kwargs['vendor'].name

#         return kwargs

#     def post(self, request,sub_type=None, *args, **kwargs):

#         filter_kwargs = {}
#         action = request.data.get('action')
#         type = request.data.get('type')
#         company = request.data.get('company')
#         sub_company = request.data.get('sub_company')
#         ban = request.data.get('ban')
#         month = request.data.get('month')
#         report_type = request.data.get('report_type')
#         year = request.data.get('year')
#         vendor = request.data.get('vendor')
#         if not (company and sub_company and ban and report_type and month and vendor and year):
#             return Response({"message":"Missing required fields."},status=status.HTTP_400_BAD_REQUEST)
        
#         if company:
#             filter_kwargs['company'] = Company.objects.filter(Company_name=company)[0]
#         if sub_company:
#             filter_kwargs['organization'] = Organizations.objects.filter(Organization_name=sub_company)[0]
#         if ban:
#             filter_kwargs['Account_Number'] = ban
#         if report_type:
#             filter_kwargs['Report_Type'] = report_type
#         if month:
#             filter_kwargs['Month'] = month
#         if year:
#             filter_kwargs['Year'] = year
#         if vendor:
#             filter_kwargs['vendor'] = Vendors.objects.filter(name=vendor)[0]

#         filtered_data = Report_Billed_Data.objects.filter(company=filter_kwargs['company'])
#         if not filtered_data.exists():
#             return Response({"message":f"no billed report found for {filter_kwargs['company'].Company_name}"},status=status.HTTP_400_BAD_REQUEST)
        
#         filtered_data = filtered_data.filter(organization=filter_kwargs['organization'])
#         if not filtered_data.exists():
#             return Response({"message":f"no billed report found for {filter_kwargs['organization'].Organization_name}"},status=status.HTTP_400_BAD_REQUEST)
        
#         filtered_data = filtered_data.filter(Account_Number=filter_kwargs['Account_Number'])
#         if not filtered_data.exists():
#             return Response({"message":f"no billed report found for {filter_kwargs['Account_Number']}"},status=status.HTTP_400_BAD_REQUEST)

#         filtered_data = filtered_data.filter(Report_Type=filter_kwargs['Report_Type'])
#         if not filtered_data.exists():
#             return Response({"message":f"no billed report found for {filter_kwargs['Report_Type']}"},status=status.HTTP_400_BAD_REQUEST)
        
#         filtered_data = filtered_data.filter(Month=filter_kwargs['Month'])
#         if not filtered_data.exists():
#             return Response({"message":f"no billed report found for {filter_kwargs['Month']}"},status=status.HTTP_400_BAD_REQUEST)
        
#         filtered_data = filtered_data.filter(Year=filter_kwargs['Year'])
#         if not filtered_data.exists():
#             return Response({"message":f"no billed report found for {filter_kwargs['Year']}"},status=status.HTTP_400_BAD_REQUEST)
        
#         filtered_data = filtered_data.filter(vendor=filter_kwargs['vendor'])
#         if not filtered_data.exists():
#             return Response({"message":f"no billed report found for {filter_kwargs['vendor'].name}"},status=status.HTTP_400_BAD_REQUEST)
#         if not filtered_data.exists():
#             return Response({"message": "No data found for the given filters."}, status=200)
        
#         if not sub_type:
#             try:
#                 filtered_data = Report_Billed_Data.objects.filter(company=filter_kwargs['company'])
#                 if not filtered_data.exists():
#                     return Response({"message":f"no billed report found for {filter_kwargs['company'].Company_name}"},status=status.HTTP_400_BAD_REQUEST)
                
#                 filtered_data = filtered_data.filter(organization=filter_kwargs['organization'])
#                 if not filtered_data.exists():
#                     return Response({"message":f"no billed report found for {filter_kwargs['organization'].Organization_name}"},status=status.HTTP_400_BAD_REQUEST)
                
#                 filtered_data = filtered_data.filter(Report_Type=filter_kwargs['Report_Type'])
#                 if not filtered_data.exists():
#                     return Response({"message":f"no billed report found for {filter_kwargs['Report_Type']}"},status=status.HTTP_400_BAD_REQUEST)
                
#                 filtered_data = filtered_data.filter(Month=filter_kwargs['Month'])
#                 if not filtered_data.exists():
#                     return Response({"message":f"no billed report found for {filter_kwargs['Month']}"},status=status.HTTP_400_BAD_REQUEST)
                
#                 filtered_data = filtered_data.filter(Year=filter_kwargs['Year'])
#                 if not filtered_data.exists():
#                     return Response({"message":f"no billed report found for {filter_kwargs['Year']}"},status=status.HTTP_400_BAD_REQUEST)
                
#                 filtered_data = filtered_data.filter(vendor=filter_kwargs['vendor'])
#                 if not filtered_data.exists():
#                     return Response({"message":f"no billed report found for {filter_kwargs['vendor'].name}"},status=status.HTTP_400_BAD_REQUEST)
#                 if not filtered_data.exists():
#                     return Response({"message": "No data found for the given filters."}, status=200)

#                 # Create a list of dictionaries to store the data in the required format
#                 report_data = []
#                 for record in filtered_data:
#                     report_data.append({
#                         "accountNumber": record.Account_Number,
#                         'wirelessNumber': record.Wireless_Number,
#                         'userName': record.User_Name,
#                         'vendor': record.vendor.name,
#                         'voicePlanUsage': float(record.Voice_Plan_Usage),  # New column
#                         'messagingUsage': float(record.Messaging_Usage),  # New column
#                         'usage': float(record.Data_Usage_GB),
#                     })

#                 # Convert the data to a pandas DataFrame
#                 df = pd.DataFrame(report_data)

#                 # Categorize data usage into specified ranges
#                 bins = [0, 0.001, 1, 5, 7.5, 12.5, 22, float('inf')]
#                 labels = ['Equal to 0', 'Less than 1 GB', '1-5 GB', '5-7.5 GB', '7.5-12.5 GB', '12.5-22 GB', '22 GB plus']
#                 df['usage_category'] = pd.cut(df['usage'], bins=bins, labels=labels, right=False)

#                 # Reorder columns and include new fields, removing 'date'
#                 df = df[['accountNumber', 'wirelessNumber', 'userName', 'vendor', 'voicePlanUsage', 'messagingUsage', 'usage', 'usage_category']]

#                 # Generate a bar graph for data usage categories
#                 if vendor:
#                     df_vendor = df[df['vendor'] == vendor]
#                 else:
#                     df_vendor = df

#                 usage_counts = df_vendor['usage_category'].value_counts().reindex(labels, fill_value=0)
#                 plt.figure(figsize=(12, 6))
#                 usage_counts.plot(kind='bar')
#                 plt.title(f'Data Usage Frequency for Vendor: {vendor}')
#                 plt.xlabel('Data Usage Category')
#                 plt.ylabel('Frequency')
#                 plt.xticks(rotation=45, ha='right')
#                 plt.tight_layout()  # Adjust the padding to fit the labels

#                 # Save the plot to a BytesIO object
#                 plot_output = BytesIO()
#                 plt.savefig(plot_output, format='png')
#                 plot_output.seek(0)

#                 # Create an Excel writer object and write the DataFrame to it
#                 output = BytesIO()
#                 writer = pd.ExcelWriter(output, engine='openpyxl')
#                 df.to_excel(writer, index=False, sheet_name='Billed Data Report')
#                 writer.close()
#                 output.seek(0)

#                 # Load the workbook and insert the image
#                 workbook = load_workbook(output)
#                 worksheet = workbook['Billed Data Report']
#                 img = Image(plot_output)
#                 worksheet.add_image(img, 'K2')  # Adjust the position as needed

#                 # Insert a dynamic heading
#                 heading = f'{month} Usage Report'
#                 worksheet.insert_rows(1)  # Shift everything down to make space for the header
#                 worksheet.merge_cells('A1:H1')  # Adjust the merge to fit the new columns
#                 heading_cell = worksheet['A1']
#                 heading_cell.value = heading
#                 heading_cell.font = Font(bold=True, color='FFFFFF', size=14)
#                 heading_cell.fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')
#                 heading_cell.alignment = Alignment(horizontal='center', vertical='center')

#                 # Define the background fill colors for 'usage' column based on conditions
#                 red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
#                 light_yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
#                 dark_navy_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")

#                 # Define light blue fill for Voice_Plan_Usage, Messaging_Usage, and usage_category columns
#                 light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

#                 # Apply specific formatting for the 'usage' column header (set text color to black)
#                 usage_header_cell = worksheet["G2"]  # Assuming 'G2' is the header for the usage column
#                 usage_header_cell.font = Font(bold=True, color="000000")  # Set usage header text to black

#                 # Adjust column widths based on the content and header
#                 for col_num, column_cells in enumerate(worksheet.columns, 1):
#                     max_length = 0
#                     column = get_column_letter(col_num)

#                     for cell in column_cells:
#                         try:
#                             if cell.value:
#                                 max_length = max(max_length, len(str(cell.value)))

#                             # Apply background color and text formatting to specific columns
#                             if column == 'G':  # Assuming 'G' is usage column
#                                 usage_value = cell.value
#                                 if isinstance(usage_value, float) or isinstance(usage_value, int):
#                                     # Set content text color to white
#                                     cell.font = Font(color="FFFFFF")
#                                     # Apply conditional formatting based on usage values
#                                     if usage_value == 0:
#                                         cell.fill = red_fill
#                                     elif 0 < usage_value <= 1:
#                                         cell.fill = light_yellow_fill
#                                     elif usage_value > 1:
#                                         cell.fill = dark_navy_fill
#                             elif column in ['E', 'F', 'H']:  # Voice_Plan_Usage, Messaging_Usage, usage_category columns
#                                 cell.fill = light_blue_fill
#                         except:
#                             pass

#                     adjusted_width = (max_length + 2) if max_length < 30 else 30
#                     worksheet.column_dimensions[column].width = adjusted_width

#                 # Save the modified workbook to the output
#                 output_with_image = BytesIO()
#                 workbook.save(output_with_image)
#                 output_with_image.seek(0)
                
#                 report_file = Downloaded_reports(
#                     report_type=report_type,
#                     kwargs=self.change_kwargs(filter_kwargs)
#                 )
#                 report_file.file.save(f'billed_data_report_{month}_{year}.xlsx', ContentFile(output_with_image.getvalue()))

#                 saveuserlog(request.user, f"Billed data report excel file of {month}-{year} downloaded.")
#                 return Response({"data":report_file.file.url, "message" : "Excel file generated sucessfully!"}, status=status.HTTP_200_OK)

#             except Exception as e:
#                 print(e)
#                 return Response({"message": "Unable to view report."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
#         else:
#             if sub_type == "getZeroUsageReportbilledExcel":
#                 try:
#                     filtered_data = Report_Billed_Data.objects.filter(
#                         Q(Data_Usage_GB="0") | Q(Data_Usage_GB="0.0"),
#                         **filter_kwargs
#                     )
#                     if not filtered_data.exists():
#                         return Response({"message": "No data found for the given filters."}, status=status.HTTP_404_NOT_FOUND)

#                     # Extract the unique vendor name from the filtered data
#                     unique_vendor = filtered_data.values_list('vendor', flat=True).distinct()
#                     if unique_vendor:
#                         vendor_name = unique_vendor[0]  # Get the first unique vendor name
#                     else:
#                         vendor_name = "Unknown Vendor"  # Fallback in case there's no vendor

#                     # Create a list of dictionaries to store the data in the required format
#                     report_data = []
#                     for record in filtered_data:
#                         report_data.append({
#                             "accountNumber": record.Account_Number,
#                             'wirelessNumber': record.Wireless_Number,
#                             'userName': record.User_Name,
#                             'dataUsageGB': record.Data_Usage_GB,
#                         })

#                     # Convert the data to a pandas DataFrame
#                     df = pd.DataFrame(report_data)

#                     # Create an Excel writer object and write the DataFrame to it
#                     output = BytesIO()
#                     writer = pd.ExcelWriter(output, engine='openpyxl')
#                     df.to_excel(writer, index=False, sheet_name='Zero Usage Billed Data Report')
#                     writer.close()
#                     output.seek(0)

#                     # Load the workbook to apply formatting
#                     workbook = load_workbook(output)
#                     worksheet = workbook['Zero Usage Billed Data Report']

#                     # Add the main heading for the report (dynamic based on month and vendor)
#                     main_heading = f'{month} Zero Usage Report ({vendor_name})'
#                     worksheet.insert_rows(1)
#                     worksheet.merge_cells('A1:D1')  # Adjust the merge to cover the required columns
#                     heading_cell = worksheet['A1']
#                     heading_cell.value = main_heading
#                     heading_cell.font = Font(bold=True, color="FFFFFF", size=14)
#                     heading_cell.alignment = Alignment(horizontal="center", vertical="center")
#                     heading_cell.fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')  # Dark navy blue background

#                     # Define the background fill for 'dataUsageGB' column
#                     light_yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

#                     # Get the index of the 'dataUsageGB' column (4th column in this case)
#                     data_usage_column_index = df.columns.get_loc('dataUsageGB') + 1  # Pandas index is 0-based, Excel is 1-based
                    
#                     # Apply background color to 'dataUsageGB' column (4th column)
#                     for row in worksheet.iter_rows(min_row=3, min_col=data_usage_column_index, max_col=data_usage_column_index):
#                         for cell in row:
#                             cell.fill = light_yellow_fill

#                     # Adjust column widths based on content and header
#                     for col_num, column_cells in enumerate(worksheet.columns, 1):
#                         max_length = 0
#                         column = get_column_letter(col_num)
                        
#                         for cell in column_cells:
#                             try:
#                                 if cell.value:
#                                     max_length = max(max_length, len(str(cell.value)))
#                             except:
#                                 pass

#                         adjusted_width = (max_length + 2) if max_length < 30 else 30  # Adjust maximum column width to 30 characters
#                         worksheet.column_dimensions[column].width = adjusted_width
#                     output_with_formatting = BytesIO()
#                     workbook.save(output_with_formatting)
#                     output_with_formatting.seek(0)

#                     report_file = Downloaded_reports(
#                         report_type=report_type,
#                         kwargs=self.change_kwargs(filter_kwargs)
#                     )
#                     report_file.file.save(f'zero_usage_billed_data_report_{month}_{year}.xlsx', ContentFile(output_with_formatting.getvalue()))
#                     saveuserlog(request.user, f"Billed data report excel file of {month}-{year} downloaded.")
#                     return Response({"data": report_file.file.url, "message" : "Excel file generated sucessfully!"}, status=status.HTTP_200_OK)

#                     # response = HttpResponse(output_with_formatting, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#                     # response['Content-Disposition'] = f'attachment; filename=zero_usage_billed_data_report_{month}_{year}.xlsx'

#                     # return response

#                 except Exception as e:
#                     print(e)
#                     return Response({"message": "Internal Server Error."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
#             elif sub_type == "getTop10BilledUsers":
#                 try:
#                     filtered_data = Report_Billed_Data.objects.filter(**filter_kwargs).annotate(
#                         usage_float=Cast('Data_Usage_GB', FloatField())
#                     ).order_by('-usage_float')[:10]

#                     # Create a list of dictionaries to store the data in the required format
#                     report_data = []
#                     for record in filtered_data:
#                         report_data.append({
#                             "accountNumber": record.Account_Number,
#                             'wirelessNumber': record.Wireless_Number,
#                             'userName': record.User_Name,
#                             'dataUsageGB': float(record.Data_Usage_GB),
#                         })

#                     # Convert the data to a pandas DataFrame
#                     df = pd.DataFrame(report_data, columns=['accountNumber', 'wirelessNumber', 'userName', 'dataUsageGB'])

#                     # Create a bar graph for the top 10 data usage
#                     plt.figure(figsize=(10, 6))
#                     df.plot(kind='bar', x='userName', y='dataUsageGB', legend=False)
#                     plt.title(f'Top 10 Data Usage for Vendor: {vendor}')
#                     plt.xlabel('User Name')
#                     plt.ylabel('Data Usage (GB)')
#                     plt.xticks(rotation=45, ha='right')
#                     plt.tight_layout()

#                     # Save the plot to a BytesIO object
#                     plot_output = BytesIO()
#                     plt.savefig(plot_output, format='png')
#                     plot_output.seek(0)

#                     # Create a new BytesIO output for the Excel file
#                     output = BytesIO()

#                     # Write the DataFrame to an Excel file using pandas
#                     with pd.ExcelWriter(output, engine='openpyxl') as writer:
#                         df.to_excel(writer, index=False, sheet_name='Top 10 Usage Billed Data Report')

#                     # Reopen the Excel file for additional formatting with openpyxl
#                     output.seek(0)
#                     workbook = load_workbook(output)
#                     worksheet = workbook['Top 10 Usage Billed Data Report']

#                     # Insert the plot image into the worksheet
#                     img = Image(plot_output)
#                     worksheet.add_image(img, 'F2')

#                     # Insert a dynamic heading for the top of the report
#                     main_heading = f'{month} Top 10 Data Usage ({vendor})'
#                     worksheet.insert_rows(1)
#                     worksheet.merge_cells('A1:D1')
#                     heading_cell = worksheet['A1']
#                     heading_cell.value = main_heading
#                     heading_cell.font = Font(bold=True, color="FFFFFF", size=14)
#                     heading_cell.alignment = Alignment(horizontal="center", vertical="center")
#                     heading_cell.fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')

#                     # Apply formatting to the 'dataUsageGB' column and adjust column widths
#                     dark_navy_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
#                     white_font = Font(color="FFFFFF")
#                     for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=4, max_col=4):
#                         for cell in row:
#                             cell.fill = dark_navy_fill
#                             cell.font = white_font

#                     for col_num, column_cells in enumerate(worksheet.columns, 1):
#                         max_length = 0
#                         column = get_column_letter(col_num)
#                         for cell in column_cells:
#                             if cell.value:
#                                 max_length = max(max_length, len(str(cell.value)))
#                         adjusted_width = (max_length + 2) if max_length < 30 else 30
#                         worksheet.column_dimensions[column].width = adjusted_width

#                     # Save the workbook with formatting
#                     output_with_formatting = BytesIO()
#                     workbook.save(output_with_formatting)
#                     output_with_formatting.seek(0)

#                     report_file = Downloaded_reports(
#                         report_type = report_type,
#                         kwargs = self.change_kwargs(filter_kwargs)
#                     )
#                     report_file.file.save(f'top_10_usage_billed_data_report_{month}_{year}.xlsx', ContentFile(output_with_formatting.getvalue()))
#                     saveuserlog(request.user, f"Billed data report excel file of {month}-{year} downloaded.")
#                     return Response({"data": report_file.file.url, "message" : "Excel file generated successfully!"}, status=status.HTTP_200_OK)

#                 except Exception as e:
#                     return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



# views.py

from datetime import datetime
from io import BytesIO

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")  # headless backend for servers
import matplotlib.pyplot as plt

from django.core.files.base import ContentFile
from django.db.models import Q, FloatField
from django.db.models.functions import Cast

from rest_framework.views import APIView
from rest_framework import status, permissions
from rest_framework.response import Response

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from authenticate.views import saveuserlog
from OnBoard.Organization.models import Organizations
from OnBoard.Company.models import Company
from Dashboard.ModelsByPage.DashAdmin import Vendors
from OnBoard.Ban.models import BaseDataTable

from ..models import Report_Billed_Data, Downloaded_reports
from ..ser import (
    showBanSerializer,
    showBilledReport,
    OrganizationShowSerializer,
    VendorShowSerializer,
)


# helper: consistent month ordering (handles numeric "1"/"01" or names "Jan"/"January")
_MONTH_ORDER = [
    "1","01","January","Jan",
    "2","02","February","Feb",
    "3","03","March","Mar",
    "4","04","April","Apr",
    "5","05","May",
    "6","06","June","Jun",
    "7","07","July","Jul",
    "8","08","August","Aug",
    "9","09","September","Sep","Sept",
    "10","October","Oct",
    "11","November","Nov",
    "12","December","Dec",
]
def _month_sort_key(m):
    # try numeric
    try:
        return int(m)
    except Exception:
        pass
    # map by first occurrence in known list; fallback lexicographic
    if m in _MONTH_ORDER:
        # position group index by month “slot” 1..12
        # find the first index where that slot starts
        slots = [0,4,8,12,16,20,23,26,29,33,36,39]  # starts of each month group in _MONTH_ORDER
        for idx, start in enumerate(slots, start=1):
            if m in _MONTH_ORDER[start:(slots[idx] if idx < 12 else None)]:
                return idx
    return 99


class ViewBilledReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def __init__(self, **kwargs):
        self.report_type = "Billed_Data_Usage"
        self.current_year = datetime.now().year
        self.data = None
        self.unique_dates = None
        super().__init__(**kwargs)

    # -----------------------------
    # GET (unchanged list behavior)
    # -----------------------------
    def get(self, request, sub_type=None, *args, **kwargs):

        company = request.GET.get('company')
        sub_company = request.GET.get('sub_company')
        ban = request.GET.get('ban')
        report_type = request.GET.get('report_type')
        month = request.GET.get('month')
        vendor = request.GET.get('vendor')
        year = request.GET.get('year')
        print(company, sub_company, ban, report_type, month, vendor, year, sub_type)
        if not (company and sub_company and ban and report_type and month and vendor and year):
            return Response({"message":"Missing required fields."},status=status.HTTP_400_BAD_REQUEST)
        
        filter_kwargs = {}
        if company:
            filter_kwargs['company'] = Company.objects.filter(Company_name=company)[0]
        if sub_company:
            filter_kwargs['organization'] = Organizations.objects.filter(Organization_name=sub_company)[0]
        if ban:
            filter_kwargs['Account_Number'] = ban
        if report_type:
            filter_kwargs['Report_Type'] = report_type
        if month:
            filter_kwargs['Month'] = month
        if year:
            filter_kwargs['Year'] = year
        if vendor:
            filter_kwargs['vendor'] = Vendors.objects.filter(name=vendor)[0]

        filtered_data = Report_Billed_Data.objects.filter(company=filter_kwargs['company'])
        if not filtered_data.exists():
            return Response({"message":f"no unbilled report found for {filter_kwargs['organization'].Organization_name}"},status=status.HTTP_400_BAD_REQUEST)
        filtered_data = filtered_data.filter(organization=filter_kwargs['organization'])
        if not filtered_data.exists():
            return Response({"message":f"no unbilled report found for {filter_kwargs['organization'].Organization_name}"},status=status.HTTP_400_BAD_REQUEST)
        filtered_data = filtered_data.filter(vendor=filter_kwargs['vendor'])
        if not filtered_data.exists():
            return Response({"message":f"no unbilled report found for {filter_kwargs['vendor'].name}"},status=status.HTTP_400_BAD_REQUEST)
        filtered_data = filtered_data.filter(Account_Number=filter_kwargs['Account_Number'])
        if not filtered_data.exists():
            return Response({"message":f"no unbilled report found for {filter_kwargs['Account_Number']}"},status=status.HTTP_400_BAD_REQUEST)
        filtered_data = filtered_data.filter(Report_Type=filter_kwargs['Report_Type'])
        if not filtered_data.exists():
            return Response({"message":f"no unbilled report found for {filter_kwargs['Report_Type']}"},status=status.HTTP_400_BAD_REQUEST)
        filtered_data = filtered_data.filter(Month=filter_kwargs['Month'])
        if not filtered_data.exists():
            return Response({"message":f"no unbilled report found for {filter_kwargs['Month']}"},status=status.HTTP_400_BAD_REQUEST)
        filtered_data = filtered_data.filter(Year=filter_kwargs['Year'])
        if not filtered_data.exists():
            return Response({"message":f"no unbilled report found for {filter_kwargs['Year']}"},status=status.HTTP_400_BAD_REQUEST)

        if not filtered_data.exists():
            return Response({"message": "No data found for the given filters."}, status=status.HTTP_200_OK)
        if not sub_type:
            self.data = showBilledReport(filtered_data, many=True)
        else:
            

            if sub_type == "Top10BilledUsers":
                
                try:
                    top10 = (
                        filtered_data
                        .annotate(usage_float=Cast("Data_Usage_GB", FloatField()))
                        .filter(usage_float__gt=0)  # ✅ ignore 0 or null values
                        .order_by("-usage_float")[:10]
                    )
                    self.data = showBilledReport(top10, many=True)
                except Exception:
                    return Response(
                        {"message": "Unable to view report."},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR
                    )

            elif sub_type == "getZeroUsageReportbilledData":
                try:
                    zero = filtered_data.filter(Q(Data_Usage_GB="0") | Q(Data_Usage_GB="0.0"))
                    if not zero.exists():
                        return Response({"message": "No data found for the given filters."}, status=status.HTTP_200_OK)
                    self.data = showBilledReport(zero, many=True)
                except Exception:
                    return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            else:
                return Response({"message": "Invalid sub_type provided."}, status=status.HTTP_400_BAD_REQUEST)

        return Response(
            {
                "data": self.data.data if self.data else None,
                "unique_dates": self.unique_dates if self.report_type == "Unbilled_Data_Usage" else None,
            },
            status=status.HTTP_200_OK,
        )

    # -----------------------------
    # helpers for Excel
    # -----------------------------
    def change_kwargs(self, kwargs):
        if "organization" in kwargs and str(kwargs["organization"]) != str:
            kwargs["organization"] = kwargs["organization"].Organization_name
        if "company" in kwargs and str(kwargs["company"]) != str:
            kwargs["company"] = kwargs["company"].Company_name
        if "vendor" in kwargs and str(kwargs["vendor"]) != str:
            kwargs["vendor"] = kwargs["vendor"].name
        return kwargs

    def _auto_fit_columns(self, ws, max_width=30):
        for col_idx, _ in enumerate(ws.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for c in ws[col_letter]:
                if c.value is not None:
                    max_len = max(max_len, len(str(c.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

    def _month_sort_key(m):
        # Try numeric first
        try:
            return int(str(m).lstrip("0") or "0")
        except Exception:
            pass
        # Map common names/abbrevs → month number
        m_str = str(m).strip().lower()
        name_map = {
            "jan":1,"january":1,"feb":2,"february":2,"mar":3,"march":3,"apr":4,"april":4,"may":5,
            "jun":6,"june":6,"jul":7,"july":7,"aug":8,"august":8,"sep":9,"sept":9,"september":9,
            "oct":10,"october":10,"nov":11,"november":11,"dec":12,"december":12
        }
        return name_map.get(m_str, 99)

    def _generate_combined_workbook(self, base_qs, filter_kwargs, year, vendor_name, report_type):
        """
        ONE workbook with 3 sheets (monthwise):
        - Billed Data Report: months as columns; colored cells; Average (GB) column
        - Zero Usage Billed Data Report: ONLY zero rows; months as columns with '0 GB'
        - Top 10 Usage Billed Data Report: global top 10 (non-zero), deduped by wirelessNumber;
            header shows actual month span used
        """
        import pandas as pd
        import matplotlib.pyplot as plt
        from io import BytesIO
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from django.core.files.base import ContentFile

        # ---- helpers ----
        def _month_sort_key(m):
            try:
                return int(str(m).lstrip("0") or "0")
            except Exception:
                pass
            m_str = str(m).strip().lower()
            name_map = {
                "jan":1,"january":1,"feb":2,"february":2,"mar":3,"march":3,"apr":4,"april":4,"may":5,
                "jun":6,"june":6,"jul":7,"july":7,"aug":8,"august":8,"sep":9,"sept":9,"september":9,
                "oct":10,"october":10,"nov":11,"november":11,"dec":12,"december":12
            }
            return name_map.get(m_str, 99)

        def _nanlike(val):
            if val is None:
                return True
            s = str(val).strip().lower()
            return s in {"", "nan", "none", "null"}

        def _drop_nanlike(df_in, cols):
            if df_in.empty:
                return df_in
            mask_bad = False
            for c in cols:
                if c in df_in.columns:
                    _bad = df_in[c].astype(str).str.strip().str.lower().isin(["", "nan", "none", "null"])
                    mask_bad = _bad if isinstance(mask_bad, bool) else (mask_bad | _bad)
            return df_in[~mask_bad].copy()

        def _col_offset(col_index, offset=3):
            return get_column_letter(col_index + offset)

        def _format_header_row(ws, header_row=2):
            light_gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            for j in range(1, ws.max_column + 1):
                cell = ws.cell(row=header_row, column=j)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = light_gray

        # ---------- Build base DataFrame (drop junk wireless numbers upfront) ----------
        rows = []
        for r in base_qs:
            if _nanlike(r.Wireless_Number):
                continue
            rows.append({
                "Month": str(r.Month),
                "accountNumber": r.Account_Number,
                "wirelessNumber": r.Wireless_Number,
                "userName": r.User_Name,
                "vendor": r.vendor.name,
                "usage": float(r.Data_Usage_GB),
            })

        df = pd.DataFrame(rows)
        if not df.empty:
            df = _drop_nanlike(df, ["wirelessNumber", "userName"])

        if df.empty:
            months_sorted = []
            df_billed_pivot = pd.DataFrame(columns=["Account Number","Wireless Number","User Name","Vendor"])
            df_zero_pivot   = pd.DataFrame(columns=["Account Number","Wireless Number","User Name"])
            df_top10        = pd.DataFrame(columns=["Month","Account Number","Wireless Number","User Name","Data Usage (GB)"])
            top10_month_span_str = "No Non-zero Months"
        else:
            months_sorted = sorted(df["Month"].unique().tolist(), key=_month_sort_key)

            # ----- Billed: pivot usage by month -----
            idx_cols = ["accountNumber","wirelessNumber","userName","vendor"]
            billed = (
                df.pivot_table(
                    index=idx_cols,
                    columns="Month",
                    values="usage",
                    aggfunc="sum",
                    fill_value=0.0,
                )
                .reindex(columns=months_sorted)
                .reset_index()
            )
            # Average across available months (replace "Total")
            if months_sorted:
                billed["Average (GB)"] = billed[months_sorted].mean(axis=1)

            rename_map = {
                "accountNumber": "Account Number",
                "wirelessNumber": "Wireless Number",
                "userName": "User Name",
                "vendor": "Vendor",
            }
            df_billed_pivot = billed.rename(columns=rename_map)
            df_billed_pivot = _drop_nanlike(df_billed_pivot, ["Wireless Number", "User Name"])

            # ----- Zero Usage: ONLY zero rows, show "0 GB" monthwise -----
            df_zero_only = df[df["usage"] == 0].copy()
            if df_zero_only.empty:
                df_zero_pivot = pd.DataFrame(columns=["Account Number","Wireless Number","User Name"] + months_sorted)
            else:
                df_zero_only["zero_display"] = "0 GB"
                zero = (
                    df_zero_only.pivot_table(
                        index=["accountNumber","wirelessNumber","userName"],
                        columns="Month",
                        values="zero_display",
                        aggfunc="first",
                        fill_value="",
                    )
                    .reindex(columns=months_sorted)
                    .reset_index()
                )
                df_zero_pivot = zero.rename(columns={
                    "accountNumber":"Account Number",
                    "wirelessNumber":"Wireless Number",
                    "userName":"User Name",
                })
                df_zero_pivot = _drop_nanlike(df_zero_pivot, ["Wireless Number", "User Name"])

            # ----- Global Top 10 (non-zero, unique by wirelessNumber) -----
            df_work = df[df["usage"] > 0].copy()
            if df_work.empty:
                df_top10 = pd.DataFrame(columns=["Month","Account Number","Wireless Number","User Name","Data Usage (GB)"])
                top10_month_span_str = "No Non-zero Months"
            else:
                months_used_sorted = sorted(df_work["Month"].unique().tolist(), key=_month_sort_key)
                top10_month_span_str = months_used_sorted[0] if len(months_used_sorted) == 1 else f"{months_used_sorted[0]}–{months_used_sorted[-1]}"
                df_work["_month_order"] = df_work["Month"].apply(_month_sort_key).astype(int)
                df_work = df_work.sort_values(
                    ["wirelessNumber", "usage", "_month_order"],
                    ascending=[True, False, False]
                )
                df_unique = df_work.drop_duplicates(subset=["wirelessNumber"], keep="first")
                df_top10 = (
                    df_unique.nlargest(10, "usage")
                    [["Month","accountNumber","wirelessNumber","userName","usage"]]
                    .rename(columns={
                        "accountNumber":"Account Number",
                        "wirelessNumber":"Wireless Number",
                        "userName":"User Name",
                        "usage":"Data Usage (GB)",
                    })
                )
                df_top10 = _drop_nanlike(df_top10, ["Wireless Number", "User Name"])

        # ---------- Charts ----------
        billed_plot = None
        if not df.empty and months_sorted:
            monthly_total = df.groupby("Month")["usage"].sum().reindex(months_sorted)
            plt.figure(figsize=(12, 6))
            monthly_total.plot(kind="bar")
            plt.title(f"Total Usage by Month ({year}) - Vendor: {vendor_name}")
            plt.xlabel("Month"); plt.ylabel("Total GB")
            plt.xticks(rotation=45, ha="right"); plt.tight_layout()
            billed_plot = BytesIO(); plt.savefig(billed_plot, format="png"); plt.close(); billed_plot.seek(0)

        zero_plot = None
        if not df.empty and months_sorted:
            monthly_zero = (df["usage"] == 0).groupby(df["Month"]).sum().reindex(months_sorted)
            plt.figure(figsize=(10, 6))
            monthly_zero.plot(kind="bar")
            plt.title(f"Zero-Usage Counts by Month ({year}) - Vendor: {vendor_name}")
            plt.xlabel("Month"); plt.ylabel("Zero Lines")
            plt.xticks(rotation=45, ha="right"); plt.tight_layout()
            zero_plot = BytesIO(); plt.savefig(zero_plot, format="png"); plt.close(); zero_plot.seek(0)

        top_plot = None
        if not df_top10.empty:
            plt.figure(figsize=(10, 6))
            labels = df_top10.apply(lambda r: f"{r['User Name']} ({r['Month']})", axis=1)
            plt.bar(labels, df_top10["Data Usage (GB)"])
            title_span = top10_month_span_str if 'top10_month_span_str' in locals() else "All Months"
            plt.title(f"Global Top 10 Usage ({title_span} {year}) - {vendor_name}")
            plt.xlabel("User (Month)"); plt.ylabel("Data Usage (GB)")
            plt.xticks(rotation=45, ha="right"); plt.tight_layout()
            top_plot = BytesIO(); plt.savefig(top_plot, format="png"); plt.close(); top_plot.seek(0)

        # ---------- Write sheets ----------
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            billed_cols = ["Account Number","Wireless Number","User Name","Vendor"] + months_sorted + (["Average (GB)"] if months_sorted else [])
            (df_billed_pivot if not df_billed_pivot.empty else pd.DataFrame(columns=billed_cols)).to_excel(
                writer, index=False, sheet_name="Billed Data Report"
            )
            zero_cols = ["Account Number","Wireless Number","User Name"] + months_sorted
            (df_zero_pivot if not df_zero_pivot.empty else pd.DataFrame(columns=zero_cols)).to_excel(
                writer, index=False, sheet_name="Zero Usage Billed Data Report"
            )
            top_cols = ["Month","Account Number","Wireless Number","User Name","Data Usage (GB)"]
            (df_top10 if not df_top10.empty else pd.DataFrame(columns=top_cols)).to_excel(
                writer, index=False, sheet_name="Top 10 Usage Billed Data Report"
            )
        output.seek(0)

        # ---------- Format + place charts ----------
        wb = load_workbook(output)
        dark_navy  = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
        red_fill   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        yellow_fill= PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        blue_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        light_yellow = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
        white_font = Font(color="FFFFFF")

        # A) Billed sheet
        ws_billed = wb["Billed Data Report"]
        ws_billed.insert_rows(1)
        last_col_letter = get_column_letter(ws_billed.max_column or 1)
        ws_billed.merge_cells(f"A1:{last_col_letter}1")
        h = ws_billed["A1"]
        h.value = f"Usage Report (All Months {year}) - Monthwise"
        h.font = Font(bold=True, color="FFFFFF", size=14)
        h.alignment = Alignment(horizontal="center", vertical="center")
        h.fill = dark_navy
        _format_header_row(ws_billed, header_row=2)

        # numeric format for month & average columns
        if ws_billed.max_row >= 3 and ws_billed.max_column >= 5:
            # month cols start at 5 (after 4 ID cols)
            month_col_start = 5
            # average is the last column if months exist
            avg_col = ws_billed.max_column
            for r in range(3, ws_billed.max_row + 1):
                # month columns
                for c in range(month_col_start, avg_col):
                    ws_billed.cell(row=r, column=c).number_format = "0.00"
                # average column
                ws_billed.cell(row=r, column=avg_col).number_format = "0.00"

            # color rules for month columns ONLY
            for r in range(3, ws_billed.max_row + 1):
                for c in range(month_col_start, avg_col):  # exclude avg column
                    cell = ws_billed.cell(row=r, column=c)
                    try:
                        v = float(cell.value)
                    except Exception:
                        continue
                    if v == 0:
                        cell.fill = red_fill
                        cell.font = white_font
                    elif 0 < v < 5:
                        cell.fill = yellow_fill
                    elif v > 15:
                        cell.fill = blue_fill
                        cell.font = white_font

        if billed_plot:
            ws_billed.add_image(XLImage(billed_plot), f"{_col_offset(ws_billed.max_column, 3)}2")
        self._auto_fit_columns(ws_billed)

        # B) Zero usage sheet
        ws_zero = wb["Zero Usage Billed Data Report"]
        ws_zero.insert_rows(1)
        last_col_letter = get_column_letter(ws_zero.max_column or 1)
        ws_zero.merge_cells(f"A1:{last_col_letter}1")
        z = ws_zero["A1"]
        z.value = f"Zero Usage (All Months {year}) - Monthwise"
        z.font = Font(bold=True, color="FFFFFF", size=14)
        z.alignment = Alignment(horizontal="center", vertical="center")
        z.fill = dark_navy
        _format_header_row(ws_zero, header_row=2)
        if ws_zero.max_row >= 3 and ws_zero.max_column >= 4:
            for r in range(3, ws_zero.max_row + 1):
                for c in range(4, ws_zero.max_column + 1):
                    if ws_zero.cell(row=r, column=c).value == "0 GB":
                        ws_zero.cell(row=r, column=c).fill = light_yellow
        if zero_plot:
            ws_zero.add_image(XLImage(zero_plot), f"{_col_offset(ws_zero.max_column, 3)}2")
        self._auto_fit_columns(ws_zero)

        # C) Top 10 sheet
        ws_top = wb["Top 10 Usage Billed Data Report"]
        ws_top.insert_rows(1)
        last_col_letter = get_column_letter(ws_top.max_column or 1)
        ws_top.merge_cells(f"A1:{last_col_letter}1")
        t = ws_top["A1"]
        title_span = top10_month_span_str if 'top10_month_span_str' in locals() else "All Months"
        t.value = f"Global Top 10 Usage ({title_span} {year})"
        t.font = Font(bold=True, color="FFFFFF", size=14)
        t.alignment = Alignment(horizontal="center", vertical="center")
        t.fill = dark_navy
        _format_header_row(ws_top, header_row=2)
        # numeric format for GB column (last col)
        if ws_top.max_row >= 3 and ws_top.max_column >= 5:
            gb_col = ws_top.max_column
            for r in range(3, ws_top.max_row + 1):
                ws_top.cell(row=r, column=gb_col).number_format = "0.00"
        if top_plot:
            ws_top.add_image(XLImage(top_plot), f"{_col_offset(ws_top.max_column, 3)}2")
        self._auto_fit_columns(ws_top)

        # ---------- Save ----------
        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        report_file = Downloaded_reports(
            report_type=report_type,
            kwargs=self.change_kwargs(filter_kwargs),
        )
        report_file.file.save(f"usage_bundle_all_months_{year}.xlsx", ContentFile(final_output.getvalue()))
        return report_file





    # -----------------------------
    # POST: return combined Excel (3 tabs, ALL months, pivoted)
    # -----------------------------
    def post(self, request, sub_type=None, *args, **kwargs):
        # keep params for FE compatibility (month is ignored for filtering now)
        company = request.data.get("company")
        sub_company = request.data.get("sub_company")
        ban = request.data.get("ban")
        month = request.data.get("month")          # ignored in filter
        report_type = request.data.get("report_type")
        year = request.data.get("year")
        vendor = request.data.get("vendor")

        if not (company and sub_company and ban and report_type and month and vendor and year):
            return Response({"message": "Missing required fields."}, status=status.HTTP_400_BAD_REQUEST)

        company_obj = Company.objects.filter(Company_name=company).first()
        org_obj = Organizations.objects.filter(Organization_name=sub_company).first()
        vendor_obj = Vendors.objects.filter(name=vendor).first()
        if not (company_obj and org_obj and vendor_obj):
            return Response({"message": "Invalid company / organization / vendor."}, status=status.HTTP_400_BAD_REQUEST)

        # IMPORTANT: no Month filter → get ALL months for the year
        filter_kwargs = {
            "company": company_obj,
            "organization": org_obj,
            "Account_Number": ban,
            "Report_Type": report_type,
            "Year": year,
            "vendor": vendor_obj,
        }

        base_qs = Report_Billed_Data.objects.filter(**filter_kwargs)
        if not base_qs.exists():
            return Response({"message": "No data found for the given filters."}, status=status.HTTP_200_OK)

        try:
            report_file = self._generate_combined_workbook(
                base_qs=base_qs,
                filter_kwargs=filter_kwargs,
                year=year,
                vendor_name=vendor_obj.name,
                report_type=report_type,
            )
            saveuserlog(request.user, f"Combined billed data report (ALL months {year}, pivoted) generated.")
            return Response(
                {"data": report_file.file.url, "message": "Excel file (all months, pivoted) generated successfully!"},
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            print("Combined (all months, pivot) report error:", e)
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)