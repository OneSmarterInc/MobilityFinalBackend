# from rest_framework.views import APIView
# from rest_framework import status
# from rest_framework.response import Response
# from rest_framework import permissions
# from authenticate.views import saveuserlog
# from OnBoard.Organization.models import Organizations
# from OnBoard.Company.models import Company
# from Dashboard.ModelsByPage.DashAdmin import Vendors
# from OnBoard.Ban.models import BaseDataTable


# from django.db.models.functions import Cast
# from django.db.models import FloatField

# import matplotlib.pyplot as plt
# from openpyxl import load_workbook
# from openpyxl.drawing.image import Image
# from openpyxl.styles import PatternFill
# from io import BytesIO
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import Font, Alignment, PatternFill
# from openpyxl.styles import Font, Alignment, PatternFill
# from openpyxl.styles import Font, Alignment, PatternFill
# from openpyxl.styles import Font, Alignment, PatternFill
# from openpyxl.styles import Font, Alignment, PatternFill
# from django.db.models import Q
# from django.core.files.base import ContentFile
# from ..models import Report_Unbilled_Data
# from ..ser import showBanSerializer,showUnbilledReport, OrganizationShowSerializer, VendorShowSerializer

# from openpyxl import Workbook
# from datetime import datetime
# from ..models import Downloaded_reports
# from collections import defaultdict
# import pandas as pd

# class ViewUnbilledReportView(APIView):
#     permission_classes = [permissions.IsAuthenticated]

#     def __init__(self, **kwargs):
#         self.report_type = "Unbilled_Data_Usage"
#         self.current_year = datetime.now().year
#         self.data = None
#         self.unique_dates = None
#         super().__init__(**kwargs)
#     def get(self, request, sub_type=None, *args, **kwargs):
#         if request.user.designation.name == 'Admin':
#             orgs = OrganizationShowSerializer(Organizations.objects.filter(company=request.user.company), many=True)
#             vendors = VendorShowSerializer(Vendors.objects.all(), many=True)
#             bans = showBanSerializer(BaseDataTable.objects.filter(company=request.user.company, viewuploaded=None, viewpapered=None), many=True)
#         else:
#             orgs = OrganizationShowSerializer(Organizations.objects.all(), many=True)
#             vendors = VendorShowSerializer(Vendors.objects.all(), many=True)
#             bans = showBanSerializer(BaseDataTable.objects.filter(viewuploaded=None, viewpapered=None), many=True)

#         if not sub_type:
#             filter_kwargs = {}
#             company = Company.objects.get(Company_name=request.user.company.Company_name) if request.user.company else None
#             if company:
#                 filter_kwargs['company'] = company
#             if self.report_type:
#                 filter_kwargs['Report_Type'] = self.report_type
#             filtered_data = Report_Unbilled_Data.objects.filter(**filter_kwargs)

#             report_data = []
#             for record in filtered_data:
#                 existing_report = next((r for r in report_data if r['Wireless_Number'] == record.Wireless_Number), None)
#                 if not existing_report:
#                     report_data.append({
#                         "Account_Number": record.Account_Number,
#                         'Wireless_Number': record.Wireless_Number,
#                         'user_Name': record.User_Name,
#                         'vendor': record.vendor,
#                         record.Date: record.Usage,
#                     })
#                 else:
#                     existing_report[record.Date] = record.Usage

#             self.unique_dates = sorted(
#                 list(set(
#                     datetime.strptime(record.Date, '%Y-%m-%d %H:%M:%S').date()
#                     for record in filtered_data
#                 )),
#                 key=lambda date: date
#             )
#             self.data = showUnbilledReport(Report_Unbilled_Data.objects.all(), many=True)
#         else:
#             company = request.GET.get('company')
#             sub_company = request.GET.get('sub_company')
#             ban = request.GET.get('ban')
#             report_type = request.GET.get('report_type')
#             month = request.GET.get('month')
#             vendor = request.GET.get('vendor')
#             year = request.GET.get('year')
#             week = request.GET.get('week')
#             print("company==", company,sub_company, ban, report_type,month,vendor,year)
#             if not (company and sub_company and ban and report_type and month and vendor and year):
#                 return Response({"message":"Missing required fields."},status=status.HTTP_400_BAD_REQUEST)
#             filter_kwargs = {}
#             if company:
#                 filter_kwargs['company'] = Company.objects.filter(Company_name=company)[0]
#             if sub_company:
#                 filter_kwargs['organization'] = Organizations.objects.filter(Organization_name=sub_company)[0]

#             if ban:
#                 filter_kwargs['Account_Number'] = ban
#             if report_type:
#                 filter_kwargs['Report_Type'] = report_type
#             if month:
#                 filter_kwargs['Month'] = month
#             if year:
#                 filter_kwargs['Year'] = year
#             if vendor:
#                 filter_kwargs['vendor'] = Vendors.objects.filter(name=vendor)[0]
#             filtered_data = Report_Unbilled_Data.objects.filter(company=filter_kwargs['company'])
#             if not filtered_data.exists():
#                 return Response({"message":f"no unbilled report found for {filter_kwargs['company'].Company_name}"},status=status.HTTP_400_BAD_REQUEST)
            
#             filtered_data = filtered_data.filter(organization=filter_kwargs['organization'])
#             if not filtered_data.exists():
#                 return Response({"message":f"no unbilled report found for {filter_kwargs['organization'].Organization_name}"},status=status.HTTP_400_BAD_REQUEST)
            
#             filtered_data = filtered_data.filter(vendor=filter_kwargs['vendor'])
#             if not filtered_data.exists():
#                 return Response({"message":f"no unbilled report found for {filter_kwargs['vendor'].name}"},status=status.HTTP_400_BAD_REQUEST)
            
#             filtered_data = filtered_data.filter(Account_Number=filter_kwargs['Account_Number'])
#             if not filtered_data.exists():
#                 return Response({"message":f"no unbilled report found for {filter_kwargs['Account_Number']}"},status=status.HTTP_400_BAD_REQUEST)
            
#             filtered_data = filtered_data.filter(Report_Type=filter_kwargs['Report_Type'])
#             if not filtered_data.exists():
#                 return Response({"message":f"no unbilled report found for {filter_kwargs['Report_Type']}"},status=status.HTTP_400_BAD_REQUEST)
            
#             filtered_data = filtered_data.filter(Month=filter_kwargs['Month'])
#             if not filtered_data.exists():
#                 return Response({"message":f"no unbilled report found for {filter_kwargs['Month']}"},status=status.HTTP_400_BAD_REQUEST)
            
#             filtered_data = filtered_data.filter(Year=filter_kwargs['Year'])
#             if not filtered_data.exists():
#                 return Response({"message":f"no unbilled report found for {filter_kwargs['Year']}"},status=status.HTTP_400_BAD_REQUEST)
            
            
            

#             if sub_type == "getTop10byWeek":
#                 try:
#                     if week:
#                         filter_kwargs['Week'] = week
#                     friday_date = None
#                     thursday_date = None
#                     monday_date = None
#                     for record in filtered_data:
#                         date_obj = datetime.strptime(record.Date, '%Y-%m-%d %H:%M:%S').date()
#                         if date_obj.weekday() == 4:  # Friday
#                             friday_date = date_obj
#                         elif date_obj.weekday() == 3 and not thursday_date:  # Thursday
#                             thursday_date = date_obj
#                         elif date_obj.weekday() == 0 and not monday_date:  # Monday
#                             monday_date = date_obj

#                     selected_date = friday_date or thursday_date or monday_date
#                     if not selected_date:
#                         return Response({"message": "No valid dates found for the specified week"}, status=status.HTTP_404_NOT_FOUND)
#                     top_data = filtered_data.filter(Date=selected_date.strftime('%Y-%m-%d %H:%M:%S')).annotate(
#                         usage_float=Cast('Usage', FloatField())
#                     ).order_by('-usage_float')[:10]

#                     if not top_data:
#                         return Response({"message": "No data found for unbilled top 10 users!"}, status=status.HTTP_404_NOT_FOUND)
#                     self.data = showUnbilledReport(top_data, many=True)
#                 except Exception as e:
#                     print(e)
#                     return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)
#             elif sub_type == "getZeroUsageReportUnbilledData":
#                 try:
#                     filtered_data = filtered_data.filter(Month=filter_kwargs['Month'])
#                     if not filtered_data.exists():
#                         return Response({"message":f"no unbilled report found for {filter_kwargs['Month']}"},status=status.HTTP_400_BAD_REQUEST)

#                     weeks = defaultdict(set)  
#                     for record in filtered_data:
#                         week = record.Week
#                         date = record.Date
#                         weeks[week].add(date)

            
#                     self.data = showUnbilledReport(filtered_data, many=True)
#                 except Exception as e:
#                     print(e)
#                     return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)
#             saveuserlog(request.user, f"Unilled data report excel file of {month}-{year} downloaded.")
#         return Response(
#             {"orgs": orgs.data, "vendors": vendors.data, "bans":bans.data, "data":self.data.data if self.data else None, "unique_dates":self.unique_dates if self.report_type == "Unbilled_Data_Usage" else None},
#             status=status.HTTP_200_OK,
#         )
#     def change_kwargs(self,kwargs):
#         if 'organization' in kwargs and str(kwargs['organization']) != str:
#             kwargs['organization'] = kwargs['organization'].Organization_name
#         if 'company' in kwargs and str(kwargs['company']) != str:
#             kwargs['company'] = kwargs['company'].Company_name
#         if 'vendor' in kwargs and str(kwargs['vendor']) != str:
#             kwargs['vendor'] = kwargs['vendor'].name

#         return kwargs
#     def post(self, request,sub_type=None, *args, **kwargs):
#         filter_kwargs = {}
#         action = request.data.get('action')
#         type = request.data.get('type')
#         company = request.data.get('company')
#         sub_company = request.data.get('sub_company')
#         ban = request.data.get('ban')
#         month = request.data.get('month')
#         report_type = request.data.get('report_type')
#         year = request.data.get('year')
#         vendor = request.data.get('vendor')
#         print("company==", company,sub_company, ban, report_type,month,vendor,year)
#         if not (company and sub_company and ban and report_type and month and vendor and year):
#             return Response({"message":"Missing required fields."},status=status.HTTP_400_BAD_REQUEST)

#         if company:
#             filter_kwargs['company'] = Company.objects.filter(Company_name=company)[0]
#         if sub_company:
#             filter_kwargs['organization'] = Organizations.objects.filter(Organization_name=sub_company)[0]
#         if ban:
#             filter_kwargs['Account_Number'] = ban
#         if report_type:
#             filter_kwargs['Report_Type'] = report_type
#         if month:
#             filter_kwargs['Month'] = month
#         if year:
#             filter_kwargs['Year'] = year
#         if vendor:
#             filter_kwargs['vendor'] = Vendors.objects.filter(name=vendor)[0]
#         filtered_data = Report_Unbilled_Data.objects.filter(company=filter_kwargs['company'])
#         if not filtered_data.exists():
#             return Response({"message":f"no unbilled report found for {filter_kwargs['company'].Company_name}"},status=status.HTTP_400_BAD_REQUEST)
        
#         filtered_data = filtered_data.filter(organization=filter_kwargs['organization'])
#         if not filtered_data.exists():
#             return Response({"message":f"no unbilled report found for {filter_kwargs['organization'].Organization_name}"},status=status.HTTP_400_BAD_REQUEST)
        
#         filtered_data = filtered_data.filter(Account_Number=filter_kwargs['Account_Number'])
#         if not filtered_data.exists():
#             return Response({"message":f"no unbilled report found for {filter_kwargs['Account_Number']}"},status=status.HTTP_400_BAD_REQUEST)

#         filtered_data = filtered_data.filter(Report_Type=filter_kwargs['Report_Type'])
#         if not filtered_data.exists():
#             return Response({"message":f"no unbilled report found for {filter_kwargs['Report_Type']}"},status=status.HTTP_400_BAD_REQUEST)
        
#         filtered_data = filtered_data.filter(Month=filter_kwargs['Month'])
#         if not filtered_data.exists():
#             return Response({"message":f"no unbilled report found for {filter_kwargs['Month']}"},status=status.HTTP_400_BAD_REQUEST)
        
#         filtered_data = filtered_data.filter(Year=filter_kwargs['Year'])
#         if not filtered_data.exists():
#             return Response({"message":f"no unbilled report found for {filter_kwargs['Year']}"},status=status.HTTP_400_BAD_REQUEST)
        
#         filtered_data = filtered_data.filter(vendor=filter_kwargs['vendor'])
#         if not filtered_data.exists():
#             return Response({"message":f"no unbilled report found for {filter_kwargs['vendor'].name}"},status=status.HTTP_400_BAD_REQUEST)
#         if not filtered_data.exists():
#             return Response({"message": "No data found for the given filters."}, status=200)
#         if not sub_type:
#             try:
#                 filtered_data = Report_Unbilled_Data.objects.filter(**filter_kwargs)

#                 # Create a dictionary to store the data in the required format
#                 report_data = []
#                 week_dates = {}

#                 for record in filtered_data:
#                     week = int(record.Week)
#                     date = record.Date
#                     # Add dates for each week
#                     if week not in week_dates:
#                         week_dates[week] = set()
#                     week_dates[week].add(date)

#                     existing_report = next((r for r in report_data if r['wirelessNumber'] == record.Wireless_Number), None)
#                     if not existing_report:
#                         report_data.append({
#                             "accountNumber": record.Account_Number,
#                             'wirelessNumber': record.Wireless_Number,
#                             'userName': record.User_Name,
#                             'vendor': record.vendor.name,
#                             record.Date: record.Usage,
#                         })
#                     else:
#                         existing_report[record.Date] = record.Usage

#                 # Sort the weeks and dates within each week
#                 sorted_weeks = sorted(week_dates.keys())
#                 sorted_week_dates = {week: sorted(dates) for week, dates in week_dates.items()}

#                 # Flatten the sorted dates into a list of columns for the DataFrame
#                 all_weeks_flat = []
#                 for week in sorted_weeks:
#                     all_weeks_flat.extend(sorted_week_dates[week])

#                 # Convert the data to a pandas DataFrame
#                 df = pd.DataFrame(report_data)
#                 if not df.empty:
#                     df = df[['accountNumber', 'wirelessNumber', 'userName', 'vendor'] + all_weeks_flat]

#                 # Create an Excel writer object and write the DataFrame to it
#                 output = BytesIO()
#                 writer = pd.ExcelWriter(output, engine='openpyxl')
#                 df.to_excel(writer, index=False, startrow=1, sheet_name='Unbilled Data Report')  # Start writing from row 2

#                 # Get the workbook and worksheet objects
#                 workbook = writer.book
#                 worksheet = writer.sheets['Unbilled Data Report']

#                 # Add Week 1, Week 2 headers in the first row (for merged cells)
#                 current_col = 5  # Start after the first 4 columns (accountNumber, wirelessNumber, etc.)
#                 for week_num in sorted_weeks:
#                     num_dates_in_week = len(sorted_week_dates[week_num])
#                     worksheet.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + num_dates_in_week - 1)
#                     worksheet.cell(1, current_col).value = f"Week {week_num}"
#                     worksheet.cell(1, current_col).alignment = Alignment(horizontal="center", vertical="center")
                    
#                     current_col += num_dates_in_week  # Move to the next week's set of dates

#                 # Adjust column widths based on the max length of each column (content, not headers)
#                 for col_num, column in enumerate(df.columns, 1):
#                     max_length = max(df[column].astype(str).map(len).max(), len(column))  # Find max content length
#                     worksheet.column_dimensions[get_column_letter(col_num)].width = max_length + 2  # Add some padding

#                 # Define specific colors for each week
#                 week_colors = {
#                     1: "FFFF99",  # Light Yellow for Week 1
#                     2: "CCFFCC",  # Light Green for Week 2
#                     3: "FFCCCC",  # Light Pink for Week 3
#                     4: "CCCCFF",  # Light Blue for Week 4
#                     5: "FFCC99",  # Light Orange for Week 5
#                 }

#                 # Apply the background color for each week uniformly across all columns for that week
#                 current_col = 5
#                 for week_num in sorted_weeks:
#                     num_dates_in_week = len(sorted_week_dates[week_num])
#                     fill_color = PatternFill(start_color=week_colors[week_num], end_color=week_colors[week_num], fill_type="solid")
                    
#                     # Apply color to Week headers
#                     worksheet.cell(1, current_col).fill = fill_color

#                     for col_offset in range(num_dates_in_week):
#                         for row in range(2, len(df) + 2):  # Start from row 2 (Actual data headers)
#                             worksheet.cell(row=row, column=current_col + col_offset).fill = fill_color
                    
#                     current_col += num_dates_in_week  # Move to the next week's set of dates

#                 # Add blue background with white text to headers (second row)
#                 header_fill = PatternFill(start_color="FF000066", end_color="FF000066", fill_type="solid")  # Blue
#                 header_font = Font(color="FFFFFF")  # White text
#                 for col_num in range(1, len(df.columns) + 1):
#                     cell = worksheet.cell(2, col_num)
#                     cell.fill = header_fill
#                     cell.font = header_font
#                     cell.alignment = Alignment(horizontal="center", vertical="center")

#                 writer.close()
#                 output.seek(0)
#                 report_file = Downloaded_reports(
#                     report_type=self.report_type,
#                     kwargs=self.change_kwargs(filter_kwargs)
#                 )
#                 report_file.file.save(f'unbilled_data_report_{month}.xlsx', ContentFile(output.getvalue()))
#                 saveuserlog(request.user, f"Unilled data report excel file of {month} downloaded.")

#                 return Response({"data":report_file.file.url, "message" : "Excel file generated sucessfully!"}, status=status.HTTP_200_OK)

#             except Exception as e:
#                 print(e)
#                 return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)
#         else:
#             if sub_type == "getZeroUsageReportUnbilledExcel":
#                 try:
#                     filtered_data = Report_Unbilled_Data.objects.filter(**filter_kwargs)
#                     # Group data by week
#                     week_data = defaultdict(list)
#                     unique_vendors = set()
#                     for record in filtered_data:
#                         week_data[record.Week].append(record)
#                         unique_vendors.add(record.vendor.name)  # Collect unique vendors

#                     # Create an Excel workbook and worksheet
#                     output = BytesIO()
#                     workbook = Workbook()

#                     # Apply formatting for titles and headers
#                     title_font = Font(bold=True, size=12)
#                     header_font = Font(bold=True, color="FFFFFF", size=14)
#                     medium_orange_fill = PatternFill(start_color="FFFF8503", end_color="FFFF8503", fill_type="solid")  # Medium orange header
#                     light_yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Light yellow for odd rows
#                     light_grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light grey for even rows
#                     main_header_fill = PatternFill(start_color="FF000066", end_color="FF000066", fill_type="solid")  # Dark blue bg for the main header

#                     worksheet = workbook.active
#                     worksheet.title = 'Zero Usage Report (Weekly)'

#                     worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)  # Merge across several columns
#                     worksheet.cell(row=1, column=1).value = f"{month} Zero Usage Report (Weekly)"
#                     worksheet.cell(row=1, column=1).font = header_font
#                     worksheet.cell(row=1, column=1).fill = main_header_fill
#                     worksheet.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
#                     def get_valid_week_date(week_records):
#                         week_dates = {'Monday': None, 'Wednesday': None, 'Friday': None}
#                         for rec in week_records:
#                             date_obj = datetime.strptime(rec.Date, '%Y-%m-%d %H:%M:%S')
#                             weekday = date_obj.strftime('%A')
#                             if weekday in week_dates and not week_dates[weekday]:
#                                 week_dates[weekday] = rec.Date
#                         return week_dates['Monday'] or week_dates['Wednesday'] or week_dates['Friday']
    
#                     start_column = 1
#                     row_offset = 3  # Start from the 3rd row (after the main header)
#                     for week, records in week_data.items():
#                         selected_date = get_valid_week_date(records)
#                         report_data = [
#                             {'Wireless_Number': rec.Wireless_Number, 'User_Name': rec.User_Name, 'Device': rec.Device, 'Upgrade_Eligibility_Date': rec.Upgrade_Eligibilty_Date}
#                             for rec in records
#                         ]
#                         df = pd.DataFrame(report_data)

#                         # Write the title, week, and date
#                         worksheet.merge_cells(start_row=row_offset, start_column=start_column, end_row=row_offset, end_column=start_column + 2)
#                         worksheet.cell(row=row_offset, column=start_column).value = f"Week {week} - {selected_date}"
#                         worksheet.cell(row=row_offset, column=start_column).font = title_font

#                         # Show unique vendor in the next column to the date
#                         worksheet.cell(row=row_offset, column=start_column + 3).value = ', '.join(unique_vendors)
#                         worksheet.cell(row=row_offset, column=start_column + 3).font = title_font

#                         # Write the headers
#                         headers = ['Wireless_Number', 'User_Name', 'Device', 'Upgrade_Eligibility_Date']
#                         for col_num, header in enumerate(headers, start=start_column):
#                             cell = worksheet.cell(row=row_offset + 1, column=col_num)
#                             cell.value = header
#                             cell.fill = medium_orange_fill
#                             cell.font = Font(bold=True, color="FFFFFF")
#                             cell.alignment = Alignment(horizontal="center")

#                         # Fill the data into the table with alternating row colors
#                         for row_num, row_data in enumerate(df.itertuples(index=False), row_offset + 2):
#                             fill_color = light_yellow_fill if row_num % 2 == 0 else light_grey_fill  # Alternate row colors
#                             for col_num, value in enumerate(row_data, start=start_column):
#                                 cell = worksheet.cell(row=row_num, column=col_num)
#                                 cell.value = value
#                                 cell.fill = fill_color

#                         # Adjust column widths dynamically based on content
#                         for col_num in range(start_column, start_column + len(headers)):
#                             max_len = max(len(str(cell.value)) for cell in worksheet[get_column_letter(col_num)])
#                             worksheet.column_dimensions[get_column_letter(col_num)].width = max_len + 5

#                         # Move start_column to the next set of columns for the next week's table
#                         start_column += len(headers) + 3  # Shift the starting column by enough space to place tables side by side
#                     # Save the modified workbook to the output stream
#                     workbook.save(output)
#                     output.seek(0)

#                     report_file = Downloaded_reports(
#                         report_type=report_type,
#                         kwargs=self.change_kwargs(filter_kwargs)
#                     )
#                     report_file.file.save(f'zero_usage_report_{month}.xlsx', ContentFile(output.getvalue()))
#                     saveuserlog(request.user, f"Unilled data report excel file of {month} downloaded.")

#                     return Response({"data": report_file.file.url, "message" : "Excel file generated sucessfully!"}, status=status.HTTP_200_OK)

#                 except Exception as e:
#                     print(e)
#                     return Response({
#                         'message': str(e),
#                     }, status=status.HTTP_400_BAD_REQUEST)
#             elif sub_type == "getTop10byWeekExcel":
#                 try:
#                     filtered_data = Report_Unbilled_Data.objects.filter(**filter_kwargs)

#                     week_data = {}
#                     for record in filtered_data:
#                         week = record.Week
#                         if week not in week_data:
#                             week_data[week] = []
#                         week_data[week].append(record)

#                     # Create an Excel workbook and worksheet
#                     output = BytesIO()
#                     workbook = Workbook()

#                     # Apply formatting for titles and headers
#                     title_font = Font(bold=True, size=12)
#                     header_font = Font(bold=True, color="FFFFFF", size=14)  # White text for the main header
#                     light_yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
#                     light_blue_fill = PatternFill(start_color="FF000066", end_color="FF000066", fill_type="solid")  # Blue header
#                     main_header_fill = PatternFill(start_color="FF000066", end_color="FF000066", fill_type="solid")  # Dark blue bg for the main header

#                     # Create the worksheet and set column offset for side-by-side tables
#                     worksheet = workbook.active
#                     worksheet.title = 'Top 10 Usage Report'
#                     start_column = 1  # Initial starting column for the first table
#                     row_offset = 3  # Start from the 3rd row (after the main header)

#                     # Add Main Header for the report (e.g., September Usage Report)
#                     worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)  # Merge across several columns
#                     worksheet.cell(row=1, column=1).value = f"{month} Usage Report (Weekly)"
#                     worksheet.cell(row=1, column=1).font = header_font
#                     worksheet.cell(row=1, column=1).fill = main_header_fill
#                     worksheet.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

#                     # Function to get the date with data for Monday, Wednesday, or Friday of a specific week
#                     def get_valid_week_date(week_records):
#                         week_dates = {'Monday': None, 'Wednesday': None, 'Friday': None}

#                         for rec in week_records:
#                             date_obj = datetime.strptime(rec.Date, '%Y-%m-%d %H:%M:%S')
#                             weekday = date_obj.strftime('%A')
#                             if weekday in week_dates and not week_dates[weekday]:
#                                 week_dates[weekday] = rec.Date
                        
#                         # Return the first available date in order: Monday, Wednesday, Friday
#                         return week_dates['Monday'] or week_dates['Wednesday'] or week_dates['Friday']

#                     for week, records in week_data.items():
#                         # Sort and get top 10 usage records for the week
#                         records_sorted = sorted(records, key=lambda x: float(x.Usage), reverse=True)[:10]

#                         # Get the valid date for the week (Monday, Wednesday, or Friday)
#                         selected_date = get_valid_week_date(records_sorted)

#                         if not selected_date:
#                             continue  # If no valid date, skip this week

#                         # Prepare data for the table
#                         report_data = [
#                             {'Wireless_Number': rec.Wireless_Number, 'User_Name': rec.User_Name, 'Usage': rec.Usage}
#                             for rec in records_sorted
#                         ]
#                         df = pd.DataFrame(report_data)

#                         # Write the title, week date, and vendor
#                         worksheet.merge_cells(start_row=row_offset, start_column=start_column, end_row=row_offset, end_column=start_column)
#                         worksheet.cell(row=row_offset, column=start_column).value = f"Top 10 Users"
#                         worksheet.cell(row=row_offset, column=start_column).font = title_font

#                         worksheet.cell(row=row_offset, column=start_column + 1).value = f"Week {week} - {selected_date}"
#                         worksheet.cell(row=row_offset, column=start_column + 1).font = title_font
#                         worksheet.cell(row=row_offset, column=start_column + 1).alignment = Alignment(horizontal="center")

#                         # Write the vendor in the third column
#                         worksheet.cell(row=row_offset, column=start_column + 2).value = f"Vendor: {vendor}"
#                         worksheet.cell(row=row_offset, column=start_column + 2).font = title_font
#                         worksheet.cell(row=row_offset, column=start_column + 2).alignment = Alignment(horizontal="center")

#                         # Write the headers
#                         headers = ['Wireless_Number', 'User_Name', 'Usage']
#                         for col_num, header in enumerate(headers, start=start_column):
#                             cell = worksheet.cell(row=row_offset + 1, column=col_num)
#                             cell.value = header
#                             cell.fill = light_blue_fill
#                             cell.font = Font(bold=True, color="FFFFFF")
#                             cell.alignment = Alignment(horizontal="center")

#                         # Fill the data into the table
#                         for row_num, row_data in enumerate(df.itertuples(index=False), row_offset + 2):
#                             for col_num, value in enumerate(row_data, start=start_column):
#                                 cell = worksheet.cell(row=row_num, column=col_num)
#                                 cell.value = value
#                                 cell.fill = light_yellow_fill

#                         # Adjust column widths dynamically based on content
#                         for col_num in range(start_column, start_column + len(headers)):
#                             max_len = max(len(str(cell.value)) for cell in worksheet[get_column_letter(col_num)])
#                             worksheet.column_dimensions[get_column_letter(col_num)].width = max_len + 5

#                         # Move start_column to the next set of columns for the next week's table
#                         start_column += 5  # Shift the starting column by 5 to leave space between tables

#                     # Save the workbook to the output stream
#                     workbook.save(output)
#                     output.seek(0)

#                     report_file = Downloaded_reports(
#                         report_type = report_type,
#                         kwargs = self.change_kwargs(filter_kwargs)
#                     )
#                     report_file.file.save(f'top_10_usage_report_{month}.xlsx', ContentFile(output.getvalue()))
#                     saveuserlog(request.user, f"Unilled data report excel file of {month} downloaded.")

#                     return Response({"data": report_file.file.url, "message" : "Excel file generated successfully!"}, status=status.HTTP_200_OK)
#                     # Create the HTTP response with the Excel file
#                     # response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#                     # response['Content-Disposition'] = f'attachment; filename=top_10_usage_report_{month}.xlsx'

#                     # return response
                
#                 except Exception as e:
#                     return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)


from rest_framework.views import APIView
from rest_framework import status, permissions
from rest_framework.response import Response

from authenticate.views import saveuserlog
from OnBoard.Organization.models import Organizations
from OnBoard.Company.models import Company
from Dashboard.ModelsByPage.DashAdmin import Vendors
from OnBoard.Ban.models import BaseDataTable

from django.core.files.base import ContentFile

import pandas as pd
from collections import defaultdict
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

from ..models import Report_Unbilled_Data, Downloaded_reports
from ..ser import (
    showBanSerializer,
    showUnbilledReport,
    OrganizationShowSerializer,
    VendorShowSerializer,
)

def _fmt_db_date(val: str) -> str:
    """
    Convert 'YYYY-MM-DD HH:MM:SS' (or 'YYYY-MM-DD') → 'MM-dd-YYYY'.
    If parsing fails, returns the original string.
    """
    if not val:
        return ""
    s = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime("%m-%d-%Y")   # <- dash format
        except Exception:
            pass
    # fallback: try first 10 chars as YYYY-MM-DD
    try:
        dt = datetime.strptime(s[:10], "%Y-%m-%d")
        return dt.strftime("%m-%d-%Y")
    except Exception:
        return s




HEADER_BLUE = "FF000066"        # dark blue ARGB
WEEK_COLORS = {1: "FFFF99", 2: "CCFFCC", 3: "FFCCCC", 4: "CCCCFF", 5: "FFCC99"}
MEDIUM_ORANGE = "FFFF8503"
ALT_ROW_LIGHT_YELLOW = "FFFFE0"
ALT_ROW_LIGHT_GREY = "FFD3D3D3"

def _nanlike(val) -> bool:
    """True if value is None/blank/'nan'/'null'/'none' (case-insensitive)."""
    if val is None:
        return True
    s = str(val).strip().lower()
    return s in {"", "nan", "null", "none"}

def _auto_width(ws, start_col=1, end_col=None):
    if end_col is None:
        end_col = ws.max_column
    for col in range(start_col, end_col + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for cell in ws[col_letter]:
            v = cell.value
            if v is not None:
                ln = len(str(v))
                if ln > max_len:
                    max_len = ln
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

def _style_header_row(ws, row_idx: int, start_col: int = 1, end_col: int | None = None):
    """Bold, centered, blue background, white text for a header row."""
    end_col = end_col or ws.max_column
    fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill

def _write_df(ws, df, start_row=1, start_col=1):
    # headers
    for j, col_name in enumerate(df.columns, start_col):
        ws.cell(start_row, j, col_name)
    # rows
    for i, row in enumerate(df.itertuples(index=False), start_row + 1):
        for j, val in enumerate(row, start_col):
            ws.cell(i, j, val)

def _get_valid_week_date(records):
    # Monday -> Wednesday -> Friday; return 'MM-dd-YYYY'
    pick = {'Monday': None, 'Wednesday': None, 'Friday': None}
    for r in records:
        try:
            d = datetime.strptime(r.Date, '%Y-%m-%d %H:%M:%S')
        except Exception:
            continue
        wd = d.strftime('%A')
        if wd in pick and not pick[wd]:
            pick[wd] = r.Date
    chosen = pick['Monday'] or pick['Wednesday'] or pick['Friday']
    return _fmt_db_date(chosen) if chosen else None



def _write_df(ws, df, start_row=1, start_col=1):
    # write headers
    for j, col_name in enumerate(df.columns, start_col):
        ws.cell(start_row, j, col_name)
    # write rows
    for i, row in enumerate(df.itertuples(index=False), start_row + 1):
        for j, val in enumerate(row, start_col):
            ws.cell(i, j, val)

# -------------------------------
# Sheet 1:






def _build_sheet_unbilled_on(ws, qs):
    """
    'Unbilled Data Report' on existing ws:
    - Row 1: merged week headers (only top-left cell gets value)
    - Row 2: column headers (styled)
    - Row 3..: data
    - Skips records with blank/'nan' Wireless_Number
    """
    report_rows, week_dates = [], {}

    for record in qs:
        if _nanlike(record.Wireless_Number):
            continue  # 🚫 skip bad wireless numbers

        week = int(record.Week)
        date = record.Date
        week_dates.setdefault(week, set()).add(date)

        row_obj = next((r for r in report_rows if r['wirelessNumber'] == record.Wireless_Number), None)
        if not row_obj:
            report_rows.append({
                "accountNumber": record.Account_Number,
                "wirelessNumber": record.Wireless_Number,
                "userName": record.User_Name,
                "vendor": record.vendor.name if getattr(record, "vendor", None) else "",
                record.Date: record.Usage,
            })
        else:
            row_obj[record.Date] = record.Usage

    if not report_rows:
        ws.cell(1, 1, "No data available for Unbilled Data Report").font = Font(bold=True)
        return

    sorted_weeks = sorted(week_dates.keys())
    sorted_week_dates = {w: sorted(dates) for w, dates in week_dates.items()}
    all_dates = [d for w in sorted_weeks for d in sorted_week_dates[w]]

    df = pd.DataFrame(report_rows)
    if df.empty:
        ws.cell(1, 1, "No data available for Unbilled Data Report").font = Font(bold=True)
        return

    # extra guard
    bad = df['wirelessNumber'].astype(str).str.strip().str.lower().isin(["", "nan", "null", "none"])
    df = df[~bad].copy()

    base_cols = ['accountNumber', 'wirelessNumber', 'userName', 'vendor']
    df = df[base_cols + all_dates]
    date_rename = {d: _fmt_db_date(d) for d in all_dates}
    df = df.rename(columns={
        "accountNumber": "Account Number",
        "wirelessNumber": "Wireless Number",
        "userName": "User Name",
        "vendor": "Vendor",
        **date_rename

    })

    # Write headers (row 2) + data (row 3..)
    _write_df(ws, df, start_row=2, start_col=1)

    # Row 1: merged + styled week headers from col 5 onwards
    current_col = 5
    for w in sorted_weeks:
        count = len(sorted_week_dates[w])
        if count <= 0:
            continue
        ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + count - 1)

        # set value only in top-left of the merge
        top = ws.cell(1, current_col)
        top.value = f"Week {w}"
        top.alignment = Alignment(horizontal="center", vertical="center")
        top.font = Font(bold=True, color="FFFFFF")
        top.fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")

        # style the rest (no .value writes!)
        for c in range(current_col + 1, current_col + count):
            mc = ws.cell(1, c)
            mc.alignment = Alignment(horizontal="center", vertical="center")
            mc.font = Font(bold=True, color="FFFFFF")
            mc.fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")

        current_col += count

    # Week color bands for data rows only
    current_col = 5
    for w in sorted_weeks:
        count = len(sorted_week_dates[w])
        if count <= 0:
            continue
        band_color = WEEK_COLORS.get(w, "FFFFFFFF")
        fill = PatternFill(start_color=band_color, end_color=band_color, fill_type="solid")
        for off in range(count):
            for r in range(3, ws.max_row + 1):
                ws.cell(r, current_col + off).fill = fill
        current_col += count

    # Style the table header row (row 2)
    _style_header_row(ws, row_idx=2, start_col=1, end_col=ws.max_column)

    ws.freeze_panes = "A3"
    _auto_width(ws)






def _build_sheet_zero_usage(workbook, qs, month, gap_cols: int = 1):
    """
    Zero Usage sheet — WEEK blocks with DATE columns:
      [Wireless Number | User Name | Device | Upgrade Eligibility Date | <date1> | <date2> | ...]
    - Vendor is shown inside the merged week title.
    - Fixed widths per block so columns (esp. Wireless Number) stay tight.
    - Skips any record with Wireless_Number blank/'nan'/'null'/'none'.
    - Only includes rows where Usage == 0.
    """
    ws = workbook.create_sheet(title="Zero Usage Report (Weekly)")

    # Title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
    t = ws.cell(1, 1, f"{month} Zero Usage Report (Weekly • Date-wise)")
    t.font = Font(bold=True, color="FFFFFF", size=14)
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")

    from collections import defaultdict
    week_dates = defaultdict(set)                 # week -> {yyyy-mm-dd}
    week_rows  = defaultdict(lambda: defaultdict(dict))  # week -> {(wln, user, dev, upg) -> {date:"0 GB"}}
    vendors    = set()

    # collect rows (zero usage only, valid wireless only)
    for rec in qs:
        if _nanlike(rec.Wireless_Number):
            continue

        # only zero usage
        try:
            usage_val = float(rec.Usage)
        except Exception:
            try:
                usage_val = float(str(rec.Usage).strip())
            except Exception:
                continue
        if usage_val != 0:
            continue

        # date key
        try:
            dt = datetime.strptime(rec.Date, "%Y-%m-%d %H:%M:%S")
            date_key = dt.strftime("%Y-%m-%d")
        except Exception:
            date_key = str(rec.Date).split()[0]

        # week number
        try:
            week_no = int(rec.Week)
        except Exception:
            week_no = 0

        week_dates[week_no].add(date_key)
        row_key = (rec.Wireless_Number, rec.User_Name, rec.Device, rec.Upgrade_Eligibilty_Date)
        week_rows[week_no][row_key][date_key] = "0 GB"

        if getattr(rec, "vendor", None):
            vendors.add(rec.vendor.name)

    if not week_rows:
        ws.cell(3, 1, "No zero-usage rows for the selected filters.").font = Font(bold=True)
        ws.freeze_panes = "A3"
        return

    # styles
    title_font        = Font(bold=True, size=12)
    light_yellow_fill = PatternFill(start_color=ALT_ROW_LIGHT_YELLOW, end_color=ALT_ROW_LIGHT_YELLOW, fill_type="solid")
    light_grey_fill   = PatternFill(start_color=ALT_ROW_LIGHT_GREY,   end_color=ALT_ROW_LIGHT_GREY,   fill_type="solid")

    start_col = 1
    row_off   = 3
    base_headers = ['Wireless Number', 'User Name', 'Device', 'Upgrade Eligibility Date']
    vendor_str = ', '.join(sorted(vendors)) if vendors else ""

    for week_no in sorted(week_rows.keys()):
        dates_sorted_raw = sorted(week_dates[week_no])              # 'YYYY-MM-DD'
        date_headers      = [_fmt_db_date(d) for d in dates_sorted_raw]  # 'MM dd yyyy'
        headers = base_headers + date_headers

        # Merge the full block width for the week title (no separate vendor column)
        block_start = start_col
        block_end   = start_col + len(headers) - 1
        ws.merge_cells(start_row=row_off, start_column=block_start, end_row=row_off, end_column=block_end)
        top = ws.cell(row_off, block_start)
        top.value = f"Week {week_no}" + (f" • {vendor_str}" if vendor_str else "")
        top.font = title_font
        top.alignment = Alignment(horizontal="left", vertical="center")

        # Column headers (styled)
        for idx, h in enumerate(headers, start=start_col):
            ws.cell(row=row_off + 1, column=idx, value=h)
        _style_header_row(ws, row_idx=row_off + 1, start_col=start_col, end_col=start_col + len(headers) - 1)

        # Data rows (alternating fill)
        for r_i, (wireless, user, device, upgrade) in enumerate(week_rows[week_no].keys(), start=row_off + 2):
            upgrade_fmt = _fmt_db_date(upgrade)
            vals = [wireless, user, device, upgrade_fmt]
            date_map = week_rows[week_no][(wireless, user, device, upgrade)]
            vals.extend(date_map.get(d, "") for d in dates_sorted_raw)  # use raw keys; headers are pretty


            fill = light_yellow_fill if (r_i - (row_off + 2)) % 2 == 0 else light_grey_fill
            for c_idx, val in enumerate(vals, start=start_col):
                ws.cell(row=r_i, column=c_idx, value=val).fill = fill

        # ---- Fixed widths per block (prevents "Wireless Number" from expanding) ----
        def _set_w(col_idx, width):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Base columns
        _set_w(block_start + 0, 14)  # Wireless Number
        _set_w(block_start + 1, 18)  # User Name
        _set_w(block_start + 2, 18)  # Device
        _set_w(block_start + 3, 22)  # Upgrade Eligibility Date

        # Date columns (tight)
        for i in range(len(dates_sorted_raw)):
            _set_w(block_start + 4 + i, 10)  # "2025-09-01" fits

        # Add a tiny gap, then move to next block
        if gap_cols > 0:
            gap_start = block_end + 1
            for gc in range(gap_start, gap_start + gap_cols):
                ws.column_dimensions[get_column_letter(gc)].width = 2
            start_col = gap_start + gap_cols
        else:
            start_col = block_end + 1

    ws.freeze_panes = "A3"







def _build_sheet_top10(workbook, qs, month, vendor_label):
    ws = workbook.create_sheet(title="Top 10 Usage Report")

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
    t = ws.cell(1, 1, f"{month} Usage Report (Weekly)")
    t.font = Font(bold=True, color="FFFFFF", size=14)
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")

    # group by week (skip bad wireless)
    week_map = defaultdict(list)
    for rec in qs:
        if _nanlike(rec.Wireless_Number):
            continue
        week_map[rec.Week].append(rec)

    title_font = Font(bold=True, size=12)
    start_col = 1
    row_off = 3

    for week, recs in week_map.items():
        # sort by usage desc, then unique by wireless number
        recs_sorted = sorted(recs, key=lambda x: float(x.Usage), reverse=True)
        seen = set()
        top10 = []
        for r in recs_sorted:
            if r.Wireless_Number in seen:
                continue
            seen.add(r.Wireless_Number)
            top10.append(r)
            if len(top10) == 10:
                break

        selected_date = _get_valid_week_date(top10)
        df = pd.DataFrame([{
            'Wireless Number': r.Wireless_Number,
            'User Name': r.User_Name,
            'Data Usage (GB)': float(r.Usage),
        } for r in top10])

        # sub-title row
        ws.merge_cells(start_row=row_off, start_column=start_col, end_row=row_off, end_column=start_col)
        ws.cell(row_off, start_col, "Top 10 Users").font = title_font
        ws.cell(row_off, start_col + 1, f"Week {week}").font = title_font
        ws.cell(row_off, start_col + 1).alignment = Alignment(horizontal="center")
        ws.cell(row_off, start_col + 2, f"Vendor: {vendor_label}").font = title_font
        ws.cell(row_off, start_col + 2).alignment = Alignment(horizontal="center")

        # headers (row_off+1) + style
        headers = ['Wireless Number', 'User Name', 'Data Usage (GB)']
        for idx, h in enumerate(headers, start=start_col):
            ws.cell(row=row_off + 1, column=idx, value=h)
        _style_header_row(ws, row_idx=row_off + 1, start_col=start_col, end_col=start_col + len(headers) - 1)

        # data
        for r_idx, row in enumerate(df.itertuples(index=False), row_off + 2):
            for c_idx, val in enumerate(row, start=start_col):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                # number format for usage col
                if c_idx == start_col + 2:
                    cell.number_format = "0.00"

        # widths
        for c in range(start_col, start_col + len(headers)):
            col_letter = get_column_letter(c)
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in ws[col_letter])
            ws.column_dimensions[col_letter].width = min(max_len + 5, 60)

        start_col += 5

    ws.freeze_panes = "A3"

from django.db.models import Q, FloatField
from django.db.models.functions import Cast
from rest_framework.viewsets import ViewSet


class ViewUnbilledReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def __init__(self, **kwargs):
        self.report_type = "Unbilled_Data_Usage"
        self.current_year = datetime.now().year
        self.data = None
        self.unique_dates = None
        super().__init__(**kwargs)

    def get(self, request, sub_type=None, *args, **kwargs):

        company = request.GET.get('company')
        sub_company = request.GET.get('sub_company')
        ban = request.GET.get('ban')
        report_type = request.GET.get('report_type')
        month = request.GET.get('month')
        vendor = request.GET.get('vendor')
        year = request.GET.get('year')
        print(company, sub_company, ban, report_type, month, vendor, year, sub_type)
        if not (company and sub_company and ban and report_type and month and vendor and year):
            return Response({"message":"Missing required fields."},status=status.HTTP_400_BAD_REQUEST)
        
        filter_kwargs = {}
        if company:
            filter_kwargs['company'] = Company.objects.filter(Company_name=company)[0]
        if sub_company:
            filter_kwargs['organization'] = Organizations.objects.filter(Organization_name=sub_company)[0]
        if ban:
            filter_kwargs['Account_Number'] = ban
        if report_type:
            filter_kwargs['Report_Type'] = report_type
        if month:
            filter_kwargs['Month'] = month
        if year:
            filter_kwargs['Year'] = year
        if vendor:
            filter_kwargs['vendor'] = Vendors.objects.filter(name=vendor)[0]

        filtered_data = Report_Unbilled_Data.objects.filter(company=filter_kwargs['company'])
        if not filtered_data.exists():
            return Response({"message":f"no unbilled report found for {filter_kwargs['organization'].Organization_name}"},status=status.HTTP_400_BAD_REQUEST)
        filtered_data = filtered_data.filter(organization=filter_kwargs['organization'])
        if not filtered_data.exists():
            return Response({"message":f"no unbilled report found for {filter_kwargs['organization'].Organization_name}"},status=status.HTTP_400_BAD_REQUEST)
        filtered_data = filtered_data.filter(vendor=filter_kwargs['vendor'])
        if not filtered_data.exists():
            return Response({"message":f"no unbilled report found for {filter_kwargs['vendor'].name}"},status=status.HTTP_400_BAD_REQUEST)
        filtered_data = filtered_data.filter(Account_Number=filter_kwargs['Account_Number'])
        if not filtered_data.exists():
            return Response({"message":f"no unbilled report found for {filter_kwargs['Account_Number']}"},status=status.HTTP_400_BAD_REQUEST)
        filtered_data = filtered_data.filter(Report_Type=filter_kwargs['Report_Type'])
        if not filtered_data.exists():
            return Response({"message":f"no unbilled report found for {filter_kwargs['Report_Type']}"},status=status.HTTP_400_BAD_REQUEST)
        filtered_data = filtered_data.filter(Month=filter_kwargs['Month'])
        if not filtered_data.exists():
            return Response({"message":f"no unbilled report found for {filter_kwargs['Month']}"},status=status.HTTP_400_BAD_REQUEST)
        filtered_data = filtered_data.filter(Year=filter_kwargs['Year'])
        if not filtered_data.exists():
            return Response({"message":f"no unbilled report found for {filter_kwargs['Year']}"},status=status.HTTP_400_BAD_REQUEST)
        if not sub_type:
            # filter_kwargs = {}
            # company = Company.objects.get(Company_name=request.user.company.Company_name) if request.user.company else None
            # if company:
            #     filter_kwargs['company'] = company
            # filter_kwargs['Report_Type'] = self.report_type

            # filtered_data = Report_Unbilled_Data.objects.filter(**filter_kwargs)

            report_data = []
            for record in filtered_data:
                existing_report = next((r for r in report_data if r['Wireless_Number'] == record.Wireless_Number), None)
                if not existing_report:
                    report_data.append({
                        "Account_Number": record.Account_Number,
                        'Wireless_Number': record.Wireless_Number,
                        'user_Name': record.User_Name,
                        'vendor': record.vendor,
                        record.Date: record.Usage,
                    })
                else:
                    existing_report[record.Date] = record.Usage

            self.unique_dates = sorted(
                list(set(
                    datetime.strptime(record.Date, '%Y-%m-%d %H:%M:%S').date()
                    for record in filtered_data
                )),
                key=lambda date: date
            )
            self.data = showUnbilledReport(filtered_data, many=True)
        else:
            if sub_type == "getTop10byWeek":
                try:
                    top10 = (
                        filtered_data
                        .annotate(usage_float=Cast("Usage", FloatField()))
                        .filter(usage_float__gt=0) 
                        .order_by("-usage_float")[:10]
                    )
                    self.data = showUnbilledReport(top10, many=True)
                except Exception as e:
                    return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            elif sub_type == "getZeroUsageReportUnbilledData":
                try:
                    filtered_data = filtered_data.filter(Month=filter_kwargs['Month'])
                    if not filtered_data.exists():
                        return Response({"message":f"no unbilled report found for {filter_kwargs['Month']}"},status=status.HTTP_400_BAD_REQUEST)
                    self.data = showUnbilledReport(filtered_data, many=True)
                except Exception:
                    return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            saveuserlog(request.user, f"Unilled data report excel file of {month}-{year} downloaded.")
        return Response(
            {
                "data": self.data.data if self.data else None,
                "unique_dates": self.unique_dates if self.report_type == "Unbilled_Data_Usage" else None
            },
            status=status.HTTP_200_OK,
        )

    def change_kwargs(self, kwargs):
        if 'organization' in kwargs and str(kwargs['organization']) != str:
            kwargs['organization'] = kwargs['organization'].Organization_name
        if 'company' in kwargs and str(kwargs['company']) != str:
            kwargs['company'] = kwargs['company'].Company_name
        if 'vendor' in kwargs and str(kwargs['vendor']) != str:
            kwargs['vendor'] = kwargs['vendor'].name
        return kwargs

    def post(self, request, sub_type=None, *args, **kwargs):
        """
        No frontend or URL changes required.
        Any sub_type now returns ONE Excel with THREE sheets.
        """
        filter_kwargs = {}
        action = request.data.get('action')
        _type = request.data.get('type')
        company = request.data.get('company')
        sub_company = request.data.get('sub_company')
        ban = request.data.get('ban')
        month = request.data.get('month')
        report_type = request.data.get('report_type')
        year = request.data.get('year')
        vendor = request.data.get('vendor')

        if not (company and sub_company and ban and report_type and month and vendor and year):
            return Response({"message": "Missing required fields."}, status=status.HTTP_400_BAD_REQUEST)

        # filters
        filter_kwargs['company'] = Company.objects.filter(Company_name=company)[0]
        filter_kwargs['organization'] = Organizations.objects.filter(Organization_name=sub_company)[0]
        filter_kwargs['Account_Number'] = ban
        filter_kwargs['Report_Type'] = report_type
        filter_kwargs['Month'] = month
        filter_kwargs['Year'] = year
        filter_kwargs['vendor'] = Vendors.objects.filter(name=vendor)[0]

        # validate chain
        qs = Report_Unbilled_Data.objects.filter(company=filter_kwargs['company'])
        if not qs.exists():
            return Response({"message": f"no unbilled report found for {filter_kwargs['company'].Company_name}"}, status=status.HTTP_400_BAD_REQUEST)
        qs = qs.filter(organization=filter_kwargs['organization'])
        if not qs.exists():
            return Response({"message": f"no unbilled report found for {filter_kwargs['organization'].Organization_name}"}, status=status.HTTP_400_BAD_REQUEST)
        qs = qs.filter(Account_Number=filter_kwargs['Account_Number'])
        if not qs.exists():
            return Response({"message": f"no unbilled report found for {filter_kwargs['Account_Number']}"}, status=status.HTTP_400_BAD_REQUEST)
        qs = qs.filter(Report_Type=filter_kwargs['Report_Type'])
        if not qs.exists():
            return Response({"message": f"no unbilled report found for {filter_kwargs['Report_Type']}"}, status=status.HTTP_400_BAD_REQUEST)
        qs = qs.filter(Month=filter_kwargs['Month'])
        if not qs.exists():
            return Response({"message": f"no unbilled report found for {filter_kwargs['Month']}"}, status=status.HTTP_400_BAD_REQUEST)
        qs = qs.filter(Year=filter_kwargs['Year'])
        if not qs.exists():
            return Response({"message": f"no unbilled report found for {filter_kwargs['Year']}"}, status=status.HTTP_400_BAD_REQUEST)
        qs = qs.filter(vendor=filter_kwargs['vendor'])
        if not qs.exists():
            return Response({"message": f"no unbilled report found for {filter_kwargs['vendor'].name}"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Always produce a single workbook with three sheets
            output = BytesIO()
            workbook = Workbook()

            # Keep default sheet; rename and use it for Unbilled
            ws0 = workbook.active
            ws0.title = "Unbilled Data Report"
            _build_sheet_unbilled_on(ws0, qs)

            # Other two tabs
            _build_sheet_zero_usage(workbook, qs, month)
            _build_sheet_top10(workbook, qs, month, filter_kwargs['vendor'].name)

            # Finalize
            workbook.save(output)
            output.seek(0)

            report_file = Downloaded_reports(
                report_type=report_type,
                kwargs=self.change_kwargs(filter_kwargs.copy())
            )
            report_file.file.save(f'unbilled_combined_{month}_{year}.xlsx', ContentFile(output.getvalue()))
            saveuserlog(request.user, f"Unbilled data combined excel ({month}-{year}) downloaded.")
            return Response({"data": report_file.file.url, "message": "Combined Excel generated successfully!"}, status=status.HTTP_200_OK)

        except Exception as e:
            # Log/print if you like
            print("Export error:", e)
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)
