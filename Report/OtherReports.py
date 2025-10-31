
from django.shortcuts import render
from django.shortcuts import render, HttpResponse
from rest_framework.views import APIView
from rest_framework import status
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Alignment, Font
from OnBoard.Ban.models import UniquePdfDataTable, BaseDataTable, BaselineDataTable
from OnBoard.Ban.ser import baseallserializer
from openpyxl.utils import get_column_letter
from datetime import datetime
from OnBoard.Organization.models import Division, Organizations, Links
from OnBoard.Location.models import Location
from rest_framework.decorators import api_view, permission_classes
from django.db.models import Q
import json
from django.db.models import Q, Func, F
from django.db.models.functions import Cast
from django.db.models import DateField
from django.db import models

def is_within_range(raw_date, start, end):
    bill_date = parse_bill_date(raw_date)
    return bill_date and start <= bill_date <= end


def parse_bill_date(raw_date: str):
    """
    Tries to parse a date string in various common formats like:
    'Jan 15, 2024', 'July 15 2025', 'Jan 15 2024', 'Jun 03 2025'.
    Returns a datetime object or None if parsing fails.
    """
    if not raw_date:
        return None

    raw_date = raw_date.strip()
    possible_formats = [
        '%b %d, %Y',   # Jan 15, 2024
        '%B %d, %Y',   # January 15, 2024
        '%b %d %Y',    # Jan 15 2024
        '%B %d %Y'     # January 15 2024
    ]

    for fmt in possible_formats:
        try:
            return datetime.strptime(raw_date, fmt)
        except ValueError:
            continue
    return None

class Get_Asset_Tracking_asset_master_report(APIView):
    def get(self,request):
        sub_company = request.query_params.get('sub_company')
        company = request.query_params.get('company')
        location = request.query_params.get('location')
        data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(sub_company=sub_company,company=company,location=location)
        serializer = baseallserializer(data,many=True)
        df = pd.DataFrame(serializer.data)
        df = df[['Asset_Type', 'Asset_Tag', 'Serial', 'InvoiceNumber', 'Date_Acquired', 'Acquired_For', 'location']]
        required_data = df.to_dict(orient="records")
        return Response(required_data)
    
class Download_Asset_Tracking_asset_master_report(APIView):
    def get(self, request):
        sub_company = request.query_params.get('sub_company')
        company = request.query_params.get('company')
        location = request.query_params.get('location')
        data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(sub_company=sub_company, company=company, location=location)
        serializer = baseallserializer(data, many=True)

        try:
            df = pd.DataFrame(serializer.data)
            df = df[['Asset_Type', 'Asset_Tag', 'Serial', 'InvoiceNumber', 'Date_Acquired', 'Acquired_For', 'location']]

            with BytesIO() as buffer:
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Asset Tracking Asset Master Report', startrow=1)
                    workbook = writer.book
                    sheet = writer.sheets['Asset Tracking Asset Master Report']
                    
                    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
                    title_cell = sheet.cell(row=1, column=1)
                    title_cell.value = "Asset Tracking Asset Master Report"
                    title_cell.font = Font(size=18, bold=True)
                    title_cell.alignment = Alignment(horizontal="center", vertical="center")
                    title_cell.fill = PatternFill(start_color="A8CD89", end_color="A8CD89", fill_type="solid")
                    
                    column_fill = PatternFill(start_color="F4E0AF", end_color="F4E0AF", fill_type="solid")
                    for cell in sheet[2]:
                        cell.fill = column_fill
                        cell.font = Font(bold=True, color="000000")
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    for col_num, _ in enumerate(df.columns, 1):
                        column_width = max(len(df.columns[col_num - 1]) + 2, 15)
                        sheet.column_dimensions[get_column_letter(col_num)].width = column_width

                buffer.seek(0)
                response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="Asset_Tracking_Asset_Master_Report.xlsx"'
                return response

        except Exception as e:
            print(str(e))
            return Response("Data not found", status=404)
        

class Get_Asset_Transactions(APIView):
    def get(self,request):
        sub_company = request.query_params.get('sub_company')
        company = request.query_params.get('company')
        location = request.query_params.get('location')
        Asset = request.query_params.get('Asset')
        Transacton_Type = request.query_params.get('Transacton_Type')
        data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(sub_company=sub_company,company=company,location=location,Asset=Asset,Transaction_Type=Transacton_Type)
        serializer = baseallserializer(data,many=True)
        return Response(serializer.data)
    
class Download_Asset_Transactions(APIView):
    def get(self, request):
        sub_company = request.query_params.get('sub_company')
        company = request.query_params.get('company')
        location = request.query_params.get('location')
        Asset = request.query_params.get('Asset')
        Transacton_Type = request.query_params.get('Transacton_Type')
        data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
            sub_company=sub_company, 
            company=company, 
            location=location, 
            Asset=Asset, 
            Transaction_Type=Transacton_Type
        )
        serializer = baseallserializer(data, many=True)

        try:
            df = pd.DataFrame(serializer.data)

            with BytesIO() as buffer:
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Asset Transaction Report', startrow=1)
                    workbook = writer.book
                    sheet = writer.sheets['Asset Transaction Report']
                    
                    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
                    title_cell = sheet.cell(row=1, column=1)
                    title_cell.value = "Asset Transaction Report"
                    title_cell.font = Font(size=18, bold=True)
                    title_cell.alignment = Alignment(horizontal="center", vertical="center")
                    title_cell.fill = PatternFill(start_color="A8CD89", end_color="A8CD89", fill_type="solid")
                    
                    column_fill = PatternFill(start_color="F4E0AF", end_color="F4E0AF", fill_type="solid")
                    for cell in sheet[2]:
                        cell.fill = column_fill
                        cell.font = Font(bold=True, color="000000")
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    for col_num, _ in enumerate(df.columns, 1):
                        column_width = max(len(df.columns[col_num - 1]) + 2, 15)
                        sheet.column_dimensions[get_column_letter(col_num)].width = column_width

                buffer.seek(0)
                response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="Asset_Transaction_Report.xlsx"'
                return response

        except Exception as e:
            print(str(e))
            return Response("Data not found", status=404)


class Get_Outstanding_Loaners_report(APIView):
    def get(self,request):
        sub_company = request.query_params.get('sub_company')
        company = request.query_params.get('company')
        location = request.query_params.get('location')
        data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(sub_company=sub_company,company=company,location=location)
        serializer = baseallserializer(data,many=True)
        df = pd.DataFrame(serializer.data)
        df = df[['location', 'Asset_Type', 'Novus_Tag', 'Serial', 'InvoiceNumber', 'Date_Acquired', 'Acquired_For', 'Employee', 'Trans_Date']]
        required_data = df.to_dict(orient="records")
        return Response(required_data)
    
class Download_Outstanding_Loaners_report(APIView):
    def get(self, request):
        sub_company = request.query_params.get('sub_company')
        company = request.query_params.get('company')
        location = request.query_params.get('location')
        data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(sub_company=sub_company, company=company, location=location)
        serializer = baseallserializer(data, many=True)

        try:
            df = pd.DataFrame(serializer.data)
            df = df[['location', 'Asset_Type', 'Novus_Tag', 'Serial', 'InvoiceNumber', 'Date_Acquired', 
                     'Acquired_For', 'Employee', 'Trans_Date']]

            with BytesIO() as buffer:
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Outstanding Loaners Report', startrow=1)
                    workbook = writer.book
                    sheet = writer.sheets['Outstanding Loaners Report']
                    
                    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
                    title_cell = sheet.cell(row=1, column=1)
                    title_cell.value = "Outstanding Loaners Report"
                    title_cell.font = Font(size=18, bold=True)
                    title_cell.alignment = Alignment(horizontal="center", vertical="center")
                    title_cell.fill = PatternFill(start_color="A8CD89", end_color="A8CD89", fill_type="solid")
                    
                    peach_fill = PatternFill(start_color="F4E0AF", end_color="F4E0AF", fill_type="solid")
                    for cell in sheet[2]:
                        cell.fill = peach_fill
                        cell.font = Font(bold=True, color="000000")
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    for col_num, _ in enumerate(df.columns, 1):
                        column_width = max(len(df.columns[col_num - 1]) + 2, 15)
                        sheet.column_dimensions[get_column_letter(col_num)].width = column_width

                buffer.seek(0)
                response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="Outstanding_Loaners_Report.xlsx"'
                return response

        except Exception as e:
            print(str(e))
            return Response("Data not found", status=404)
        


class Get_Order_Tracking_report(APIView):
    def get(self,request):
        sub_company = request.query_params.get('sub_company')
        company = request.query_params.get('company')
        vendor = request.query_params.get('vendor')
        data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(sub_company=sub_company,company=company,vendor=vendor)
        serializer = baseallserializer(data,many=True)
        df = pd.Dataframe(serializer.data)
        df = df[['Order_Title', 'location', 'Order_Type', 'vendor', 'Service_Product', 'Requestor_Name', 'accountnumber', 'Order_Date', 'Requested_Date', 'Date_Due', 'Completed_Date']]
        required_data = df.to_dict(orient="records")
        return Response(required_data)

    
class Download_Order_Tracking_report(APIView):
    def get(self, request):
        sub_company = request.query_params.get('sub_company')
        company = request.query_params.get('company')
        vendor = request.query_params.get('vendor')
        data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(sub_company=sub_company, company=company, vendor=vendor)
        serializer = baseallserializer(data, many=True)

        try:
            df = pd.DataFrame(serializer.data)
            df = df[['Order_Title', 'location', 'Order_Type', 'vendor', 'Service_Product', 
                     'Requestor_Name', 'accountnumber', 'Order_Date', 'Requested_Date', 'Date_Due', 'Completed_Date']]

            with BytesIO() as buffer:
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Order Tracking Report', startrow=1)
                    workbook = writer.book
                    sheet = writer.sheets['Order Tracking Report']
                    
                    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
                    title_cell = sheet.cell(row=1, column=1)
                    title_cell.value = "Order Tracking Report"
                    title_cell.font = Font(size=18, bold=True)
                    title_cell.alignment = Alignment(horizontal="center", vertical="center")
                    title_cell.fill = PatternFill(start_color="A8CD89", end_color="A8CD89", fill_type="solid")
                    
                    column_fill = PatternFill(start_color="F4E0AF", end_color="F4E0AF", fill_type="solid")
                    for cell in sheet[2]:
                        cell.fill = column_fill
                        cell.font = Font(bold=True, color="000000")
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    for col_num, _ in enumerate(df.columns, 1):
                        column_width = max(len(df.columns[col_num - 1]) + 2, 15)
                        sheet.column_dimensions[get_column_letter(col_num)].width = column_width

                buffer.seek(0)
                response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="Order_Tracking_Report.xlsx"'
                return response

        except Exception as e:
            print(str(e))
            return Response("Data not found", status=404)
        
class Get_Contact_report(APIView):
    def get(self,request):
        sub_company = request.query_params.get('sub_company')
        company = request.query_params.get('company')
        vendor = request.query_params.get('vendor')
        data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(sub_company=sub_company,company=company,vendor=vendor)
        serializer = baseallserializer(data,many=True)
        df = pd.Dataframe(serializer.data)
        df = df[["vendor", "Contract_Person", "Date_Signed", "Start_Date", "End_Date", "Term", "Contract", "accountnumber", "SVC_Type", "Auto_Renew", "Attachments"]]
        required_data = df.to_dict(orient="records")
        return Response(required_data)
    
class Download_Contact_report(APIView):
    def get(self, request):
        sub_company = request.query_params.get('sub_company')
        company = request.query_params.get('company')
        vendor = request.query_params.get('vendor')
        data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(sub_company=sub_company, company=company, vendor=vendor)
        serializer = baseallserializer(data, many=True)

        try:
            df = pd.DataFrame(serializer.data)
            df = df[["vendor", "Contract_Person", "Date_Signed", "Start_Date", "End_Date", "Term", 
                     "Contract", "accountnumber", "SVC_Type", "Auto_Renew", "Attachments"]]

            with BytesIO() as buffer:
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Contact Report', startrow=1)
                    workbook = writer.book
                    sheet = writer.sheets['Contact Report']
                    
                    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
                    title_cell = sheet.cell(row=1, column=1)
                    title_cell.value = "Contact Report"
                    title_cell.font = Font(size=18, bold=True)
                    title_cell.alignment = Alignment(horizontal="center", vertical="center")
                    title_cell.fill = PatternFill(start_color="A8CD89", end_color="A8CD89", fill_type="solid")
                    
                    column_fill = PatternFill(start_color="F4E0AF", end_color="F4E0AF", fill_type="solid")
                    for cell in sheet[2]:
                        cell.fill = column_fill
                        cell.font = Font(bold=True, color="000000")
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    for col_num, _ in enumerate(df.columns, 1):
                        column_width = max(len(df.columns[col_num - 1]) + 2, 15)
                        sheet.column_dimensions[get_column_letter(col_num)].width = column_width

                buffer.seek(0)
                response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="Contact_Report.xlsx"'
                return response

        except Exception as e:
            print(str(e))
            return Response("Data not found", status=404)
        
class Get_BAN_Autopay_Enabled_Listing_report(APIView):
    def get(self,request):
        sub_company = request.query_params.get('sub_company')
        company = request.query_params.get('company')
        data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(sub_company=sub_company,company=company)
        serializer = baseallserializer(data,many=True)
        df = pd.DataFrame(serializer.data)
        df = df[["accountnumber", "Location_Code", "location", "vendor"]]
        required_data = df.to_dict(orient="records")
        return Response(required_data)
    
class Download_BAN_Autopay_Enabled_Listing_report(APIView):
    def get(self, request):
        sub_company = request.query_params.get('sub_company')
        company = request.query_params.get('company')
        data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(sub_company=sub_company, company=company)
        serializer = baseallserializer(data, many=True)

        try:
            df = pd.DataFrame(serializer.data)
            df = df[["accountnumber", "Location_Code", "location", "vendor"]]

            with BytesIO() as buffer:
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Ban Autopay Enabled Report', startrow=1)
                    workbook = writer.book
                    sheet = writer.sheets['Ban Autopay Enabled Report']
                    
                    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
                    title_cell = sheet.cell(row=1, column=1)
                    title_cell.value = "BAN Autopay Enabled Listing Report"
                    title_cell.font = Font(size=18, bold=True)
                    title_cell.alignment = Alignment(horizontal="center", vertical="center")
                    title_cell.fill = PatternFill(start_color="A8CD89", end_color="A8CD89", fill_type="solid")
                    
                    column_fill = PatternFill(start_color="F4E0AF", end_color="F4E0AF", fill_type="solid")
                    for cell in sheet[2]:
                        cell.fill = column_fill
                        cell.font = Font(bold=True, color="000000")
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    for col_num, _ in enumerate(df.columns, 1):
                        column_width = max(len(df.columns[col_num - 1]) + 2, 15)
                        sheet.column_dimensions[get_column_letter(col_num)].width = column_width

                buffer.seek(0)
                response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="BAN_Autopay_Enabled_Listing_Report.xlsx"'
                return response

        except Exception as e:
            print(str(e))
            return Response("Data not found", status=404)
        
class Get_Phone_Numbers_List_By_Organization_report(APIView):
    def get(self,request):
        sub_company = request.query_params.get('sub_company')
        company = request.query_params.get('company')
        data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(sub_company=sub_company,company=company)
        serializer = baseallserializer(data,many=True)
        df = pd.DataFrame(serializer.data)
        df = df[["location","Services_Used","BTN_WTN","DID","vendor"]]
        required_data = df.to_dict(orient="records")
        return Response(required_data)
 

class Download_Phone_Numbers_List_By_Organization_report(APIView):
    def get(self, request):
        sub_company = request.query_params.get('sub_company')
        company = request.query_params.get('company')
        data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(sub_company=sub_company, company=company)
        serializer = baseallserializer(data, many=True)

        try:
            df = pd.DataFrame(serializer.data)
            df = df[["location", "Services_Used", "BTN_WTN", "DID", "vendor"]]

            with BytesIO() as buffer:
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Phone Numbers List By Organization Report', startrow=1)
                    workbook = writer.book
                    sheet = writer.sheets['Phone Numbers List By Organization Report']
                    
                    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
                    title_cell = sheet.cell(row=1, column=1)
                    title_cell.value = "Phone Numbers List By Organization Report"
                    title_cell.font = Font(size=18, bold=True)
                    title_cell.alignment = Alignment(horizontal="center", vertical="center")
                    title_cell.fill = PatternFill(start_color="A8CD89", end_color="A8CD89", fill_type="solid")
                    
                    column_fill = PatternFill(start_color="F4E0AF", end_color="F4E0AF", fill_type="solid")
                    for cell in sheet[2]:
                        cell.fill = column_fill
                        cell.font = Font(bold=True, color="000000")
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    for col_num, _ in enumerate(df.columns, 1):
                        column_width = max(len(df.columns[col_num - 1]) + 2, 15)
                        sheet.column_dimensions[get_column_letter(col_num)].width = column_width

                buffer.seek(0)
                response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="Phone_Numbers_List_By_Organization_Report.xlsx"'
                return response

        except Exception as e:
            print(str(e))
            return Response("Data not found", status=404)
        
class Approve_all_data(APIView):
    def put(self,request):
        try:
            date = request.data.get('bill_date')
            data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(bill_date=date)
            data.update(status="Approved")
            return Response("status saved sucessfully")
        except:
            return Response("No record found")
        

from django.http import JsonResponse,FileResponse,HttpResponse

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_vendor_summary_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    company = request.GET.get('company')

    print(start_date, end_date)

    # Validate required parameters
    if not start_date or not end_date or not sub_company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    # Parse date range
    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return JsonResponse({'message': 'Invalid date format (expected YYYY-MM-DD)'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d{1,2},?\s\d{4}$'),
        sub_company=sub_company,
        company=company
    )


    result_list = []

    for item in filtered_data:
        bill_date_parsed = parse_bill_date(item.bill_date)
        if not bill_date_parsed:
            continue

        if start_date_parsed <= bill_date_parsed <= end_date_parsed:
            result_list.append({
                'id': item.id,
                'bill_date': item.bill_date,
                'sub_company': item.sub_company,
                'Total_Current_Charges': item.Total_Current_Charges,
                'vendor': item.vendor
            })

    return JsonResponse(result_list, safe=False)


def vendor_summary_report(request):
    # Extract query parameters from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    company = request.GET.get('company')

    print(sub_company, end_date, start_date)

    if not start_date or not end_date or not sub_company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return JsonResponse({'message': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter by sub_company and ensure bill_date format is valid
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),sub_company=sub_company,company=company)

    print(filtered_data)
    result_list = []
    for item in filtered_data:
        print(item)
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                result_list.append({
                    'id': item.id,  
                    'bill_date': item.bill_date,
                    'sub_company': item.sub_company,
                    'Total_Current_Charges': item.Total_Current_Charges,
                    'vendor': item.vendor
                })
        except ValueError:
            # Skip entries that don't match the expected date format
            continue

    return JsonResponse(result_list, safe=False)
import os
from django.conf import settings

def vendor_summary_report_download(request):
    # Get query parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    company = request.GET.get('company')

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return HttpResponse("Invalid date format", status=status.HTTP_400_BAD_REQUEST)

    # Generate the full list of months within the date range
    all_months = generate_month_range(start_date, end_date)

    # Fetch and filter data from the database
    data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        company=company,
        sub_company=sub_company,
        bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'  # Ensure valid date format
    )

    filtered_data = []
    for item in data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                filtered_data.append(item)
        except ValueError:
            continue

    # Group data by vendor and month/year
    grouped_data = {}
    for item in filtered_data:
        formatted_date = datetime.strptime(item.bill_date, '%B %d %Y').strftime('%B %Y')
        if item.vendor not in grouped_data:
            grouped_data[item.vendor] = {}
        grouped_data[item.vendor][formatted_date] = float(item.Total_Current_Charges.replace(',', '').replace('$', ''))

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Vendor Summary Report'

    # Add main heading
    main_heading = 'Expense Management - Total Expense By Vendor'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_months) + 1)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Adjust column widths
    for col in range(1, len(all_months) + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # Add headers
    headers = ['Vendor'] + all_months
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='ADD8E6', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Fill rows
    for vendor, charges in grouped_data.items():
        row = [vendor]
        for date in all_months:
            value = charges.get(date, 0.00)
            row.append(value)
        ws.append(row)

    # Add total row
    total_row = ['Total']
    for col_index in range(2, len(all_months) + 2):
        sum_value = sum(
            cell.value for col in ws.iter_cols(min_col=col_index, max_col=col_index, min_row=3, max_row=ws.max_row)
            for cell in col if isinstance(cell.value, (int, float))
        )
        total_row.append(round(sum_value, 2))
    ws.append(total_row)

    # Style total row
    last_row_idx = ws.max_row
    for cell in ws[last_row_idx]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFFFE0', fill_type='solid')
        cell.alignment = Alignment(horizontal='right')

    # Add borders to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    # ==========================
    # 📁 SAVE REPORT LOCALLY
    # ==========================
    reports_dir = os.path.join(settings.BASE_DIR, 'reports')
    os.makedirs(reports_dir, exist_ok=True)

    filename = f"{sub_company.replace(' ', '_')}_Total_Expense_By_Vendor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = os.path.join(reports_dir, filename)

    wb.save(file_path)  # Save locally

    # ==========================
    # 📤 RETURN AS DOWNLOAD
    # ==========================
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def division_summary_report(request):
    # Extract query parameters from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    division = request.GET.get('division')

    print(sub_company, end_date, start_date)

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return JsonResponse({'message': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter by sub_company and ensure bill_date format is valid
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),sub_company=sub_company,division=division)


    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                result_list.append({
                    'id': item.id,  # Include relevant fields
                    'bill_date': item.bill_date,
                    'sub_company': item.sub_company,
                    'Total_Current_Charges': item.Total_Current_Charges,
                    'vendor': item.vendor,
                    'division': item.division  # Include relevant fields
                })
        except ValueError:
            # Skip entries that don't match the expected date format
            continue

    return JsonResponse(result_list, safe=False)





def division_summary_report_download(request):
    # Get query parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    division = request.GET.get('division')

    # Validate input parameters
    if not start_date or not end_date or not sub_company or not division:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return HttpResponse("Invalid date format", status=status.HTTP_400_BAD_REQUEST)

    # Fetch and filter data from the database
    data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        sub_company=sub_company,
        division=division,
        bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'  # Ensure valid date format
    )

    filtered_data = []
    for item in data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                filtered_data.append(item)
        except ValueError:
            continue

    # Group data by unique vendors and sum their Total_Current_Charges
    grouped_data = {}
    for item in filtered_data:
        vendor_key = (item.vendor, item.division)
        if vendor_key not in grouped_data:
            grouped_data[vendor_key] = {
                'vendor': item.vendor,
                'division': item.division,
                'Total_Current_Charges': float(item.Total_Current_Charges.replace(',', ''))
            }
        else:
            grouped_data[vendor_key]['Total_Current_Charges'] += float(item.Total_Current_Charges.replace(',', ''))

    # Convert the grouped data to a list
    grouped_data_list = list(grouped_data.values())

    # Calculate total sum for the Total_Current_Charges
    total_sum = sum(item['Total_Current_Charges'] for item in grouped_data_list)

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Division Summary Report'

    # Add main heading
    main_heading = 'T.E.M. - Division Summary'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)  # Adjusted to span more columns
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Set column width to ensure main heading fits well
    ws.column_dimensions['A'].width = 40  # Adjust the width as needed

    # Add table headers
    headers = ['Vendor', 'Division', 'Total Bill']
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='ADD8E6', fill_type='solid')  # Light blue background
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in grouped_data_list:
        ws.append([
            item['vendor'],
            item['division'],
            "{:,.2f}".format(item['Total_Current_Charges'])  # Format with commas and two decimal places
        ])

    # Add a total row at the end
    total_row = ws.max_row + 1
    ws[f'A{total_row}'] = 'Total'
    ws[f'A{total_row}'].font = Font(bold=True)
    ws[f'A{total_row}'].alignment = Alignment(horizontal='right')
    ws[f'A{total_row}'].border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000')
    )

    ws[f'C{total_row}'] = "{:,.2f}".format(total_sum)  # Format with commas and two decimal places
    ws[f'C{total_row}'].font = Font(bold=True)
    ws[f'C{total_row}'].alignment = Alignment(horizontal='right')
    ws[f'C{total_row}'].fill = PatternFill(start_color='FFFFE0', fill_type='solid')  # Light yellow background

    # Style all cells with borders
    for row in ws.iter_rows(min_row=1, max_row=total_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    # Create HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_T.E.M_Division_Report.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



def get_vendors_by_organization(request):
    # Get the 'organization_name' parameter from the GET request
    organization_name = request.GET.get('organization_name')

    # Check if the parameter is provided
    if not organization_name:
        return JsonResponse({'message': 'Missing required parameter: organization_name'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter the data from Organizations by organization_name and return as a simpler JSON response
    try:
        filtered_data = Organizations.objects.filter(organization_name=organization_name).values()

        # Check if there is data
        if not filtered_data:
            return JsonResponse({'message': 'No data found for the specified organization'}, status=404)

        # Return the filtered data as a simple JSON response
        return JsonResponse(list(filtered_data), safe=False)

    except Exception as e:
        return JsonResponse({'message': str(e)}, status=500)



def get_bans_by_organization_vendor(request):
    # Get the 'organization_name' parameter from the GET request
    organization_name = request.GET.get('organization_name')
    vendor = request.GET.get('vendor')

    # Check if the parameter is provided
    if not organization_name:
        return JsonResponse({'message': 'Missing required parameter: organization_name'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter the data from Organizations by organization_name and return as a simpler JSON response
    try:
        filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(sub_company=organization_name, vendor=vendor).values()

        # Check if there is data
        if not filtered_data:
            return JsonResponse({'message': 'No data found for the specified organization'}, status=404)

        # Return the filtered data as a simple JSON response
        return JsonResponse(list(filtered_data), safe=False)

    except Exception as e:
        return JsonResponse({'message': str(e)}, status=500)



def get_entrytype_bans_by_organization(request):
    # Get the 'organization_name' parameter from the GET request
    organization_name = request.GET.get('organization_name')
    company = request.GET.get('company')
    Entry_type = request.GET.get('Entry_type')

    # Check if the parameter is provided
    if not organization_name:
        return JsonResponse({'message': 'Missing required parameter: organization_name'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter the data from Organizations by organization_name and return as a simpler JSON response
    try:
        filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(sub_company=organization_name, company=company, Entry_type=Entry_type).values()

        # Check if there is data
        if not filtered_data:
            return JsonResponse({'message': 'No data found for the specified organization'}, status=404)

        # Return the filtered data as a simple JSON response
        return JsonResponse(list(filtered_data), safe=False)

    except Exception as e:
        return JsonResponse({'message': str(e)}, status=500)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def un_orapprovedbills_base_tem_report(request):
    # Extract query parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    vendor = request.GET.get('vendor')
    ban = request.GET.get('ban')
    company = request.GET.get('company')
    approval_status = request.GET.get('approval_status')
    Entry_type = request.GET.get('entry_type')

    # Validate required parameters
    if not start_date or not end_date or not sub_company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    # Parse date range
    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return JsonResponse({'message': 'Invalid date format (expected YYYY-MM-DD)'}, status=status.HTTP_400_BAD_REQUEST)

    # Common date parsing formats


    # Filter BaseDataTable
    print(Entry_type, vendor, sub_company, company)
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d{1,2},?\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        vendor=vendor,
        is_baseline_approved=approval_status,
        # Entry_type=Entry_type
    )
    if ban:
        filtered_data = filtered_data.filter(accountnumber=ban)


    valid_filtered_data = [
        item for item in filtered_data
        if (bill_date := parse_bill_date(item.bill_date))
        and start_date_parsed <= bill_date <= end_date_parsed
    ]
    # Convert to dict
    simple_response = [model_to_dict(item, fields=["accountnumber", "invoicenumber", "vendor", "location", "bill_date", "variance", "net_amount", "created_at", "bill_approved_date"]) for item in valid_filtered_data]
    print(simple_response)

    return JsonResponse(simple_response, safe=False)

from OnBoard.Location.ser import showDivisions, LocationShowSerializer
class GetDivisionView(APIView):
    def get(self, request):
        company_name = request.query_params.get('company_name')
        sub_company_name = request.query_params.get('sub_company_name')
        if not company_name or not sub_company_name:
            return Response({"error": "company_name and sub_company_name are required parameters"}, status=status.HTTP_400_BAD_REQUEST)
        
        divisions = showDivisions.objects.filter(company_name=company_name, sub_company_name=sub_company_name)
        if not divisions.exists():
            return Response({"error": "No records found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = showDivisions(divisions, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)




class GetLocationView(APIView):
    def get(self, request):
        company_name = request.query_params.get('company_name')
        sub_company_name = request.query_params.get('sub_company_name')
        if not company_name or not sub_company_name:
            return Response({"error": "company_name and sub_company_name are required parameters"}, status=status.HTTP_400_BAD_REQUEST)
        
        locations = Location.objects.filter(company_name=company_name, sub_company_name=sub_company_name)
        if not locations.exists():
            return Response({"error": "No records found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = LocationShowSerializer(locations, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

def approved_bills_base_tem_report_download(request):
    # Get query parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    vendor = request.GET.get('vendor')
    ban = request.GET.get('ban')
    company = request.GET.get('company')
    Entry_type = request.GET.get('entry_type')

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return HttpResponse("Invalid date format", status=status.HTTP_400_BAD_REQUEST)

    # Filter data from BaseDataTable
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        vendor=vendor,
        approval_status='True',
        Entry_type=Entry_type
    )
    if ban:
        filtered_data = filtered_data.filter(accountnumber=ban)

    # Filter by date range
    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                result_list.append(item)
        except ValueError:
            continue

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Approved Bills Report'

    # Main heading
    main_heading = 'Base T.E.M. - Approved Bills'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        "Account #", "Invoice #", "Vendor", "Site", "Address",
        "Bill Date", "BBA", "BBA Var", "Bill Amount", "Approved By",
        "Date Entered", "Date Approved"
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')  # Blue background
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Filter data from BaselineDataTable and calculate total from category_object
    baseline_filtered_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        vendor=vendor
    )
    if ban:
        baseline_filtered_data = baseline_filtered_data.filter(account_number=ban)

    # Create a mapping of bill_date to total sum from category_object
    bba_sum_map = {}
    for item in baseline_filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                date_key = bill_date.strftime('%B %d %Y')
                category_object = item.category_object or {}

                if date_key not in bba_sum_map:
                    bba_sum_map[date_key] = 0

                # Sum all values in the category_object for this bill_date
                if isinstance(category_object, dict):
                    for section, values in category_object.items():
                        if isinstance(values, dict):
                            bba_sum_map[date_key] += sum(
                                float(value) for value in values.values() if isinstance(value, (int, float))
                            )
        except (ValueError, TypeError):
            continue

    # Populate data rows and apply alignment
    for item in result_list:
        bill_date_str = item.bill_date
        bba_value = bba_sum_map.get(bill_date_str, 0)  # Get the BBA value for the bill date or 0 if not present
        row = [
            item.accountnumber,
            item.InvoiceNumber,
            item.vendor,
            item.location,
            item.Remidence_Address,
            item.bill_date,
            format(bba_value, ',.2f'),  # Format BBA value with comma and two decimal places
            "",  # Placeholder for BBA Var
            item.Total_Current_Charges,
            "",
            "",
            ""
        ]
        ws.append(row)
        last_row = ws.max_row
        ws[f'A{last_row}'].alignment = Alignment(horizontal='center')  # Align accountnumber
        ws[f'B{last_row}'].alignment = Alignment(horizontal='center')  # Align InvoiceNumber
        ws[f'I{last_row}'].alignment = Alignment(horizontal='right')   # Align Total_Current_Charges

    # Style the entire worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=12):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
            if isinstance(cell.value, (int, float)) and cell.column == 9:  # Check for Bill Amount column
                cell.alignment = Alignment(horizontal='right')

    # Total row at the bottom
    total_row = [""] * 7 + ["Total", sum(
        float(item.Total_Current_Charges.replace(',', '')) for item in result_list if item.Total_Current_Charges), "", ""]
    ws.append(total_row)
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFFFE0', fill_type='solid')  # Light yellow background
        cell.alignment = Alignment(horizontal='right')

    # Adjust column widths safely by skipping merged cells
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Get the column letter using openpyxl's utility function
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Check if cell is not a MergedCell
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_Approved_Bills_Base_Report.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response







def un_or_approvedbills_consoliated_tem_report(request):
    # Extract query parameters from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    ban = request.GET.get('ban')
    company = request.GET.get('company')
    approval_status = request.GET.get('approval_status')
    Entry_type = request.GET.get('Entry_type')

    print(sub_company, end_date, start_date, ban, company, approval_status, Entry_type)

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return JsonResponse({'message': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter by sub_company and ensure bill_date format is valid
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        is_baseline_approved=approval_status,
        # Entry_type=Entry_type,
    )
    if ban:
        filtered_data = filtered_data.filter(accountnumber=ban)


    valid_filtered_data = [
        item for item in filtered_data
        if (bill_date := parse_bill_date(item.bill_date))
        and start_date_parsed <= bill_date <= end_date_parsed
    ]

    simple_response = [model_to_dict(item, fields=["account_number", "invoicenumber", "vendor", "location", "bill_date", "variance", "net_amount", "created_at", "bill_approved_date"]) for item in valid_filtered_data]
    print(simple_response)

    # Return the updated response
    return JsonResponse(simple_response, safe=False)




def approved_bills_consoliated_tem_report_download(request):
    # Get query parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    ban = request.GET.get('ban')
    company = request.GET.get('company')
    approval_status = request.GET.get('approval_status')
    Entry_type = request.GET.get('Entry_type')

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return HttpResponse("Invalid date format", status=status.HTTP_400_BAD_REQUEST)

    # Filter data from BaseDataTable
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        approval_status=approval_status,
        Entry_type=Entry_type,
    )
    if ban:
        filtered_data = filtered_data.filter(accountnumber=ban)

    # Filter by date range
    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                result_list.append(item)
        except ValueError:
            continue

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Approved Bills Report'

    # Main heading
    main_heading = 'Consoliated T.E.M. - Approved Bills'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        "Account #", "Invoice #", "Vendor", "Site", "Address",
        "Bill Date", "BBA", "BBA Var", "Bill Amount", "Approved By",
        "Date Entered", "Date Approved"
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')  # Blue background
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows and include BBA sum from BaselineDataTable
    for item in result_list:
        bill_date_str = item.bill_date
        bba_sum = 0

        # Filter BaselineDataTable for the current bill_date
        baseline_filtered_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
            Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
            sub_company=sub_company,
            company=company,
            account_number=item.accountnumber
        )
        for baseline_item in baseline_filtered_data:
            try:
                baseline_bill_date = datetime.strptime(baseline_item.bill_date, '%B %d %Y')
                if baseline_bill_date == datetime.strptime(bill_date_str, '%B %d %Y'):
                    category_object = baseline_item.category_object or {}
                    if isinstance(category_object, dict):
                        for section, values in category_object.items():
                            if isinstance(values, dict):
                                bba_sum += sum(
                                    float(value) for value in values.values() if isinstance(value, (int, float))
                                )
            except (ValueError, TypeError):
                continue

        # Populate the row with data
        row = [
            item.accountnumber,
            item.InvoiceNumber,
            item.vendor,
            item.location,
            item.Remidence_Address,
            item.bill_date,
            bba_sum,  # Populate BBA with the calculated sum
            "",  # Placeholder for BBA Var
            item.Total_Current_Charges,
            "",
            "",
            ""
        ]
        ws.append(row)
        last_row = ws.max_row
        ws[f'A{last_row}'].alignment = Alignment(horizontal='center')  # Align accountnumber
        ws[f'B{last_row}'].alignment = Alignment(horizontal='center')  # Align InvoiceNumber
        ws[f'I{last_row}'].alignment = Alignment(horizontal='right')   # Align Total_Current_Charges
        ws[f'G{last_row}'].alignment = Alignment(horizontal='right')   # Align BBA

    # Style the entire worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=12):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
            if isinstance(cell.value, (int, float)) and cell.column in [7, 9]:  # Check for BBA and Bill Amount columns
                cell.alignment = Alignment(horizontal='right')

    # Total row at the bottom
    total_row = [""] * 7 + ["Total", sum(
        float(item.Total_Current_Charges.replace(',', '')) for item in result_list if item.Total_Current_Charges), "", ""]
    ws.append(total_row)
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFFFE0', fill_type='solid')  # Light yellow background
        cell.alignment = Alignment(horizontal='right')

    # Adjust column widths safely by skipping merged cells
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_Approved_Bills_Consoliated_Report.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response





def unapprovedbills_base_tem_report(request):
    # Extract query parameters from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    vendor = request.GET.get('vendor')
    ban = request.GET.get('ban')
    company = request.GET.get('company')
    approval_status = request.GET.get('approval_status')
    Entry_type = request.GET.get('entry_type')

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return JsonResponse({'message': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter by sub_company and ensure bill_date format is valid
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        vendor=vendor,
        approval_status=approval_status,
        Entry_type=Entry_type
    )
    if ban:
        filtered_data = filtered_data.filter(accountnumber=ban)

    # Filter the data by the date range
    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                result_list.append(item)
        except ValueError:
            continue

    # Convert filtered_data to a list of dictionaries
    simple_response = list(filtered_data.values())

    # Filter data from BaselineDataTable and calculate total from category_object
    baseline_filtered_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        vendor=vendor
    )
    if ban:
        baseline_filtered_data = baseline_filtered_data.filter(account_number=ban)

    # Calculate total sum for each date and include it in the response
    total_sums_by_date = {}
    for item in baseline_filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                category_object = item.category_object or {}
                date_key = bill_date.strftime('%B %d %Y')
                total_sum = sum(
                    float(value) for section in category_object.values()
                    if isinstance(section, dict)
                    for value in section.values() if isinstance(value, (int, float))
                )
                total_sums_by_date[date_key] = total_sums_by_date.get(date_key, 0) + total_sum
        except (ValueError, TypeError, json.JSONDecodeError):
            continue

    # Add the calculated totals to the response
    for entry in simple_response:
        entry_date = entry['bill_date']
        entry['Total_Category_Sum'] = total_sums_by_date.get(entry_date, 0)

    # Return the updated response with the sum data
    return JsonResponse(simple_response, safe=False)








def format_number(number):
    """Format a number with commas and two decimal places."""
    return f"{number:,.2f}"

def unapproved_bills_base_tem_report_download(request):
    # Get query parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    vendor = request.GET.get('vendor')
    ban = request.GET.get('ban')
    company = request.GET.get('company')
    Entry_type = request.GET.get('entry_type')

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return HttpResponse("Invalid date format", status=status.HTTP_400_BAD_REQUEST)

    # Filter data from the BaseDataTable
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        vendor=vendor,
        approval_status='False',
        Entry_type=Entry_type
    )
    if ban:
        filtered_data = filtered_data.filter(accountnumber=ban)

    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                result_list.append(item)
        except ValueError:
            continue

    # Filter data from BaselineDataTable and calculate totals from category_object
    baseline_filtered_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        vendor=vendor
    )
    if ban:
        baseline_filtered_data = baseline_filtered_data.filter(account_number=ban)

    total_sums_by_date = {}
    for item in baseline_filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                category_object = json.loads(item.category_object) if item.category_object else {}
                date_key = bill_date.strftime('%B %d %Y')
                total_sum = sum(
                    float(value) for section in category_object.values()
                    if isinstance(section, dict)
                    for value in section.values() if isinstance(value, (int, float))
                )
                total_sums_by_date[date_key] = total_sums_by_date.get(date_key, 0) + total_sum
        except (ValueError, TypeError, json.JSONDecodeError):
            continue

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Unapproved Bills Report'

    # Main heading
    main_heading = 'Base T.E.M. - Unapproved Bills'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        "Account #", "Invoice #", "Vendor", "Site", "Address",
        "Bill Date", "BBA", "BBA Var", "Bill Amount", "Approved By",
        "Date Entered", "Date Approved"
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in result_list:
        bba_value = total_sums_by_date.get(item.bill_date, 0)  # Get the BBA value for the current Bill Date
        row = [
            item.accountnumber,
            item.InvoiceNumber,
            item.vendor,
            item.location,
            item.Remidence_Address,
            item.bill_date,
            f"{bba_value:.2f}",  # Format BBA value
            "",  # Placeholder for BBA Var
            item.Total_Current_Charges,
            "",
            "",
            ""
        ]
        ws.append(row)
        last_row = ws.max_row
        ws[f'A{last_row}'].alignment = Alignment(horizontal='center')  # Align accountnumber
        ws[f'B{last_row}'].alignment = Alignment(horizontal='center')  # Align InvoiceNumber
        ws[f'I{last_row}'].alignment = Alignment(horizontal='right')   # Align Total_Current_Charges

    # Style the entire worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=12):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
            if isinstance(cell.value, (int, float)) and cell.column == 9:  # Check for Bill Amount column
                cell.alignment = Alignment(horizontal='right')

    # Total row at the bottom
    total_row = [""] * 7 + ["Total", sum(
        float(item.Total_Current_Charges.replace(',', '')) for item in result_list if item.Total_Current_Charges), "", ""]
    ws.append(total_row)
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFFFE0', fill_type='solid')
        cell.alignment = Alignment(horizontal='right')

    # Adjust column widths safely by skipping merged cells
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_Unapproved_Bills_Base_Report.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response









def unapprovedbills_consoliated_tem_report(request):
    # Extract query parameters from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    ban = request.GET.get('ban')
    company = request.GET.get('company')
    approval_status = request.GET.get('approval_status')
    Entry_type = request.GET.get('Entry_type')

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return JsonResponse({'message': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter by sub_company and ensure bill_date format is valid
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        approval_status=approval_status,
        Entry_type=Entry_type
    )
    if ban:
        filtered_data = filtered_data.filter(accountnumber=ban)

    # Filter the data by the date range
    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                result_list.append(item)
        except ValueError:
            continue

    # Convert filtered_data to a list of dictionaries
    simple_response = list(filtered_data.values())

    # Filter data from BaselineDataTable and calculate total from category_object
    baseline_filtered_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company
    )
    if ban:
        baseline_filtered_data = baseline_filtered_data.filter(account_number=ban)

    # Calculate total sum for each date and include it in the response
    total_sums_by_date = {}
    for item in baseline_filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                category_object = json.loads(item.category_object) if item.category_object else {}
                date_key = bill_date.strftime('%B %d %Y')
                total_sum = sum(
                    float(value) for section in category_object.values()
                    if isinstance(section, dict)
                    for value in section.values() if isinstance(value, (int, float))
                )
                total_sums_by_date[date_key] = total_sums_by_date.get(date_key, 0) + total_sum
        except (ValueError, TypeError, json.JSONDecodeError):
            continue

    # Add the calculated totals to the response
    for entry in simple_response:
        entry_date = entry['bill_date']
        entry['Total_Category_Sum'] = total_sums_by_date.get(entry_date, 0)

    # Return the updated response with the sum data
    return JsonResponse(simple_response, safe=False)





def unapproved_bills_consoliated_tem_report_download(request):
    # Get query parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    ban = request.GET.get('ban')
    company = request.GET.get('company')
    approval_status = request.GET.get('approval_status')
    Entry_type = request.GET.get('Entry_type')

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return HttpResponse("Invalid date format", status=status.HTTP_400_BAD_REQUEST)

    # Filter data from BaseDataTable
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        approval_status=approval_status,
        Entry_type=Entry_type,
    )
    if ban:
        filtered_data = filtered_data.filter(accountnumber=ban)

    # Filter by date range and prepare result list
    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                result_list.append(item)
        except ValueError:
            continue

    # Filter data from BaselineDataTable and calculate total from category_object
    baseline_filtered_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company
    )
    if ban:
        baseline_filtered_data = baseline_filtered_data.filter(account_number=ban)

    category_totals = {}
    for item in baseline_filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                category_object = json.loads(item.category_object) if item.category_object else {}
                total_sum = sum(
                    float(value) for section in category_object.values()
                    if isinstance(section, dict)
                    for value in section.values() if isinstance(value, (int, float))
                )
                date_key = bill_date.strftime('%B %d %Y')
                category_totals[date_key] = category_totals.get(date_key, 0) + total_sum
        except (ValueError, TypeError, json.JSONDecodeError):
            continue

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Unapproved Bills Report'

    # Main heading
    main_heading = 'Consolidated T.E.M. - Unapproved Bills'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        "Account #", "Invoice #", "Vendor", "Site", "Address",
        "Bill Date", "BBA", "BBA Var", "Bill Amount", "Approved By",
        "Date Entered", "Date Approved"
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')  # Blue background
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows and apply alignment
    for item in result_list:
        bba_value = category_totals.get(item.bill_date, 0)  # Get the total sum for the BBA field
        row = [
            item.accountnumber,
            item.InvoiceNumber,
            item.vendor,
            item.location,
            item.Remidence_Address,
            item.bill_date,
            format(bba_value, '.2f'),  # Format BBA value
            "",  # Placeholder for BBA Var
            item.Total_Current_Charges,
            "",
            "",
            ""
        ]
        ws.append(row)
        last_row = ws.max_row
        ws[f'A{last_row}'].alignment = Alignment(horizontal='center')  # Align accountnumber
        ws[f'B{last_row}'].alignment = Alignment(horizontal='center')  # Align InvoiceNumber
        ws[f'I{last_row}'].alignment = Alignment(horizontal='right')   # Align Total_Current_Charges

    # Style the entire worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=12):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
            if isinstance(cell.value, (int, float)) and cell.column == 9:  # Check for Bill Amount column
                cell.alignment = Alignment(horizontal='right')

    # Total row at the bottom
    total_row = [""] * 7 + ["Total", sum(
        float(item.Total_Current_Charges.replace(',', '')) for item in result_list if item.Total_Current_Charges), "", ""]
    ws.append(total_row)
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFFFE0', fill_type='solid')  # Light yellow background
        cell.alignment = Alignment(horizontal='right')

    # Adjust column widths safely by skipping merged cells
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Get the column letter using openpyxl's utility function
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Check if cell is not a MergedCell
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_Unapproved_Bills_Consolidated_Report.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response







def highest_expense_bills_report(request):
    # Extract query parameters from the request
    company = request.GET.get('company')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')

    print(sub_company, end_date, start_date)

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return JsonResponse({'message': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter by sub_company and ensure bill_date format is valid
    
    filtered_data = (
        BaseDataTable.objects
        .exclude(viewuploaded=None, viewpapered=None)
        .filter(
            Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
            sub_company=sub_company,
            company=company
        )
        .order_by('-net_amount')[:25]  # ⬅️ top 25 by highest net_amount
    )
    print(filtered_data)
    result_list = [model_to_dict(item, fields=["account_number", "invoicenumber", "vendor", "location", "bill_date", "net_amount", "created_at", "bill_approved_date"]) for item in filtered_data]
    # for item in filtered_data:
    #     try:
    #         bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
    #         if start_date_parsed <= bill_date <= end_date_parsed:
    #             # Convert the entire object to a dictionary
    #             item_dict = model_to_dict(item)
                
    #             # Ensure Total_Current_Charges is not None before replacing
    #             if item.Total_Current_Charges is not None:
    #                 item_dict['Total_Current_Charges'] = float(item.Total_Current_Charges.replace(',', ''))
    #             else:
    #                 item_dict['Total_Current_Charges'] = 0  # Set to 0 if None

    #             result_list.append(item_dict)
    #     except ValueError:
    #         # Skip entries that don't match the expected date format
    #         continue

    # # Sort the result list by Total_Current_Charges in descending order and get the top 25
    # result_list = sorted(result_list, key=lambda x: x['Total_Current_Charges'], reverse=True)[:25]

    # # Filter data from BaselineDataTable and calculate the total sum for each date
    # baseline_filtered_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
    #     Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
    #     sub_company=sub_company
    # )

    # # Calculate the total sum of category_object per bill_date
    # date_sum_map = {}
    # for item in baseline_filtered_data:
    #     try:
    #         bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
    #         if start_date_parsed <= bill_date <= end_date_parsed:
    #             date_key = bill_date.strftime('%B %d %Y')
    #             category_object = item.category_object or {}

    #             if category_object and isinstance(category_object, dict):
    #                 if date_key not in date_sum_map:
    #                     date_sum_map[date_key] = 0  # Initialize sum for this date

    #                 # Sum all values in the category_object for this bill_date
    #                 for section, values in category_object.items():
    #                     if isinstance(values, dict):  # Ensure section contains a dictionary
    #                         section_sum = sum(
    #                             float(value) for value in values.values() if isinstance(value, (int, float))
    #                         )
    #                         date_sum_map[date_key] += section_sum

    #                         # Debugging: Print section and section sum
    #                         print(f"Bill Date: {date_key}, Section: {section}, Section Sum: {section_sum}")
    #             else:
    #                 # If category_object is empty or not a dict, ensure it adds 0
    #                 if date_key not in date_sum_map:
    #                     date_sum_map[date_key] = 0
    #     except (ValueError, TypeError) as e:
    #         # Handle parsing errors
    #         print("Error parsing item:", e)
    #         continue

    # # Add the total sum per date to the result_list
    # for entry in result_list:
    #     bill_date_str = entry['bill_date']
    #     entry['Total_Category_Sum'] = date_sum_map.get(bill_date_str, 0)

    # Return the updated response with the total category sum
    return JsonResponse(result_list, safe=False)






def highest_expense_bills_report_download(request):
    # Get query parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    company = request.GET.get('company')

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return HttpResponse("Invalid date format", status=status.HTTP_400_BAD_REQUEST)

    # Filter data from the BaseDataTable
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company
    )

    # Filter by date range and get top 25 highest Total_Current_Charges
    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                result_list.append(item)
        except ValueError:
            continue

    # Sort the result list by Total_Current_Charges and get the top 25
    result_list.sort(key=lambda x: float(x.Total_Current_Charges.replace(',', '') if x.Total_Current_Charges else 0), reverse=True)
    top_25_data = result_list[:25]

    # Filter data from the BaselineDataTable for BBA values
    baseline_filtered_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company
    )

    # Calculate the total sum of category_object per bill_date
    date_sum_map = {}
    for item in baseline_filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                date_key = bill_date.strftime('%B %d %Y')
                category_object = item.category_object or {}

                if category_object and isinstance(category_object, dict):
                    if date_key not in date_sum_map:
                        date_sum_map[date_key] = 0  # Initialize sum for this date

                    # Sum all values in the category_object for this bill_date
                    for section, values in category_object.items():
                        if isinstance(values, dict):  # Ensure section contains a dictionary
                            section_sum = sum(
                                float(value) for value in values.values() if isinstance(value, (int, float))
                            )
                            date_sum_map[date_key] += section_sum
        except (ValueError, TypeError, json.JSONDecodeError):
            continue

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Top 25 Highest Expense Bills'

    # Main heading
    main_heading = 'T.E.M. - Top 25 Highest Expense Bills'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        "Account #", "Invoice #", "Vendor", "Site", "Address",
        "Bill Date", "BBA", "BBA Var", "Bill Amount", "Approved By",
        "Date Entered", "Date Approved"
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in top_25_data:
        bba_value = date_sum_map.get(item.bill_date, 0)  # Get the BBA value for the current Bill Date
        row = [
            item.accountnumber,
            item.InvoiceNumber,
            item.vendor,
            item.location,
            item.Remidence_Address,
            item.bill_date,
            f"{bba_value:.2f}",  # Format BBA value
            "",  # Placeholder for BBA Var
            item.Total_Current_Charges,
            "",
            
        ]
        ws.append(row)
        last_row = ws.max_row
        ws[f'A{last_row}'].alignment = Alignment(horizontal='center')  # Align accountnumber
        ws[f'B{last_row}'].alignment = Alignment(horizontal='center')  # Align InvoiceNumber
        ws[f'I{last_row}'].alignment = Alignment(horizontal='right')   # Align Total_Current_Charges

    # Style the entire worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=12):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
            if isinstance(cell.value, (int, float)) and cell.column == 9:  # Check for Bill Amount column
                cell.alignment = Alignment(horizontal='right')

    # Adjust column widths safely by skipping merged cells
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_Highest_Expense_Bills_Report.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response








class GetDivisionView(APIView):
    def get(self, request):
        company_name = request.query_params.get('company_name')
        sub_company_name = request.query_params.get('sub_company_name')
        if not company_name or not sub_company_name:
            return Response({"error": "company_name and sub_company_name are required parameters"}, status=status.HTTP_400_BAD_REQUEST)
        
        divisions = Division.objects.filter(company_name=company_name, sub_company_name=sub_company_name)
        if not divisions.exists():
            return Response({"error": "No records found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = showDivisions(divisions, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)




class GetLocationView(APIView):
    def get(self, request):
        company_name = request.query_params.get('company_name')
        sub_company_name = request.query_params.get('sub_company_name')
        if not company_name or not sub_company_name:
            return Response({"error": "company_name and sub_company_name are required parameters"}, status=status.HTTP_400_BAD_REQUEST)
        
        locations = Location.objects.filter(company_name=company_name, sub_company_name=sub_company_name)
        if not locations.exists():
            return Response({"error": "No records found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = LocationShowSerializer(locations, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)



from django.forms.models import model_to_dict

def location_level_bill_summary_report(request):
    # Extract query parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    location = request.GET.get('location')
    company = request.GET.get('company')

    # Validate required params
    if not (start_date and end_date and sub_company):
        return JsonResponse({'message': 'Missing required parameters'}, status=400)

    # Parse date range
    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return JsonResponse({'message': 'Invalid date format (expected YYYY-MM-DD)'}, status=400)

    # Filter dataset
    filtered_data = (
        BaseDataTable.objects
        .exclude(viewuploaded=None, viewpapered=None)
        .filter(
            Q(bill_date__regex=r'^[A-Za-z]+\s\d{1,2},?\s\d{4}$'),
            sub_company=sub_company,
            location=location,
            company=company
        )
    )

    # Parse and filter by date range
    result_list = []
    for item in filtered_data:
        bill_date_parsed = parse_bill_date(item.bill_date)
        if is_within_range(start=start_date_parsed,raw_date=bill_date_parsed,end=end_date_parsed):
            item_dict = model_to_dict(item, fields=[
                'id', 'bill_date', 'location', 'sub_company', 'vendor', 
                'Total_Current_Charges', 'net_amount'
            ])
            result_list.append(item_dict)

    print(result_list)

    return JsonResponse(result_list, safe=False)




def location_level_bill_summary_report_download(request):
    # Extract query parameters from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    location = request.GET.get('location')
    company = request.GET.get('company')

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return HttpResponse("Invalid date format", status=status.HTTP_400_BAD_REQUEST)

    # Filter data from BaseDataTable
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        location=location
    )

    # Filter by date range
    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                result_list.append(item)
        except ValueError:
            continue

    # Create a new workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Location Level Bill Summary'

    # Add main heading
    main_heading = 'T.E.M. - Location Level Bill Summary'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = ["Location", "Vendor", "Ban #", "Bill Date", "Bill Amount"]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in result_list:
        row = [
            item.location,
            item.vendor,
            item.accountnumber,
            item.bill_date,
            item.Total_Current_Charges
        ]
        ws.append(row)

    # Style data rows
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length

    # Create HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_Location_Bill_Summary_Report.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response






def location_level_service_summary_report(request):
    # Extract query parameters from the request
    sub_company = request.GET.get('sub_company')
    location = request.GET.get('location')
    company = request.GET.get('company')

    # Validate input parameters
    if not sub_company or not location or not company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter by sub_company, company, and location and ensure bill_date format is valid
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
       
        sub_company=sub_company,
        company=company,
        location=location
    )

    # Convert filtered_data to a list of dictionaries
    result_list = list(filtered_data.values())

    # Return the filtered data as JSON
    return JsonResponse(result_list, safe=False)




def location_level_service_summary_report_download(request):
    # Extract query parameters from the request
    sub_company = request.GET.get('sub_company')
    location = request.GET.get('location')
    company = request.GET.get('company')

    # Validate input parameters
    if not sub_company or not location or not company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    # Filter data from BaseDataTable
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        sub_company=sub_company,
        company=company,
        location=location
    )

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Location Level Service Summary'

    # Main heading
    main_heading = 'Location Level Service Summary Report'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = ["Location", "Vendor", "Account #", "Service Type", "MRC"]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows and apply alignment
    for item in filtered_data:
        try:
            mrc_value = float(item.Total_Current_Charges.replace(',', '')) if item.Total_Current_Charges else 0
        except (ValueError, AttributeError):
            mrc_value = 0

        row = [
            item.location,
            item.vendor,
            item.accountnumber,
            item.services,
            f"{mrc_value:.2f}"
        ]
        ws.append(row)

        # Align data for each cell in the row
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )

    # Adjust column widths to fit content
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Get the column letter using openpyxl's utility function
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Skip merged cells
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{sub_company}_Location_Service_Summary.xlsx"'
    wb.save(response)
    return response



def baseline_billing_report(request):
    # Extract query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')

    # Validate required parameters
    if not company or not sub_company:
        return JsonResponse({'message': 'Missing required parameters: company and/or sub_company'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter data from BaselineDataTable based on company and sub_company
    filtered_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(company=company, sub_company=sub_company)

    # Convert the queryset to a list of dictionaries
    data_list = list(filtered_data.values())

    # Calculate and add the total sum of all numerical values in category_objects to each item
    records = []
    for item in data_list:
        item_instance = filtered_data.get(id=item['id'])
        total_charge = item_instance.total_charges

        # Clean "$" and convert to float
        if isinstance(total_charge, str):
            total_charge = total_charge.replace('$', '').replace(',', '').strip()

        try:
            total_charge = float(total_charge)
        except (TypeError, ValueError):
            total_charge = 0.0

        item['total_charges'] = total_charge
        records.append(item)

    # Convert to DataFrame
    df = pd.DataFrame(records)

    # Calculate overall total
    overall_total_sum = df['total_charges'].sum()
    # Return the data list and overall total sum as JSON response
    response = {
        'data_list': data_list,
        'total_sum': overall_total_sum
    }

    return JsonResponse(response, safe=False)






def baseline_billing_report_download(request):
    # Extract query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')

    # Validate required parameters
    if not company or not sub_company:
        return HttpResponse("Missing required parameters: company and/or sub_company", status=status.HTTP_400_BAD_REQUEST)

    # Filter data from BaselineDataTable based on company and sub_company
    filtered_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(company=company, sub_company=sub_company)

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Baseline Billing Report'

    # Main heading
    main_heading = 'Baseline Billing Report'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = ["Location", "Vendor", "Account Number", "Bill Date", "Total Category Sum"]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows and calculate total sums
    overall_total_sum = 0
    for item in filtered_data:
        category_object = item.category_object or {}
        total_sum = 0
        if isinstance(category_object, dict):
            for section, values in category_object.items():
                if isinstance(values, dict):
                    total_sum += sum(
                        float(value) for value in values.values() if isinstance(value, (int, float))
                    )
        overall_total_sum += total_sum

        row = [
            item.location,
            item.vendor,
            item.account_number,
            item.bill_date,
            f"{total_sum:.2f}"
        ]
        ws.append(row)

        # Style each row
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )

    # Total row
    total_row = ["", "", "", "Total Sum", f"{overall_total_sum:.2f}"]
    ws.append(total_row)

    # Style total row
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFFFE0', fill_type='solid')
        cell.alignment = Alignment(horizontal='right')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Adjust column widths
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0) + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length

    # Create HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{sub_company}_Baseline_Billing_Report.xlsx"'
    wb.save(response)
    return response





def invoice_tracking_report(request):
    # Extract query parameters from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    billing_day_start = request.GET.get('billing_day_start')
    billing_day_end = request.GET.get('billing_day_end')
    company = request.GET.get('company')

    # Validate input parameters
    if not start_date or not end_date or not sub_company or not billing_day_start or not billing_day_end:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
        billing_day_start = int(billing_day_start)
        billing_day_end = int(billing_day_end)
    except (ValueError, TypeError):
        return JsonResponse({'message': 'Invalid date format or billing day format'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter by sub_company and ensure bill_date format is valid
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        billing_day__gte=billing_day_start,
        billing_day__lte=billing_day_end
    )

    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                # Convert the entire object to a dictionary
                item_dict = model_to_dict(item)
                result_list.append(item_dict)
        except ValueError:
            # Skip entries that don't match the expected date format
            continue

    # Calculate the total category sum from BaselineDataTable
    baseline_filtered_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        company=company,
        sub_company=sub_company,
        bill_date__range=[start_date_parsed, end_date_parsed]
    )

    date_sum_map = {}
    for item in baseline_filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            date_key = bill_date.strftime('%B %d %Y')
            category_object = item.category_object or {}
            if category_object and isinstance(category_object, dict):
                if date_key not in date_sum_map:
                    date_sum_map[date_key] = 0

                # Sum all nested values in category_object
                section_sum = sum(
                    float(value) for section, values in category_object.items()
                    if isinstance(values, dict)
                    for value in values.values() if isinstance(value, (int, float))
                )
                date_sum_map[date_key] += section_sum
        except (ValueError, TypeError):
            continue

    # Add the total category sum to each entry in result_list
    for entry in result_list:
        bill_date_str = entry['bill_date']
        entry['Total_Category_Sum'] = date_sum_map.get(bill_date_str, 0)

    # Prepare the response data
    response_data = {
        'data_list': result_list,
        'total_category_sum': sum(date_sum_map.values())
    }

    return JsonResponse(response_data, safe=False)






def invoice_tracking_report_download(request):
    # Extract query parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    billing_day_start = request.GET.get('billing_day_start')
    billing_day_end = request.GET.get('billing_day_end')
    company = request.GET.get('company')

    # Validate input parameters
    if not start_date or not end_date or not sub_company or not billing_day_start or not billing_day_end:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
        billing_day_start = int(billing_day_start)
        billing_day_end = int(billing_day_end)
    except (ValueError, TypeError):
        return HttpResponse("Invalid date or billing day format", status=status.HTTP_400_BAD_REQUEST)

    # Filter data from BaseDataTable
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        billing_day__gte=billing_day_start,
        billing_day__lte=billing_day_end
    )

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Invoice Tracking Report'

    # Main heading
    main_heading = 'Invoice Tracking Report'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = ["Location", "Address", "Vendor", "Ban #", "Invoice Date", "Due Date", "Bill Date", "BBA Total", "Bill Amount"]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                # Calculate total category sum from BaselineDataTable
                baseline_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
                    Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
                    company=company,
                    sub_company=sub_company,
                    bill_date=bill_date
                )

                total_category_sum = 0
                for baseline_item in baseline_data:
                    category_object = baseline_item.category_object or {}
                    if isinstance(category_object, dict):
                        for section_values in category_object.values():
                            if isinstance(section_values, dict):
                                total_category_sum += sum(
                                    float(value) for value in section_values.values() if isinstance(value, (int, float))
                                )

                row = [
                    item.location,
                    item.Remidence_Address,
                    item.vendor,
                    item.accountnumber,
                    "",  # Ensure this column stays as an empty string
                    item.Date_Due,
                    item.bill_date,
                    f"{total_category_sum:.2f}",
                    f"{float(item.Total_Current_Charges.replace(',', '')):.2f}" if item.Total_Current_Charges else "0.00"
                ]

                ws.append(row)
        except (ValueError, AttributeError):
            continue

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{sub_company}_Invoice_Tracking_Report.xlsx"'
    wb.save(response)
    return response






def out_of_variance_report(request):
    # Extract query parameters from the request
    company = request.GET.get('company')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return JsonResponse({'message': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter by sub_company and ensure bill_date format is valid
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company
    )

    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                # Convert the entire object to a dictionary
                item_dict = model_to_dict(item)

                # Ensure Total_Current_Charges is not None before replacing
                if item.Total_Current_Charges is not None:
                    item_dict['Total_Current_Charges'] = float(item.Total_Current_Charges.replace(',', ''))
                else:
                    item_dict['Total_Current_Charges'] = 0  # Set to 0 if None

                result_list.append(item_dict)
        except ValueError:
            continue

    # Filter data from BaselineDataTable and calculate the total sum for each date
    baseline_filtered_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company
    )

    # Calculate the total sum of category_object per bill_date
    date_sum_map = {}
    variance_map = {}
    for item in baseline_filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                date_key = bill_date.strftime('%B %d %Y')
                category_object = item.category_object or {}
                variance = item.variance or 0  # Get the variance value

                if category_object and isinstance(category_object, dict):
                    if date_key not in date_sum_map:
                        date_sum_map[date_key] = 0  # Initialize sum for this date

                    # Sum all values in the category_object for this bill_date
                    for section, values in category_object.items():
                        if isinstance(values, dict):  # Ensure section contains a dictionary
                            section_sum = sum(
                                float(value) for value in values.values() if isinstance(value, (int, float))
                            )
                            date_sum_map[date_key] += section_sum

                    # Store the variance for this date
                    variance_map[date_key] = +variance

        except (ValueError, TypeError) as e:
            print("Error parsing item:", e)
            continue

    # Filter out data that does not meet the variance condition
    final_result_list = []
    for entry in result_list:
        bill_date_str = entry['bill_date']
        total_category_sum = date_sum_map.get(bill_date_str, 0)
        variance = variance_map.get(bill_date_str, 0)

        # Calculate the lower and upper bounds for the variance
        lower_bound = total_category_sum * (1 - (variance / 100))
        upper_bound = total_category_sum * (1 + (variance / 100))

        # Check if the Total_Current_Charges is within the variance range
        if not (lower_bound <= entry['Total_Current_Charges'] <= upper_bound):
            entry['Total_Category_Sum'] = total_category_sum
            final_result_list.append(entry)

    # Return the updated response with the filtered data
    return JsonResponse(final_result_list, safe=False)





def out_of_variance_report_download(request):
    # Extract query parameters
    company = request.GET.get('company')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return HttpResponse("Invalid date format", status=status.HTTP_400_BAD_REQUEST)

    # Filter data from BaseDataTable
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company
    )

    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                item_dict = model_to_dict(item)

                # Ensure Total_Current_Charges is not None before replacing
                if item.Total_Current_Charges is not None:
                    item_dict['Total_Current_Charges'] = float(item.Total_Current_Charges.replace(',', ''))
                else:
                    item_dict['Total_Current_Charges'] = 0

                result_list.append(item_dict)
        except ValueError:
            continue

    # Filter data from BaselineDataTable
    baseline_filtered_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company
    )

    date_sum_map = {}
    variance_map = {}
    for item in baseline_filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                date_key = bill_date.strftime('%B %d %Y')
                category_object = item.category_object or {}
                variance = item.variance or 0

                if category_object and isinstance(category_object, dict):
                    if date_key not in date_sum_map:
                        date_sum_map[date_key] = 0

                    for section, values in category_object.items():
                        if isinstance(values, dict):
                            section_sum = sum(
                                float(value) for value in values.values() if isinstance(value, (int, float))
                            )
                            date_sum_map[date_key] += section_sum

                    variance_map[date_key] = +variance

        except (ValueError, TypeError):
            continue

    # Prepare data to be written to Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Highest Expense Bills Report'

    # Main heading
    main_heading = 'Highest Expense Bills Report'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = ["Location", "Address", "Vendor", "Ban #", "Invoice Date", "Due Date", "Bill Date", "BBA Total", "Bill Amount"]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for entry in result_list:
        bill_date_str = entry['bill_date']
        total_category_sum = date_sum_map.get(bill_date_str, 0)
        variance = variance_map.get(bill_date_str, 0)
        lower_bound = total_category_sum * (1 - (variance / 100))
        upper_bound = total_category_sum * (1 + (variance / 100))

        if not (lower_bound <= entry['Total_Current_Charges'] <= upper_bound):
            row = [
                entry.get('location', ''),
                entry.get('Remidence_Address', ''),
                entry.get('vendor', ''),
                entry.get('accountnumber', ''),
                "",  # Empty Invoice Date
                entry.get('Date_Due', ''),
                entry['bill_date'],
                f"{total_category_sum:.2f}",
                f"{entry['Total_Current_Charges']:.2f}"
            ]
            ws.append(row)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{sub_company}_Out_Of_Variance_Report.xlsx"'
    wb.save(response)
    return response




def missing_bills_base_tem_report(request):
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    duration = request.GET.get('duration')
    entry_type = request.GET.get('entry_type')
    custom_date = request.GET.get('custom_date')  # Expected format: 'YYYY-MM'

    if not company or not sub_company or not duration or not entry_type:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    filter_conditions = Q(company=company, sub_company=sub_company, Entry_type=entry_type)

    today = datetime.today()
    if duration == 'Current month':
        filter_month_year = today.strftime("%Y-%m")
    elif duration == 'Previous month':
        previous_month = (today.replace(day=1) - timedelta(days=1))
        filter_month_year = previous_month.strftime("%Y-%m")
    elif duration == 'Previous three months':
        last_three_months = [
            (today.replace(day=1) - timedelta(days=30 * i)).strftime("%Y-%m")
            for i in range(3)
        ]
    elif duration == 'Custom' and custom_date:
        try:
            custom_date_parsed = datetime.strptime(custom_date, "%Y-%m")
            filter_month_year = custom_date_parsed.strftime("%Y-%m")
        except ValueError:
            return JsonResponse({'message': 'Invalid custom date format'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        return JsonResponse({'message': 'Invalid duration or custom_date'}, status=status.HTTP_400_BAD_REQUEST)

    # Retrieve filtered data
    all_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(filter_conditions)

    # Filter entries by matching month and year in bill_date
    filtered_data = []
    for item in all_data:
        if item.bill_date:  # Check if bill_date is not None
            try:
                # Parse the bill_date to extract month and year
                bill_date = datetime.strptime(item.bill_date, "%B %d %Y")
                bill_month_year = bill_date.strftime("%Y-%m")

                # Check if bill_month_year matches the desired filter
                if duration == 'Previous three months' and bill_month_year in last_three_months:
                    filtered_data.append(item)
                elif duration != 'Previous three months' and bill_month_year == filter_month_year:
                    filtered_data.append(item)
            except ValueError:
                # Skip entries that don't match the expected date format
                continue

    # Convert the filtered queryset to a list of dictionaries
    data_list = [model_to_dict(entry) for entry in filtered_data]

    return JsonResponse(data_list, safe=False)




def missing_bills_base_tem_report_download(request):
    # Extract query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    duration = request.GET.get('duration')
    entry_type = request.GET.get('entry_type')
    custom_date = request.GET.get('custom_date')  # Expected format: 'YYYY-MM'

    # Validate input parameters
    if not company or not sub_company or not duration or not entry_type:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    filter_conditions = Q(company=company, sub_company=sub_company, Entry_type=entry_type)

    today = datetime.today()
    if duration == 'Current month':
        filter_month_year = today.strftime("%Y-%m")
    elif duration == 'Previous month':
        previous_month = (today.replace(day=1) - timedelta(days=1))
        filter_month_year = previous_month.strftime("%Y-%m")
    elif duration == 'Previous three months':
        last_three_months = [
            (today.replace(day=1) - timedelta(days=30 * i)).strftime("%Y-%m")
            for i in range(3)
        ]
    elif duration == 'Custom' and custom_date:
        try:
            custom_date_parsed = datetime.strptime(custom_date, "%Y-%m")
            filter_month_year = custom_date_parsed.strftime("%Y-%m")
        except ValueError:
            return HttpResponse("Invalid custom date format", status=status.HTTP_400_BAD_REQUEST)
    else:
        return HttpResponse("Invalid duration or custom_date", status=status.HTTP_400_BAD_REQUEST)

    # Retrieve filtered data
    all_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(filter_conditions)

    # Filter entries by matching month and year in bill_date
    filtered_data = []
    for item in all_data:
        if item.bill_date:  # Check if bill_date is not None
            try:
                # Parse the bill_date to extract month and year
                bill_date = datetime.strptime(item.bill_date, "%B %d %Y")
                bill_month_year = bill_date.strftime("%Y-%m")

                # Check if bill_month_year matches the desired filter
                if duration == 'Previous three months' and bill_month_year in last_three_months:
                    filtered_data.append(item)
                elif duration != 'Previous three months' and bill_month_year == filter_month_year:
                    filtered_data.append(item)
            except ValueError:
                continue

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Missing Bills Report'

    # Main heading
    main_heading = 'Missing Bills Report'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = ["Site", "Location Address", "Vendor", "Account #", "Invoice Method", "Bill Date"]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in filtered_data:
        try:
            bill_date_str = datetime.strptime(item.bill_date, "%B %d %Y").strftime('%B %d %Y')
        except (ValueError, TypeError):
            bill_date_str = item.bill_date  # Retain original if parsing fails

        row = [
            item.location,
            item.Remidence_Address,
            item.vendor,
            item.accountnumber,
            item.Invoice_Method,
            bill_date_str
        ]
        ws.append(row)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Save workbook to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{sub_company}_Missing_Bills_Base_Report.xlsx"'
    wb.save(response)
    return response





def missing_bills_consoliated_tem_report(request):
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    duration = request.GET.get('duration')
    entry_type = request.GET.get('entry_type')
    custom_date = request.GET.get('custom_date')  # Expected format: 'YYYY-MM'

    if not company or not sub_company or not duration or not entry_type:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    filter_conditions = Q(company=company, sub_company=sub_company, Entry_type=entry_type)

    today = datetime.today()
    if duration == 'Current month':
        filter_month_year = today.strftime("%Y-%m")
    elif duration == 'Previous month':
        previous_month = (today.replace(day=1) - timedelta(days=1))
        filter_month_year = previous_month.strftime("%Y-%m")
    elif duration == 'Previous three months':
        last_three_months = [
            (today.replace(day=1) - timedelta(days=30 * i)).strftime("%Y-%m")
            for i in range(3)
        ]
    elif duration == 'Custom' and custom_date:
        try:
            custom_date_parsed = datetime.strptime(custom_date, "%Y-%m")
            filter_month_year = custom_date_parsed.strftime("%Y-%m")
        except ValueError:
            return JsonResponse({'message': 'Invalid custom date format'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        return JsonResponse({'message': 'Invalid duration or custom_date'}, status=status.HTTP_400_BAD_REQUEST)

    # Retrieve filtered data
    all_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(filter_conditions)

    # Filter entries by matching month and year in bill_date
    filtered_data = []
    for item in all_data:
        if item.bill_date:  # Check if bill_date is not None
            try:
                # Parse the bill_date to extract month and year
                bill_date = datetime.strptime(item.bill_date, "%B %d %Y")
                bill_month_year = bill_date.strftime("%Y-%m")

                # Check if bill_month_year matches the desired filter
                if duration == 'Previous three months' and bill_month_year in last_three_months:
                    filtered_data.append(item)
                elif duration != 'Previous three months' and bill_month_year == filter_month_year:
                    filtered_data.append(item)
            except ValueError:
                # Skip entries that don't match the expected date format
                continue

    # Convert the filtered queryset to a list of dictionaries
    data_list = [model_to_dict(entry) for entry in filtered_data]

    return JsonResponse(data_list, safe=False)




def missing_bills_consoliated_tem_report_download(request):
    # Extract query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    duration = request.GET.get('duration')
    entry_type = request.GET.get('entry_type')
    custom_date = request.GET.get('custom_date')  # Expected format: 'YYYY-MM'

    # Validate input parameters
    if not company or not sub_company or not duration or not entry_type:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    filter_conditions = Q(company=company, sub_company=sub_company, Entry_type=entry_type)

    today = datetime.today()
    if duration == 'Current month':
        filter_month_year = today.strftime("%Y-%m")
    elif duration == 'Previous month':
        previous_month = (today.replace(day=1) - timedelta(days=1))
        filter_month_year = previous_month.strftime("%Y-%m")
    elif duration == 'Previous three months':
        last_three_months = [
            (today.replace(day=1) - timedelta(days=30 * i)).strftime("%Y-%m")
            for i in range(3)
        ]
    elif duration == 'Custom' and custom_date:
        try:
            custom_date_parsed = datetime.strptime(custom_date, "%Y-%m")
            filter_month_year = custom_date_parsed.strftime("%Y-%m")
        except ValueError:
            return HttpResponse("Invalid custom date format", status=status.HTTP_400_BAD_REQUEST)
    else:
        return HttpResponse("Invalid duration or custom_date", status=status.HTTP_400_BAD_REQUEST)

    # Retrieve filtered data
    all_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(filter_conditions)

    # Filter entries by matching month and year in bill_date
    filtered_data = []
    for item in all_data:
        if item.bill_date:  # Check if bill_date is not None
            try:
                # Parse the bill_date to extract month and year
                bill_date = datetime.strptime(item.bill_date, "%B %d %Y")
                bill_month_year = bill_date.strftime("%Y-%m")

                # Check if bill_month_year matches the desired filter
                if duration == 'Previous three months' and bill_month_year in last_three_months:
                    filtered_data.append(item)
                elif duration != 'Previous three months' and bill_month_year == filter_month_year:
                    filtered_data.append(item)
            except ValueError:
                continue

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Missing Bills Report'

    # Main heading
    main_heading = 'Missing Bills Report'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = ["Site", "Location Address", "Vendor", "Account #", "Invoice Method", "Bill Date"]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in filtered_data:
        try:
            bill_date_str = datetime.strptime(item.bill_date, "%B %d %Y").strftime('%B %d %Y')
        except (ValueError, TypeError):
            bill_date_str = item.bill_date  # Retain original if parsing fails

        row = [
            item.location,
            item.Remidence_Address,
            item.vendor,
            item.accountnumber,
            item.Invoice_Method,
            bill_date_str
        ]
        ws.append(row)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Save workbook to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{sub_company}_Missing_Bills_Consoliated_Report.xlsx"'
    wb.save(response)
    return response




def cost_center_summary_report(request):
    # Extract query parameters from the request
    company = request.GET.get('company')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return JsonResponse({'message': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter by sub_company and ensure bill_date format is valid
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company
    )

    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                item_dict = model_to_dict(item)
                
                # Ensure Total_Current_Charges is not None before replacing
                if item.Total_Current_Charges is not None:
                    item_dict['Total_Current_Charges'] = float(item.Total_Current_Charges.replace(',', ''))
                else:
                    item_dict['Total_Current_Charges'] = 0  # Set to 0 if None
                
                # Directly add cost_centers assuming it's already a list
                cost_centers = item.cost_centers if isinstance(item.cost_centers, list) else []
                item_dict['cost_centers'] = cost_centers
                result_list.append(item_dict)
        except ValueError:
            continue

    # Calculate cost center occurrences across accountnumber and vendor
    cost_center_account_count = defaultdict(set)
    cost_center_vendor_count = defaultdict(set)

    for entry in result_list:
        account_number = entry.get('accountnumber')
        vendor = entry.get('vendor')
        cost_centers = entry.get('cost_centers', [])
        
        for cost_center in cost_centers:
            cost_center_name = cost_center.get('cost_center_name')
            if cost_center_name:
                cost_center_account_count[cost_center_name].add(account_number)
                cost_center_vendor_count[cost_center_name].add(vendor)

    # Prepare final counts
    for entry in result_list:
        cost_centers = entry.get('cost_centers', [])
        for cost_center in cost_centers:
            cost_center_name = cost_center.get('cost_center_name')
            if cost_center_name:
                cost_center['account_count'] = len(cost_center_account_count[cost_center_name])
                cost_center['vendor_count'] = len(cost_center_vendor_count[cost_center_name])

    # Return the updated response
    return JsonResponse(result_list, safe=False)




def cost_center_summary_report_download(request):
    # Extract query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Validate input parameters
    if not company or not sub_company or not start_date or not end_date:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return HttpResponse("Invalid date format", status=status.HTTP_400_BAD_REQUEST)

    # Filter data based on sub_company, company, and bill_date
    filtered_data = []
    all_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        company=company,
        sub_company=sub_company
    )

    for item in all_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                filtered_data.append(item)
        except (ValueError, TypeError):
            continue

    if not filtered_data:
        return HttpResponse("No data found for the given parameters.", status=404)

    # Calculate cost center occurrences across accountnumber and vendor
    cost_center_account_count = defaultdict(set)
    cost_center_vendor_count = defaultdict(set)

    for item in filtered_data:
        account_number = item.accountnumber
        vendor = item.vendor
        cost_centers_data = json.loads(item.cost_centers) if isinstance(item.cost_centers, str) else item.cost_centers or []

        for cost_center in cost_centers_data:
            cost_center_name = cost_center.get('cost_center_name')
            if cost_center_name:
                cost_center_account_count[cost_center_name].add(account_number)
                cost_center_vendor_count[cost_center_name].add(vendor)

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cost Center Summary Report'

    # Main heading
    main_heading = 'Cost Center Summary Report'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = ["Cost Center Name", "Total Amount", "# BAN's", "# Vendor's"]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in filtered_data:
        total_amount = float(item.Total_Current_Charges.replace(',', '')) if item.Total_Current_Charges else 0
        cost_centers_data = json.loads(item.cost_centers) if isinstance(item.cost_centers, str) else item.cost_centers or []

        for cost_center in cost_centers_data:
            cost_center_name = cost_center.get('cost_center_name', 'N/A')
            account_count = len(cost_center_account_count[cost_center_name])
            vendor_count = len(cost_center_vendor_count[cost_center_name])
            
            row = [
                cost_center_name,
                f"{total_amount:.2f}",
                account_count,
                vendor_count
            ]
            ws.append(row)

    # Adjust column widths to fit content
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = (max_length + 2)

    # Save workbook to a temporary in-memory location before sending it in the response
    response_stream = BytesIO()
    wb.save(response_stream)
    response_stream.seek(0)  # Move to the beginning of the stream

    # Prepare the HTTP response with correct headers for Excel download
    response = HttpResponse(
        response_stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{sub_company}_Cost_Center_Summary_Report.xlsx"'
    
    return response




def cost_center_detailed_report(request):
    # Extract query parameters from the request
    company = request.GET.get('company')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')

    print(sub_company, end_date, start_date)

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return JsonResponse({'message': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter by sub_company and ensure bill_date format is valid
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company
    )

    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                # Convert the entire object to a dictionary
                item_dict = model_to_dict(item)
                
                # Ensure Total_Current_Charges is not None before replacing
                if item.Total_Current_Charges is not None:
                    item_dict['Total_Current_Charges'] = float(item.Total_Current_Charges.replace(',', ''))
                else:
                    item_dict['Total_Current_Charges'] = 0  # Set to 0 if None

                result_list.append(item_dict)
        except ValueError:
            # Skip entries that don't match the expected date format
            continue

    # Sort the result list by Total_Current_Charges in descending order without a limit
    result_list = sorted(result_list, key=lambda x: x['Total_Current_Charges'], reverse=True)

    # Return the updated response without the baseline table calculations
    return JsonResponse(result_list, safe=False)




def cost_center_detailed_report_download(request):
    # Extract query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Validate input parameters
    if not company or not sub_company or not start_date or not end_date:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return HttpResponse("Invalid date format", status=status.HTTP_400_BAD_REQUEST)

    # Filter data based on sub_company, company, and bill_date
    filtered_data = []
    all_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        company=company,
        sub_company=sub_company
    )

    for item in all_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                filtered_data.append(item)
        except (ValueError, TypeError):
            continue

    if not filtered_data:
        return HttpResponse("No data found for the given parameters.", status=404)

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cost Center Summary Report'

    # Main heading
    main_heading = 'Cost Center Summary Report'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = ["Cost Center Name", "Total Amount", "BAN", "Vendor", "Master Account", "Approved By", "Approved On", "Type"]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in filtered_data:
        total_amount = float(item.Total_Current_Charges.replace(',', '')) if item.Total_Current_Charges else 0
        cost_centers_data = json.loads(item.cost_centers) if isinstance(item.cost_centers, str) else item.cost_centers or []

        for cost_center in cost_centers_data:
            row = [
                cost_center.get('cost_center_name', ''),
                f"{total_amount:.2f}",
                item.accountnumber,
                item.vendor,
                item.master_account,
                "",
                "",
                cost_center.get('cost_center_type', '')
            ]
            ws.append(row)

    # Adjust column widths to fit content
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = (max_length + 2)

    # Save workbook to a temporary in-memory location before sending it in the response
    response_stream = BytesIO()
    wb.save(response_stream)
    response_stream.seek(0)

    # Prepare the HTTP response with correct headers for Excel download
    response = HttpResponse(
        response_stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{sub_company}_Cost_Center_Detailed_Report.xlsx"'
    
    return response






def location_summary_report(request):
    # Extract query parameters from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    company = request.GET.get('company')

    print(sub_company, end_date, start_date)

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return JsonResponse({'message': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter by sub_company and ensure bill_date format is valid
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),sub_company=sub_company,company=company)


    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                result_list.append({
                    'id': item.id,  # Include relevant fields
                    'bill_date': item.bill_date,
                    'sub_company': item.sub_company,
                    'Total_Current_Charges': item.Total_Current_Charges,
                    'vendor': item.vendor,
                    'Remidence_Address': item.Remidence_Address,
                })
        except ValueError:
            # Skip entries that don't match the expected date format
            continue

    return JsonResponse(result_list, safe=False)

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from OnBoard.Ban.models import BaseDataTable  # Replace with your actual model import

def generate_month_range(start_date, end_date):
    """Generate a list of all month/year combinations between start_date and end_date."""
    start = datetime.strptime(start_date, '%Y-%m-%d')
    end = datetime.strptime(end_date, '%Y-%m-%d')
    months = []

    while start <= end:
        months.append(start.strftime('%B %Y'))
        start += relativedelta(months=1)

    return months

def location_summary_report_download(request):
    # Get query parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    company = request.GET.get('company')

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return HttpResponse("Invalid date format", status=status.HTTP_400_BAD_REQUEST)

    # Generate the full list of months within the date range
    all_months = generate_month_range(start_date, end_date)

    # Fetch and filter data from the database
    data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        company=company,
        sub_company=sub_company,
        bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'  # Ensure valid date format
    )

    filtered_data = []
    for item in data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                filtered_data.append(item)
        except ValueError:
            continue

    # Group data by vendor and month/year with additional fields
    grouped_data = {}
    for item in filtered_data:
        formatted_date = datetime.strptime(item.bill_date, '%B %d %Y').strftime('%B %Y')
        if item.vendor not in grouped_data:
            grouped_data[item.vendor] = {
                "city": item.Billing_city or "",
                "state": item.Billing_State or "",
                "address": item.Remidence_Address or "",
                "charges": {}
            }
        current_charges = (
            float(item.Total_Current_Charges.replace(',', '')) if item.Total_Current_Charges else 0.0
        )
        grouped_data[item.vendor]["charges"][formatted_date] = current_charges

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Location Summary Report'

    # Add main heading
    main_heading = 'Expense Management - Total Expense By Location'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_months) + 4)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Add table headers
    headers = ['Vendor', 'City', 'State', 'Address'] + all_months
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='ADD8E6', fill_type='solid')  # Light blue background
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows with unique vendors and their charges
    for vendor, details in grouped_data.items():
        row = [vendor, details["city"], details["state"], details["address"]]
        for date in all_months:
            value = details["charges"].get(date, '0.00')
            row.append(value)
        ws.append(row)

    # Add a total row at the end with calculated sums
    total_row = ['Total', '', '', '']
    for col_index in range(5, len(all_months) + 5):
        sum_value = sum(
            cell.value for col in ws.iter_cols(min_col=col_index, max_col=col_index, min_row=3, max_row=ws.max_row)
            for cell in col if isinstance(cell.value, (int, float))
        )
        total_row.append("{:,.2f}".format(sum_value))

    ws.append(total_row)
    last_row_idx = ws.max_row
    for cell in ws[last_row_idx]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFFFE0', fill_type='solid')  # Light yellow background
        cell.alignment = Alignment(horizontal='right')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Dynamically adjust column widths and right-align amount cells
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
        adjusted_width = max_length + 2  # Add extra space for better visibility
        col_letter = openpyxl.utils.get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = adjusted_width

        # Right-align all cells in amount columns (starting from the 5th column)
        if col_idx >= 5:
            for cell in column_cells[2:]:  # Start from row 3 to avoid heading rows
                cell.alignment = Alignment(horizontal='right')

    # Create HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_Total_Expense_By_Location.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response




from django.forms.models import model_to_dict
from django.db.models import Q
from datetime import datetime, timedelta
from django.http import JsonResponse
from collections import defaultdict

def duplicate_bills_base_tem_report(request):
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    entry_type = request.GET.get('entry_type')
    custom_date = request.GET.get('date')  # Expected format: 'YYYY-MM'

    if not company or not sub_company or not entry_type or not custom_date:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    # Build filter conditions
    filter_conditions = Q(company=company, sub_company=sub_company, Entry_type=entry_type)

    # Validate and format custom date
    try:
        custom_date_parsed = datetime.strptime(custom_date, "%Y-%m")
        filter_month_year = custom_date_parsed.strftime("%Y-%m")
    except ValueError:
        return JsonResponse({'message': 'Invalid custom date format'}, status=status.HTTP_400_BAD_REQUEST)

    # Retrieve data matching the filter conditions
    all_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(filter_conditions)

    # Filter entries by matching month and year in bill_date
    filtered_data = []
    for item in all_data:
        if item.bill_date:  # Check if bill_date is not None
            try:
                # Parse the bill_date to extract month and year
                bill_date = datetime.strptime(item.bill_date, "%B %d %Y")
                bill_month_year = bill_date.strftime("%Y-%m")

                # Check if bill_month_year matches the custom date filter
                if bill_month_year == filter_month_year:
                    # Convert item to dict and add to filtered_data
                    item_dict = model_to_dict(item)
                    item_dict['Total_Current_Charges'] = float(item.Total_Current_Charges.replace(',', '')) if item.Total_Current_Charges else 0
                    filtered_data.append(item_dict)
            except ValueError:
                continue

    # Now, retrieve data from BaselineDataTable for each filtered bill_date
    for entry in filtered_data:
        bill_date_str = entry.get('bill_date')
        if bill_date_str:
            try:
                # Parse bill_date in the baseline table's format
                bill_date = datetime.strptime(bill_date_str, "%B %d %Y")

                # Get matching baseline entries by company, sub_company, and bill_date
                baseline_entries = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
                    company=company,
                    sub_company=sub_company,
                    bill_date=bill_date.strftime("%B %d %Y")  # Format as "Month Day Year"
                )

                # Sum values from category_object for each matching baseline entry
                total_category_sum = 0
                for baseline_entry in baseline_entries:
                    category_object = baseline_entry.category_object or {}
                    if isinstance(category_object, dict):
                        total_category_sum += sum(
                            float(value) for section in category_object.values() if isinstance(section, dict)
                            for value in section.values() if isinstance(value, (int, float))
                        )

                # Attach the sum to the entry
                entry['Total_Category_Sum'] = total_category_sum
            except ValueError:
                # Skip if the bill_date format does not match
                entry['Total_Category_Sum'] = 0  # Set to 0 if any parsing issues occur

    return JsonResponse(filtered_data, safe=False)

from OnBoard.Organization.ser import OrganizationShowSerializer

class GetSubCompanyByCompany(APIView):
    def get(self, request):
        # Get the 'company_name' parameter from the request
        company_name = request.GET.get('company_name')

        print("Getting Sub Company", company_name)

        # Validate that the parameter was provided
        if not company_name:
            return Response({'message': 'Missing required parameter: company_name'}, status=status.HTTP_400_BAD_REQUEST)

        # Filter the Organizations by company_name
        data_models = Organizations.objects.filter(company_name=company_name)

        # Serialize the filtered data
        serializer = OrganizationShowSerializer(data_models, many=True)
        return Response(serializer.data)    
    
def duplicate_bills_base_tem_report(request):
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    entry_type = request.GET.get('entry_type')
    custom_date = request.GET.get('date')  # Expected format: 'YYYY-MM'

    if not company or not sub_company or not entry_type or not custom_date:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    # Build filter conditions
    filter_conditions = Q(company=company, sub_company=sub_company, Entry_type=entry_type)

    # Validate and format custom date
    try:
        custom_date_parsed = datetime.strptime(custom_date, "%Y-%m")
        filter_month_year = custom_date_parsed.strftime("%Y-%m")
    except ValueError:
        return JsonResponse({'message': 'Invalid custom date format'}, status=status.HTTP_400_BAD_REQUEST)

    # Retrieve data matching the filter conditions
    all_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(filter_conditions)

    # Filter entries by matching month and year in bill_date
    filtered_data = []
    for item in all_data:
        if item.bill_date:  # Check if bill_date is not None
            try:
                # Parse the bill_date to extract month and year
                bill_date = datetime.strptime(item.bill_date, "%B %d %Y")
                bill_month_year = bill_date.strftime("%Y-%m")

                # Check if bill_month_year matches the custom date filter
                if bill_month_year == filter_month_year:
                    # Convert item to dict and add to filtered_data
                    item_dict = model_to_dict(item)
                    item_dict['Total_Current_Charges'] = float(item.Total_Current_Charges.replace(',', '')) if item.Total_Current_Charges else 0
                    filtered_data.append(item_dict)
            except ValueError:
                continue

    # Now, retrieve data from BaselineDataTable for each filtered bill_date
    for entry in filtered_data:
        bill_date_str = entry.get('bill_date')
        if bill_date_str:
            try:
                # Parse bill_date in the baseline table's format
                bill_date = datetime.strptime(bill_date_str, "%B %d %Y")

                # Get matching baseline entries by company, sub_company, and bill_date
                baseline_entries = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
                    company=company,
                    sub_company=sub_company,
                    bill_date=bill_date.strftime("%B %d %Y")  # Format as "Month Day Year"
                )

                # Sum values from category_object for each matching baseline entry
                total_category_sum = 0
                for baseline_entry in baseline_entries:
                    category_object = baseline_entry.category_object or {}
                    if isinstance(category_object, dict):
                        total_category_sum += sum(
                            float(value) for section in category_object.values() if isinstance(section, dict)
                            for value in section.values() if isinstance(value, (int, float))
                        )

                # Attach the sum to the entry
                entry['Total_Category_Sum'] = total_category_sum
            except ValueError:
                # Skip if the bill_date format does not match
                entry['Total_Category_Sum'] = 0  # Set to 0 if any parsing issues occur

    return JsonResponse(filtered_data, safe=False)




def duplicate_bills_base_tem_report_download(request):
    # Extract query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    entry_type = request.GET.get('entry_type')
    custom_date = request.GET.get('date')  # Expected format: 'YYYY-MM'

    if not company or not sub_company or not entry_type or not custom_date:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    # Validate custom date
    try:
        custom_date_parsed = datetime.strptime(custom_date, "%Y-%m")
        filter_month_year = custom_date_parsed.strftime("%Y-%m")
    except ValueError:
        return HttpResponse("Invalid custom date format", status=status.HTTP_400_BAD_REQUEST)

    # Retrieve data matching the conditions
    filter_conditions = Q(company=company, sub_company=sub_company, Entry_type=entry_type)
    data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(filter_conditions)

    # Filter entries by matching month and year in bill_date
    filtered_data = []
    for item in data:
        if item.bill_date:
            try:
                bill_date = datetime.strptime(item.bill_date, "%B %d %Y")
                bill_month_year = bill_date.strftime("%Y-%m")
                if bill_month_year == filter_month_year:
                    filtered_data.append(item)
            except ValueError:
                continue

    # Fetch category sums from BaselineDataTable for each relevant bill date
    for entry in filtered_data:
        bill_date = entry.bill_date
        category_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
            company=company,
            sub_company=sub_company,
            bill_date=bill_date
        )
        total_category_sum = sum(
            sum(float(val) for val in cat_obj.values() if isinstance(val, (int, float)))
            for item in category_data for cat_obj in (item.category_object or {}).values()
            if isinstance(cat_obj, dict)
        )
        entry.Total_Category_Sum = total_category_sum

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Duplicate Bills Report'

    # Add main heading
    main_heading = "Duplicate Bills Base TEM Report"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)  # Adjust columns as per the headers count
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light grey background

    # Set the headers
    headers = [
        "Account Number", "Invoice #", "Site", "Address", "Bill Date", "BBA", "BBA Var.",
        "Bill Amount", "Approved By", "Date Entered", "Date Approved"
    ]
    ws.append(headers)

    # Style the headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in filtered_data:
        row = [
            item.accountnumber,
            item.InvoiceNumber or "",
            item.location or "",
            item.Remidence_Address or "",
            item.bill_date or "",
            f"{getattr(item, 'Total_Category_Sum', 0):.2f}",
            "",  # BBA Var. is blank as per the frontend
            item.Total_Current_Charges or "0.00",
            "",
            "",
            ""
        ]
        ws.append(row)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Use BytesIO to create an in-memory file
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)  # Reset the stream position

    # Create the response
    filename = f"{sub_company}_Duplicate_Bills_Base_Report_{custom_date}.xlsx"
    response = HttpResponse(
        file_stream,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



def duplicate_bills_consoliated_tem_report(request):
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    entry_type = request.GET.get('entry_type')
    custom_date = request.GET.get('date')  # Expected format: 'YYYY-MM'

    if not company or not sub_company or not entry_type or not custom_date:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    # Build filter conditions
    filter_conditions = Q(company=company, sub_company=sub_company, Entry_type=entry_type)

    # Validate and format custom date
    try:
        custom_date_parsed = datetime.strptime(custom_date, "%Y-%m")
        filter_month_year = custom_date_parsed.strftime("%Y-%m")
    except ValueError:
        return JsonResponse({'message': 'Invalid custom date format'}, status=status.HTTP_400_BAD_REQUEST)

    # Retrieve data matching the filter conditions
    all_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(filter_conditions)

    # Filter entries by matching month and year in bill_date
    filtered_data = []
    for item in all_data:
        if item.bill_date:  # Check if bill_date is not None
            try:
                # Parse the bill_date to extract month and year
                bill_date = datetime.strptime(item.bill_date, "%B %d %Y")
                bill_month_year = bill_date.strftime("%Y-%m")

                # Check if bill_month_year matches the custom date filter
                if bill_month_year == filter_month_year:
                    # Convert item to dict and add to filtered_data
                    item_dict = model_to_dict(item)
                    item_dict['Total_Current_Charges'] = float(item.Total_Current_Charges.replace(',', '')) if item.Total_Current_Charges else 0
                    filtered_data.append(item_dict)
            except ValueError:
                continue

    # Now, retrieve data from BaselineDataTable for each filtered bill_date
    for entry in filtered_data:
        bill_date_str = entry.get('bill_date')
        if bill_date_str:
            try:
                # Parse bill_date in the baseline table's format
                bill_date = datetime.strptime(bill_date_str, "%B %d %Y")

                # Get matching baseline entries by company, sub_company, and bill_date
                baseline_entries = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
                    company=company,
                    sub_company=sub_company,
                    bill_date=bill_date.strftime("%B %d %Y")  # Format as "Month Day Year"
                )

                # Sum values from category_object for each matching baseline entry
                total_category_sum = 0
                for baseline_entry in baseline_entries:
                    category_object = baseline_entry.category_object or {}
                    if isinstance(category_object, dict):
                        total_category_sum += sum(
                            float(value) for section in category_object.values() if isinstance(section, dict)
                            for value in section.values() if isinstance(value, (int, float))
                        )

                # Attach the sum to the entry
                entry['Total_Category_Sum'] = total_category_sum
            except ValueError:
                # Skip if the bill_date format does not match
                entry['Total_Category_Sum'] = 0  # Set to 0 if any parsing issues occur

    return JsonResponse(filtered_data, safe=False)




def duplicate_bills_consoliated_tem_report_download(request):
    # Extract query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    entry_type = request.GET.get('entry_type')
    custom_date = request.GET.get('date')  # Expected format: 'YYYY-MM'

    if not company or not sub_company or not entry_type or not custom_date:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    # Validate custom date
    try:
        custom_date_parsed = datetime.strptime(custom_date, "%Y-%m")
        filter_month_year = custom_date_parsed.strftime("%Y-%m")
    except ValueError:
        return HttpResponse("Invalid custom date format", status=status.HTTP_400_BAD_REQUEST)

    # Retrieve data matching the conditions
    filter_conditions = Q(company=company, sub_company=sub_company, Entry_type=entry_type)
    data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(filter_conditions)

    # Filter entries by matching month and year in bill_date
    filtered_data = []
    for item in data:
        if item.bill_date:
            try:
                bill_date = datetime.strptime(item.bill_date, "%B %d %Y")
                bill_month_year = bill_date.strftime("%Y-%m")
                if bill_month_year == filter_month_year:
                    filtered_data.append(item)
            except ValueError:
                continue

    # Fetch category sums from BaselineDataTable for each relevant bill date
    for entry in filtered_data:
        bill_date = entry.bill_date
        category_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
            company=company,
            sub_company=sub_company,
            bill_date=bill_date
        )
        total_category_sum = sum(
            sum(float(val) for val in cat_obj.values() if isinstance(val, (int, float)))
            for item in category_data for cat_obj in (item.category_object or {}).values()
            if isinstance(cat_obj, dict)
        )
        entry.Total_Category_Sum = total_category_sum

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Duplicate Bills Report'

    # Add main heading
    main_heading = "Duplicate Bills Consoliated TEM Report"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)  # Adjust columns as per the headers count
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light grey background

    # Set the headers
    headers = [
        "Account Number", "Invoice #", "Site", "Address", "Bill Date", "BBA", "BBA Var.",
        "Bill Amount", "Approved By", "Date Entered", "Date Approved"
    ]
    ws.append(headers)

    # Style the headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in filtered_data:
        row = [
            item.accountnumber,
            item.InvoiceNumber or "",
            item.location or "",
            item.Remidence_Address or "",
            item.bill_date or "",
            f"{getattr(item, 'Total_Category_Sum', 0):.2f}",
            "",  # BBA Var. is blank as per the frontend
            item.Total_Current_Charges or "0.00",
            "",
            "",
            ""
        ]
        ws.append(row)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Use BytesIO to create an in-memory file
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)  # Reset the stream position

    # Create the response
    filename = f"{sub_company}_Duplicate_Bills_Consoliated_Report_{custom_date}.xlsx"
    response = HttpResponse(
        file_stream,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response




def service_by_type_report(request):
    # Extract query parameters from the request
    company = request.GET.get('company')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')

    print(sub_company, end_date, start_date)

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return JsonResponse({'message': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter by sub_company and ensure bill_date format is valid
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company
    )

    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                # Convert the entire object to a dictionary
                item_dict = model_to_dict(item)
                
                # Ensure Total_Current_Charges is not None before replacing
                if item.Total_Current_Charges is not None:
                    item_dict['Total_Current_Charges'] = float(item.Total_Current_Charges.replace(',', ''))
                else:
                    item_dict['Total_Current_Charges'] = 0  # Set to 0 if None

                result_list.append(item_dict)
        except ValueError:
            # Skip entries that don't match the expected date format
            continue

    # Sort the result list by Total_Current_Charges in descending order
    result_list = sorted(result_list, key=lambda x: x['Total_Current_Charges'], reverse=True)

    # Return the updated response with the total category sum
    return JsonResponse(result_list, safe=False)




def service_by_type_report_download(request):
    # Extract query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Validate input parameters
    if not company or not sub_company or not start_date or not end_date:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return HttpResponse("Invalid date format", status=status.HTTP_400_BAD_REQUEST)

    # Retrieve the data, ensuring dates are filtered correctly
    data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        company=company,
        sub_company=sub_company,
        bill_date__range=[start_date_parsed, end_date_parsed]
    )

    # Check if data exists
    if not data.exists():
        return HttpResponse("No data found for the given parameters.", status=404)

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Service By Type Report'

    # Main heading
    main_heading = 'Service By Type Report'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        "Vendor", "Account #", "Invoice #", "Invoice Status", "Bill Amount", "Service Type(s)"
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in data:
        row = [
            item.vendor or "",
            item.accountnumber or "",
            item.InvoiceNumber or "",
            item.status or "",
            f"{float(item.Total_Current_Charges.replace(',', '')) if item.Total_Current_Charges else 0:.2f}",
            item.services or "",
        ]
        ws.append(row)

    # Add a total row at the end
    total_bill_amount = sum(float(item.Total_Current_Charges.replace(',', '')) for item in data if item.Total_Current_Charges)
    ws.append(["", "", "", "Total", f"{total_bill_amount:.2f}", ""])
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='FFFFE0', fill_type='solid')

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Prepare HTTP response with the Excel file
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    # Create the response
    filename = f"Service_By_Type_Report_{sub_company}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        file_stream,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response



def payment_detail_report(request):
    # Extract query parameters from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    vendor = request.GET.get('vendor')
    ban = request.GET.get('ban')
    company = request.GET.get('company')
    invoice_type = request.GET.get('invoice_type')

    

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return JsonResponse({'message': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter data from BaseDataTable
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        vendor=vendor,
        invoice_type=invoice_type
        
    )
    if ban:
        filtered_data = filtered_data.filter(accountnumber=ban)

    # Filter the data by the date range
    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                result_list.append(item)
        except ValueError:
            continue

    # Convert filtered_data to a list of dictionaries
    simple_response = list(filtered_data.values())

    # Filter data from BaselineDataTable by bill_date and other parameters
    baseline_filtered_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),  # Ensure valid bill_date format
        sub_company=sub_company,
        company=company,
        vendor=vendor
    )
    if ban:
        baseline_filtered_data = baseline_filtered_data.filter(account_number=ban)

    # Aggregate sums for each bill_date
    date_sum_map = {}
    for item in baseline_filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                date_key = bill_date.strftime('%B %d %Y')
                category_object = item.category_object or {}

                if category_object and isinstance(category_object, dict):
                    if date_key not in date_sum_map:
                        date_sum_map[date_key] = 0  # Initialize sum for this date

                    # Sum all values in the category_object for this bill_date
                    for section, values in category_object.items():
                        if isinstance(values, dict):  # Ensure section contains a dictionary
                            section_sum = sum(
                                float(value) for value in values.values() if isinstance(value, (int, float))
                            )
                            date_sum_map[date_key] += section_sum

                            # Debugging: Print section and section sum
                            print(f"Bill Date: {date_key}, Section: {section}, Section Sum: {section_sum}")
                else:
                    # If category_object is empty or not a dict, ensure it adds 0
                    if date_key not in date_sum_map:
                        date_sum_map[date_key] = 0
        except (ValueError, TypeError) as e:
            # Handle parsing errors
            print("Error parsing item:", e)
            continue

    # Add the total sum per date to the response
    for entry in simple_response:
        bill_date_str = entry['bill_date']
        entry['Total_Category_Sum'] = date_sum_map.get(bill_date_str, 0)

    # Return the updated response
    return JsonResponse(simple_response, safe=False)






def payment_detail_report_download(request):
    # Get query parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    vendor = request.GET.get('vendor')
    ban = request.GET.get('ban')
    company = request.GET.get('company')
    invoice_type = request.GET.get('invoice_type')

    

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return HttpResponse("Invalid date format", status=status.HTTP_400_BAD_REQUEST)

    # Filter data from BaseDataTable
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        vendor=vendor,
        invoice_type=invoice_type
    )
    if ban:
        filtered_data = filtered_data.filter(accountnumber=ban)

    # Filter by date range
    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                result_list.append(item)
        except ValueError:
            continue

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Payment Detail Report'

    # Main heading
    main_heading = 'Payment Detail Report'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        "Vendor", "Account #", "Invoice #", "Date Approved", "Due Date",
        "Amount Due", "Payment Detail"
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')  # Blue background
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Filter data from BaselineDataTable and calculate total from category_object
    baseline_filtered_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        vendor=vendor
    )
    if ban:
        baseline_filtered_data = baseline_filtered_data.filter(account_number=ban)

    # Create a mapping of bill_date to total sum from category_object
    bba_sum_map = {}
    for item in baseline_filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                date_key = bill_date.strftime('%B %d %Y')
                category_object = item.category_object or {}

                if date_key not in bba_sum_map:
                    bba_sum_map[date_key] = 0

                # Sum all values in the category_object for this bill_date
                if isinstance(category_object, dict):
                    for section, values in category_object.items():
                        if isinstance(values, dict):
                            bba_sum_map[date_key] += sum(
                                float(value) for value in values.values() if isinstance(value, (int, float))
                            )
        except (ValueError, TypeError):
            continue

    # Populate data rows and apply alignment
    for item in result_list:
        bill_date_str = item.bill_date
        bba_value = bba_sum_map.get(bill_date_str, 0)  # Get the BBA value for the bill date or 0 if not present
        row = [
            item.vendor,
            item.accountnumber,
            item.InvoiceNumber,
            "",
            item.Date_Due,
            "",
            ""
        ]
        ws.append(row)
        last_row = ws.max_row
        ws[f'A{last_row}'].alignment = Alignment(horizontal='center')  # Align accountnumber
        ws[f'B{last_row}'].alignment = Alignment(horizontal='center')  # Align InvoiceNumber
        ws[f'I{last_row}'].alignment = Alignment(horizontal='right')   # Align Total_Current_Charges

    # Style the entire worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
            if isinstance(cell.value, (int, float)) and cell.column == 9:  # Check for Bill Amount column
                cell.alignment = Alignment(horizontal='right')

    # Adjust column widths safely by skipping merged cells
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Get the column letter using openpyxl's utility function
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Check if cell is not a MergedCell
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_Payment_Detail_Report.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response






def mobile_unapproved_bills_report(request):
    # Extract query parameters from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    vendor = request.GET.get('vendor')
    ban = request.GET.get('ban')
    company = request.GET.get('company')
    approval_status = request.GET.get('approval_status')

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return JsonResponse({'message': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter by sub_company and ensure bill_date format is valid
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        vendor=vendor,
        approval_status=approval_status,
    )
    if ban:
        filtered_data = filtered_data.filter(accountnumber=ban)

    # Filter the data by the date range
    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                result_list.append(item)
        except ValueError:
            continue

    # Convert filtered_data to a list of dictionaries
    simple_response = list(filtered_data.values())

    # Filter data from BaselineDataTable and calculate total from category_object
    baseline_filtered_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        vendor=vendor
    )
    if ban:
        baseline_filtered_data = baseline_filtered_data.filter(account_number=ban)

    # Calculate total sum for each date and include it in the response
    total_sums_by_date = {}
    for item in baseline_filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                category_object = item.category_object or {}
                date_key = bill_date.strftime('%B %d %Y')
                total_sum = sum(
                    float(value) for section in category_object.values()
                    if isinstance(section, dict)
                    for value in section.values() if isinstance(value, (int, float))
                )
                total_sums_by_date[date_key] = total_sums_by_date.get(date_key, 0) + total_sum
        except (ValueError, TypeError, json.JSONDecodeError):
            continue

    # Add the calculated totals to the response
    for entry in simple_response:
        entry_date = entry['bill_date']
        entry['Total_Category_Sum'] = total_sums_by_date.get(entry_date, 0)

    # Return the updated response with the sum data
    return JsonResponse(simple_response, safe=False)








def format_number(number):
    """Format a number with commas and two decimal places."""
    return f"{number:,.2f}"

def mobile_unapproved_bills_report_download(request):
    # Get query parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_company = request.GET.get('sub_company')
    vendor = request.GET.get('vendor')
    ban = request.GET.get('ban')
    company = request.GET.get('company')

    # Validate input parameters
    if not start_date or not end_date or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return HttpResponse("Invalid date format", status=status.HTTP_400_BAD_REQUEST)

    # Filter data from the BaseDataTable
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        vendor=vendor,
        approval_status='False',
    )
    if ban:
        filtered_data = filtered_data.filter(accountnumber=ban)

    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                result_list.append(item)
        except ValueError:
            continue

    # Filter data from BaselineDataTable and calculate totals from category_object
    baseline_filtered_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        sub_company=sub_company,
        company=company,
        vendor=vendor
    )
    if ban:
        baseline_filtered_data = baseline_filtered_data.filter(account_number=ban)

    total_sums_by_date = {}
    for item in baseline_filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                category_object = json.loads(item.category_object) if item.category_object else {}
                date_key = bill_date.strftime('%B %d %Y')
                total_sum = sum(
                    float(value) for section in category_object.values()
                    if isinstance(section, dict)
                    for value in section.values() if isinstance(value, (int, float))
                )
                total_sums_by_date[date_key] = total_sums_by_date.get(date_key, 0) + total_sum
        except (ValueError, TypeError, json.JSONDecodeError):
            continue

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mobile Unapproved Bills Report'

    # Main heading
    main_heading = 'Mobile Unapproved Bills Report'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        "Invoice #", "Location", "Account #", "Vendor",
        "Bill Date", "Bill Amount", "Status"
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in result_list:
        bba_value = total_sums_by_date.get(item.bill_date, 0)  # Get the BBA value for the current Bill Date
        row = [
            item.InvoiceNumber,
            item.location,
            item.accountnumber,
            item.vendor,
            item.bill_date,
            item.Total_Current_Charges,
            ""
        ]
        ws.append(row)

    # Style the entire worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
            if isinstance(cell.value, (int, float)) and cell.column == 6:  # Check for Bill Amount column
                cell.alignment = Alignment(horizontal='right')

    # Adjust column widths safely by skipping merged cells
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_Mobile_Unapproved_Bills_Report.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response








def entered_bills_report(request):
    # Extract query parameters from the request
    company = request.GET.get('company')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    upload_by = request.GET.get('upload_by')

    

    # Validate input parameters
    if not start_date or not end_date or not company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return JsonResponse({'message': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)

    # Filter by sub_company and ensure bill_date format is valid
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        upload_by=upload_by,
        company=company
    )

    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                # Convert the entire object to a dictionary
                item_dict = model_to_dict(item)
                
                # Ensure Total_Current_Charges is not None before replacing
                if item.Total_Current_Charges is not None:
                    item_dict['Total_Current_Charges'] = float(item.Total_Current_Charges.replace(',', ''))
                else:
                    item_dict['Total_Current_Charges'] = 0  # Set to 0 if None

                result_list.append(item_dict)
        except ValueError:
            # Skip entries that don't match the expected date format
            continue

    # Sort the result list by Total_Current_Charges in descending order and get the top 25
    result_list = sorted(result_list, key=lambda x: x['Total_Current_Charges'], reverse=True)

    # Filter data from BaselineDataTable and calculate the total sum for each date
    baseline_filtered_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        company=company
    )

    # Calculate the total sum of category_object per bill_date
    date_sum_map = {}
    for item in baseline_filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                date_key = bill_date.strftime('%B %d %Y')
                category_object = item.category_object or {}

                if category_object and isinstance(category_object, dict):
                    if date_key not in date_sum_map:
                        date_sum_map[date_key] = 0  # Initialize sum for this date

                    # Sum all values in the category_object for this bill_date
                    for section, values in category_object.items():
                        if isinstance(values, dict):  # Ensure section contains a dictionary
                            section_sum = sum(
                                float(value) for value in values.values() if isinstance(value, (int, float))
                            )
                            date_sum_map[date_key] += section_sum

                            # Debugging: Print section and section sum
                            print(f"Bill Date: {date_key}, Section: {section}, Section Sum: {section_sum}")
                else:
                    # If category_object is empty or not a dict, ensure it adds 0
                    if date_key not in date_sum_map:
                        date_sum_map[date_key] = 0
        except (ValueError, TypeError) as e:
            # Handle parsing errors
            print("Error parsing item:", e)
            continue

    # Add the total sum per date to the result_list
    for entry in result_list:
        bill_date_str = entry['bill_date']
        entry['Total_Category_Sum'] = date_sum_map.get(bill_date_str, 0)

    # Return the updated response with the total category sum
    return JsonResponse(result_list, safe=False)






def entered_bills_report_download(request):
    # Get query parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    upload_by = request.GET.get('upload_by')
    company = request.GET.get('company')

    # Validate input parameters
    if not start_date or not end_date or not company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    try:
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return HttpResponse("Invalid date format", status=status.HTTP_400_BAD_REQUEST)

    # Filter data from the BaseDataTable
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        upload_by=upload_by,
        company=company
    )

    # Filter by date range and get top 25 highest Total_Current_Charges
    result_list = []
    for item in filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                result_list.append(item)
        except ValueError:
            continue

    # Sort the result list by Total_Current_Charges and get the top 25
    result_list.sort(key=lambda x: float(x.Total_Current_Charges.replace(',', '') if x.Total_Current_Charges else 0), reverse=True)
    top_25_data = result_list

    # Filter data from the BaselineDataTable for BBA values
    baseline_filtered_data = BaselineDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
        Q(bill_date__regex=r'^[A-Za-z]+\s\d+\s\d{4}$'),
        company=company
    )

    # Calculate the total sum of category_object per bill_date
    date_sum_map = {}
    for item in baseline_filtered_data:
        try:
            bill_date = datetime.strptime(item.bill_date, '%B %d %Y')
            if start_date_parsed <= bill_date <= end_date_parsed:
                date_key = bill_date.strftime('%B %d %Y')
                category_object = item.category_object or {}

                if category_object and isinstance(category_object, dict):
                    if date_key not in date_sum_map:
                        date_sum_map[date_key] = 0  # Initialize sum for this date

                    # Sum all values in the category_object for this bill_date
                    for section, values in category_object.items():
                        if isinstance(values, dict):  # Ensure section contains a dictionary
                            section_sum = sum(
                                float(value) for value in values.values() if isinstance(value, (int, float))
                            )
                            date_sum_map[date_key] += section_sum
        except (ValueError, TypeError, json.JSONDecodeError):
            continue

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Entered Bills Report'

    # Main heading
    main_heading = 'Entered Bills Report'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        "Account #", "Invoice #", "Vendor", "Site", "Address",
        "Bill Date", "BBA", "BBA Var", "Bill Amount", "Approved By",
        "Date Entered", "Date Approved"
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in top_25_data:
        bba_value = date_sum_map.get(item.bill_date, 0)  # Get the BBA value for the current Bill Date
        row = [
            item.accountnumber,
            item.InvoiceNumber,
            item.vendor,
            item.location,
            item.Remidence_Address,
            item.bill_date,
            f"{bba_value:.2f}",  # Format BBA value
            "",  # Placeholder for BBA Var
            item.Total_Current_Charges,
            "",
            
        ]
        ws.append(row)
        last_row = ws.max_row
        ws[f'A{last_row}'].alignment = Alignment(horizontal='center')  # Align accountnumber
        ws[f'B{last_row}'].alignment = Alignment(horizontal='center')  # Align InvoiceNumber
        ws[f'I{last_row}'].alignment = Alignment(horizontal='right')   # Align Total_Current_Charges

    # Style the entire worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=12):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
            if isinstance(cell.value, (int, float)) and cell.column == 9:  # Check for Bill Amount column
                cell.alignment = Alignment(horizontal='right')

    # Adjust column widths safely by skipping merged cells
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create HTTP response with the Excel file
    filename = f"{company.replace(' ', '_')}_Entered_Bills_Report.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response




def organization_location_listing_report(request):
    # Extract query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')

    print(company,sub_company)

    # Validate input parameters
    if not company or not sub_company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    # Query the database for the filtered data
    try:
        filtered_data = Location.objects.filter(
            Q(company_name=company) & Q(sub_company_name=sub_company)
        )

        # Convert the filtered data to a list of dictionaries
        data_list = list(filtered_data.values())

        return JsonResponse(data_list, safe=False)

    except Exception as e:
        return JsonResponse({'message': str(e)}, status=500)



def organization_location_listing_report_download(request):
    # Get query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')

    # Validate required parameters
    if not company or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    # Filter data from the BaseDataTable
    filtered_data = Location.objects.filter(
        company_name=company,
        sub_company_name=sub_company
    )

    # Convert the data to a list of dictionaries
    data_list = list(filtered_data.values())

    if not data_list:
        return HttpResponse("No data found for the given parameters.", status=404)

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Organization Location Listing'

    # Main heading
    main_heading = 'Organization Location Listing Report'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        "Location", "Site Name", "Location Code", "Location Address",
        "City", "State", "Zip", "Phone Number"
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')  # Fixed border style
        )

    # Populate data rows
    for item in data_list:
        row = [
            item.get("site_name", ""),
            item.get("site_name", ""),
            "",
            item.get("physical_address", ""),
            item.get("physical_city", ""),
            item.get("physical_State", ""),
            item.get("physical_zip_code", ""),
            item.get("main_phone", "")
        ]
        ws.append(row)

    # Style the entire worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')  # Fixed border style
            )
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_Organization_Location_Listing_Report.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response







def organization_location_report(request):
    # Extract query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')

    print(company,sub_company)

    # Validate input parameters
    if not company or not sub_company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    # Query the database for the filtered data
    try:
        filtered_data = Location.objects.filter(
            Q(company_name=company) & Q(sub_company_name=sub_company)
        )

        # Convert the filtered data to a list of dictionaries
        data_list = list(filtered_data.values())

        return JsonResponse(data_list, safe=False)

    except Exception as e:
        return JsonResponse({'message': str(e)}, status=500)



def organization_location_report_download(request):
    # Get query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')

    # Validate required parameters
    if not company or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    # Filter data from the BaseDataTable
    filtered_data = Location.objects.filter(
        company_name=company,
        sub_company_name=sub_company
    )

    # Convert the data to a list of dictionaries
    data_list = list(filtered_data.values())

    if not data_list:
        return HttpResponse("No data found for the given parameters.", status=404)

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Organization Location Report'

    # Main heading
    main_heading = 'Organization Location Report'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        "Division", "Site Name", "Location Address", "Vendor",
         "Phone Number"
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')  # Fixed border style
        )

    # Populate data rows
    for item in data_list:
        row = [
            item.get("division", ""),
            item.get("site_name", ""),
            item.get("physical_address", ""),
            item.get("vendor", ""),
            item.get("main_phone", ""),
        ]
        ws.append(row)

    # Style the entire worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')  # Fixed border style
            )
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_Organization_Location_Report.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response




def inactive_location_report(request):
    # Extract query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')

    print(company,sub_company)

    # Validate input parameters
    if not company or not sub_company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    # Query the database for the filtered data
    try:
        filtered_data = Location.objects.filter(
    company_name=company,
    sub_company_name=sub_company).filter(Q(location_type="Closed") | Q(location_type="Inactive Site"))

        # Convert the filtered data to a list of dictionaries
        data_list = list(filtered_data.values())

        return JsonResponse(data_list, safe=False)

    except Exception as e:
        return JsonResponse({'message': str(e)}, status=500)



def inactive_location_report_download(request):
    # Get query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')

    # Validate required parameters
    if not company or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    # Filter data from the BaseDataTable
    filtered_data = Location.objects.filter(
    company_name=company,
    sub_company_name=sub_company).filter(Q(location_type="Closed") | Q(location_type="Inactive Site"))

    # Convert the data to a list of dictionaries
    data_list = list(filtered_data.values())

    if not data_list:
        return HttpResponse("No data found for the given parameters.", status=404)

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inactive or Closed Location Report'

    # Main heading
    main_heading = 'Inactive or Closed Location Report'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        "Division", "Site Name", "Location Address", "Vendor",
         "Phone Number"
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')  # Fixed border style
        )

    # Populate data rows
    for item in data_list:
        row = [
            item.get("division", ""),
            item.get("site_name", ""),
            item.get("physical_address", ""),
            item.get("vendor", ""),
            item.get("main_phone", ""),
        ]
        ws.append(row)

    # Style the entire worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')  # Fixed border style
            )
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_Inactive/Closed_Location_Report.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response







def location_filter_report(request):
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    location_status = request.GET.get('location_status')

    filters = {
        'company_name': company,
        'sub_company_name': sub_company,
    }

    if location_status:
        filters['location_type'] = location_status

    locations = Location.objects.filter(**filters)
    data = list(locations.values())
    return JsonResponse(data, safe=False)




def organization_ban_listing_by_location(request):
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    ban_status = request.GET.get('ban_status')
    location = request.GET.get('location')


    filters = {
        'company': company,
        'sub_company': sub_company,
        'Ban_status' : ban_status
    }

    if location:
        filters['location'] = location

    locations = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(**filters)
    data = list(locations.values())
    return JsonResponse(data, safe=False)



def organization_ban_listing_by_location_download(request):
    # Get query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    location = request.GET.get('location')
    ban_status = request.GET.get('ban_status')

    # Validate required parameters
    if not company or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    # Query the database
    try:
        filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
            company=company,
            sub_company=sub_company
        )
        if location:
            filtered_data = filtered_data.filter(location=location)
        if ban_status:
            filtered_data = filtered_data.filter(Ban_status=ban_status)

        # Convert queryset to list of dictionaries
        data_list = list(filtered_data.values())

        if not data_list:
            return HttpResponse("No data found for the given parameters.", status=404)

    except Exception as e:
        return HttpResponse(f"Error fetching data: {str(e)}", status=500)

    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Organization BAN Listing"

    # Main heading
    main_heading = "Organization BAN Listing By Location"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        "Location Name", "Site", "Location Alias", "Division", "Address",
        "BAN", "Cost Center", "Service Type", "Customer of Record",
        "Phone #(BTN/WTN)", "Phone #(DID)", "Master Account"
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in data_list:
        cost_centers = item.get('cost_centers', [])
        cost_center_names = ", ".join(
            [cc.get('cost_center_name', '') for cc in cost_centers if isinstance(cc, dict)]
        ) or "N/A"

        row = [
            item.get("location", ""),
            item.get("location", ""),
            item.get("location_alias", ""),
            item.get("division", ""),
            item.get("Remidence_Address", ""),
            item.get("accountnumber", ""),
            cost_center_names,
            item.get("services", ""),
            item.get("customer_of_Record", ""),
            "",  # Placeholder for BTN/WTN
            "",  # Placeholder for DID
            item.get("master_account", ""),
        ]
        ws.append(row)

    # Style the entire worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=12):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create an HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_BAN_Listing_By_Location.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    with BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.read())

    return response







def organization_ban_listing_by_vendor(request):
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    ban_status = request.GET.get('ban_status')
    location = request.GET.get('location')


    filters = {
        'company': company,
        'sub_company': sub_company,
        'Ban_status' : ban_status
    }

    if location:
        filters['location'] = location

    locations = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(**filters)
    data = list(locations.values())
    return JsonResponse(data, safe=False)



def organization_ban_listing_by_vendor_download(request):
    # Get query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    location = request.GET.get('location')
    ban_status = request.GET.get('ban_status')

    # Validate required parameters
    if not company or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    # Query the database
    try:
        filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
            company=company,
            sub_company=sub_company
        )
        if location:
            filtered_data = filtered_data.filter(location=location)
        if ban_status:
            filtered_data = filtered_data.filter(Ban_status=ban_status)

        # Convert queryset to list of dictionaries
        data_list = list(filtered_data.values())

        if not data_list:
            return HttpResponse("No data found for the given parameters.", status=404)

    except Exception as e:
        return HttpResponse(f"Error fetching data: {str(e)}", status=500)

    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Organization BAN Listing"

    # Main heading
    main_heading = "Organization BAN Listing By Vendor"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        "Vendor", "BAN", "Service Type", "Site", "Address",
        "CS Phone Number"
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in data_list:
        cost_centers = item.get('cost_centers', [])
        cost_center_names = ", ".join(
            [cc.get('cost_center_name', '') for cc in cost_centers if isinstance(cc, dict)]
        ) or "N/A"

        row = [
            item.get("vendor", ""),
            item.get("accountnumber", ""),
            item.get("services", ""),
            item.get("location", ""),
            item.get("Remidence_Address", ""),
            item.get("cs_number", ""),
            
        ]
        ws.append(row)

    # Style the entire worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create an HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_BAN_Listing_By_Vendor.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    with BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.read())

    return response





def organization_cost_centers_by_account(request):
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    ban_status = request.GET.get('ban_status')
    location = request.GET.get('location')


    filters = {
        'company': company,
        'sub_company': sub_company,
        'Ban_status' : ban_status
    }

    if location:
        filters['location'] = location

    locations = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(**filters)
    data = list(locations.values())
    return JsonResponse(data, safe=False)



def organization_cost_centers_by_account_download(request):
    # Get query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    location = request.GET.get('location')
    ban_status = request.GET.get('ban_status')

    # Validate required parameters
    if not company or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    # Query the database
    try:
        filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
            company=company,
            sub_company=sub_company
        )
        if location:
            filtered_data = filtered_data.filter(location=location)
        if ban_status:
            filtered_data = filtered_data.filter(Ban_status=ban_status)

        # Convert queryset to list of dictionaries
        data_list = list(filtered_data.values())

        if not data_list:
            return HttpResponse("No data found for the given parameters.", status=404)

    except Exception as e:
        return HttpResponse(f"Error fetching data: {str(e)}", status=500)

    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Organization Cost Centers"

    # Main heading
    main_heading = "Organization Cost Centers By Account"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        "BAN",
        "Vendor", "Cost Center"
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in data_list:
        cost_centers = item.get('cost_centers', [])
        cost_center_names = ", ".join(
            [cc.get('cost_center_name', '') for cc in cost_centers if isinstance(cc, dict)]
        ) or "N/A"

        row = [
            item.get("accountnumber", ""),
            item.get("vendor", ""),
            cost_center_names,
           
        ]
        ws.append(row)

    # Style the entire worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create an HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_Cost_Centers_By_Account.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    with BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.read())

    return response





def inventory_report(request):
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    ban_status = request.GET.get('ban_status')
    location = request.GET.get('location')
    vendor = request.GET.get('vendor')



    filters = {
        'company': company,
        'sub_company': sub_company,
        'Ban_status' : ban_status,
        'vendor' : vendor
    }

    if location:
        filters['location'] = location

    locations = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(**filters)
    data = list(locations.values())
    return JsonResponse(data, safe=False)



def inventory_report_download(request):
    # Get query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    location = request.GET.get('location')
    ban_status = request.GET.get('ban_status')
    vendor = request.GET.get('vendor')

    # Validate required parameters
    if not company or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    # Query the database
    try:
        filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
            company=company,
            sub_company=sub_company
        )
        if location:
            filtered_data = filtered_data.filter(location=location)
        if vendor:
            filtered_data = filtered_data.filter(vendor=vendor)
        if ban_status:
            filtered_data = filtered_data.filter(Ban_status=ban_status)

        # Convert queryset to list of dictionaries
        data_list = list(filtered_data.values())

        if not data_list:
            return HttpResponse("No data found for the given parameters.", status=404)

    except Exception as e:
        return HttpResponse(f"Error fetching data: {str(e)}", status=500)

    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Organization Inventory Report"

    # Main heading
    main_heading = "Organization Inventory Report"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        "Loc. Name", "Vendor", "BAN", "GL Code", "Ban Status", "Service Type", "Service Qty", "Phone# (BTN/WTN)", "Phone# DID", "Data Speeds", "Last Invoice Processed On" 
    ]
    										
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in data_list:
        cost_centers = item.get('cost_centers', [])
        cost_center_names = ", ".join(
            [cc.get('cost_center_name', '') for cc in cost_centers if isinstance(cc, dict)]
        ) or "N/A"

        row = [
            item.get("location", ""),
            item.get("vendor", ""),
            item.get("accountnumber", ""),
            item.get("gl_code", ""),
            item.get("services", ""),
           
        ]
        ws.append(row)

    # Style the entire worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=11):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create an HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_Inventory_Report.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    with BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.read())

    return response





def circuit_list_by_organization(request):
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    
    filters = {
        'company': company,
        'sub_company': sub_company,}

    locations = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(**filters)
    data = list(locations.values())
    return JsonResponse(data, safe=False)



def circuit_list_by_organization_download(request):
    # Get query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
   

    # Validate required parameters
    if not company or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    # Query the database
    try:
        filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
            company=company,
            sub_company=sub_company
        )
        

        # Convert queryset to list of dictionaries
        data_list = list(filtered_data.values())

        if not data_list:
            return HttpResponse("No data found for the given parameters.", status=404)

    except Exception as e:
        return HttpResponse(f"Error fetching data: {str(e)}", status=500)

    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Circuit List Report"

    # Main heading
    main_heading = "Circuit List Report"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        "Location", "Services Used", "LEC Circuit Id", "Carrier Circuit Id"
    ]
    										
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in data_list:
        

        row = [
            item.get("location", ""),
            item.get("services", ""),
            "",
            "",           
        ]
        ws.append(row)

    # Style the entire worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create an HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_Circuit_List.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    with BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.read())

    return response








def circuit_list_by_organization_contracts(request):
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    
    filters = {
        'company': company,
        'sub_company': sub_company,}

    locations = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(**filters)
    data = list(locations.values())
    return JsonResponse(data, safe=False)



def circuit_list_by_organization_contracts_download(request):
    # Get query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
   

    # Validate required parameters
    if not company or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    # Query the database
    try:
        filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
            company=company,
            sub_company=sub_company
        )
        

        # Convert queryset to list of dictionaries
        data_list = list(filtered_data.values())

        if not data_list:
            return HttpResponse("No data found for the given parameters.", status=404)

    except Exception as e:
        return HttpResponse(f"Error fetching data: {str(e)}", status=500)

    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Circuit List Contracts Report"

    # Main heading
    main_heading = "Circuit List By Organization Contracts Report"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
       "Contract Number", "Contract Term", "Vendor", "Account Number", "Site", "Service", "LEC Circuit Id", "Carrier Circuit Id"
    ]
    										
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in data_list:
        

        row = [
            item.get("contract_number", ""),
            item.get("contract_term", ""),
            item.get("vendor", ""),
            item.get("accountnumber", ""),
            item.get("location", ""),
            item.get("services", ""),
            "",
            "",           
        ]
        ws.append(row)

    # Style the entire worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create an HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_Circuit_List_Contracts.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    with BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.read())

    return response





def carrier_pic_report(request):
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
    
    filters = {
        'company': company,
        'sub_company': sub_company,}

    locations = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(**filters)
    data = list(locations.values())
    return JsonResponse(data, safe=False)



def carrier_pic_report_download(request):
    # Get query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')
   

    # Validate required parameters
    if not company or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    # Query the database
    try:
        filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
            company=company,
            sub_company=sub_company
        )
        

        # Convert queryset to list of dictionaries
        data_list = list(filtered_data.values())

        if not data_list:
            return HttpResponse("No data found for the given parameters.", status=404)

    except Exception as e:
        return HttpResponse(f"Error fetching data: {str(e)}", status=500)

    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Carrier PIC Report"

    # Main heading
    main_heading = "Carrier PIC Report"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
       "Site", "Address", "Account Number", "PICC", "LPICC"
    ]
    										
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )

    # Populate data rows
    for item in data_list:
        

        row = [
            item.get("location", ""),
            item.get("Remidence_Address", ""),
            item.get("accountnumber", ""),
            "",
            "",           
        ]
        ws.append(row)

    # Style the entire worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create an HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_Carrier_PIC_Report.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    with BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.read())

    return response





def inactive_accounts_report(request):
    # Extract query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')

    print(company,sub_company)

    # Validate input parameters
    if not company or not sub_company:
        return JsonResponse({'message': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

    # Query the database for the filtered data
    try:
        filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
    company=company,
    sub_company=sub_company).filter(Q(Ban_status="Closed") | Q(Ban_status="Inactive"))

        # Convert the filtered data to a list of dictionaries
        data_list = list(filtered_data.values())

        return JsonResponse(data_list, safe=False)

    except Exception as e:
        return JsonResponse({'message': str(e)}, status=500)



def inactive_accounts_report_download(request):
    # Get query parameters
    company = request.GET.get('company')
    sub_company = request.GET.get('sub_company')

    # Validate required parameters
    if not company or not sub_company:
        return HttpResponse("Missing required parameters", status=status.HTTP_400_BAD_REQUEST)

    # Filter data from the BaseDataTable
    filtered_data = BaseDataTable.objects.exclude(viewuploaded=None, viewpapered=None).filter(
    company=company,
    sub_company=sub_company).filter(Q(Ban_status="Closed") | Q(Ban_status="Inactive"))

    # Convert the data to a list of dictionaries
    data_list = list(filtered_data.values())

    if not data_list:
        return HttpResponse("No data found for the given parameters.", status=404)

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inactive or Closed Accounts Report'

    # Main heading
    main_heading = 'Inactive or Closed Accounts Report'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws['A1'] = main_heading
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers
    headers = [
        "Division", "Site Name", "Location Address", "Vendor",
         "Phone Number"
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0000FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')  # Fixed border style
        )

    # Populate data rows
    for item in data_list:
        row = [
            item.get("division", ""),
            item.get("site_name", ""),
            item.get("physical_address", ""),
            item.get("vendor", ""),
            item.get("main_phone", ""),
        ]
        ws.append(row)

    # Style the entire worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')  # Fixed border style
            )
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create HTTP response with the Excel file
    filename = f"{sub_company.replace(' ', '_')}_Inactive/Closed_Accounts_Report.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
        