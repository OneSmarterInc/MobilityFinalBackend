from django.shortcuts import render
from django.shortcuts import render, HttpResponse
from rest_framework.views import APIView
from rest_framework import status
from rest_framework.response import Response
from rest_framework import permissions
from authenticate.views import saveuserlog
from OnBoard.Organization.models import Organizations
from OnBoard.Company.models import Company
from Dashboard.ModelsByPage.DashAdmin import Vendors
from OnBoard.Ban.models import UploadBAN, BaseDataTable
from .ser import OrganizationShowSerializer, VendorShowSerializer, showBanSerializer
from .ser import saveBilledSerializer
from .models import Report_Billed_Data, Report_Unbilled_Data
import pandas as pd
from OnBoard.Ban.models import UniquePdfDataTable
from datetime import datetime
from .ser import showBilledReport, showUnbilledReport
from django.db.models.functions import TruncMonth

class UploadFileView(APIView):

    permission_classes = [permissions.IsAuthenticated]

    def __init__(self, **kwargs):
        self.current_year = datetime.now().year
        self.data = None
        super().__init__(**kwargs)

    def get(self, request,report_type=None, *args, **kwargs):
        if request.user.designation.name == 'Admin':
            orgs = OrganizationShowSerializer(Organizations.objects.filter(company=request.user.company), many=True)
            vendors = VendorShowSerializer(Vendors.objects.all(), many=True)
            bans = showBanSerializer(BaseDataTable.objects.filter(company=request.user.company, viewuploaded=None), many=True)
        else:
            orgs = OrganizationShowSerializer(Organizations.objects.all(), many=True)
            vendors = VendorShowSerializer(Vendors.objects.all(), many=True)
            bans = showBanSerializer(BaseDataTable.objects.filter(company=request.user.company, viewuploaded=None), many=True)
        company = Company.objects.get(Company_name=request.user.company.Company_name) if request.user.company else None
        filter_kwargs = {}
        if company:
            filter_kwargs['company'] = company
        if report_type:
            filter_kwargs['Report_Type'] = report_type
        if report_type and report_type == "Unbilled_Data_Usage":
            filtered_data = Report_Unbilled_Data.objects.filter(**filter_kwargs)
            unique_dates = {}
            for row in filtered_data:
                if row.Date not in unique_dates:
                    unique_dates[row.Date] = row
            unique_rows = list(unique_dates.values())
            self.data = showUnbilledReport(unique_rows, many=True)
        elif report_type and report_type == "Billed_Data_Usage":
            filtered_data = Report_Billed_Data.objects.filter(**filter_kwargs)
            unique_dates = {}
            for row in filtered_data:
                if row.Month not in unique_dates:
                    unique_dates[row.Month] = row
            unique_rows = list(unique_dates.values())
            self.data = showBilledReport(unique_rows, many=True)
        return Response(
            {"orgs": orgs.data, "vendors": vendors.data, "bans":bans.data, "data":self.data.data if self.data else None},
            status=status.HTTP_200_OK,
        )
    def post(self, request, *args, **kwargs):
        data = request.data
        org = Organizations.objects.filter(Organization_name=data['sub_company'])[0]
        company = Company.objects.filter(Company_name=org.company.Company_name)[0]
        vendor = Vendors.objects.filter(name=data['vendor'])[0]
        file = data['file']
        account_number = data['account_number']
        month = data['month']
        report_type = data['report_type']
        if report_type == 'Billed_Data_Usage':
            try:
                try:
                    if file.name.endswith('.csv'):
                        df = pd.read_csv(file)
                    elif file.name.endswith('.xlsx'):
                        df = pd.read_excel(file)
                    elif file.name.endswith('.txt'):
                        df = pd.read_csv(file, delimiter='\t')
                    else:
                        return Response({"message": "Unsupported file format"}, status=status.HTTP_400_BAD_REQUEST)
                except Exception as e:

                    return Response({"message": str(e)}, status=status.HTTP_400_BAD_REQUEST)
                column_mappings = {
                    'Wireless Number': 'wireless_number',
                    'User Name': 'user_name',
                    'Voice Plan Usage': 'voice_plan_usage',
                    'Messaging Usage': 'messaging_usage',
                    'Data Usage (GB)': 'data_usage_gb',
                    'Data Usage (KB)': 'data_usage_kb',
                    'Data Usage (MB)': 'data_usage_mb',
                    'Bill Cycle Date': 'Bill_Cycle_Date '
                }
                actual_columns = df.columns.tolist()
                mapped_columns = {v: k for k, v in column_mappings.items() if k in actual_columns}
                required_columns = ['wireless_number', 'user_name', 'voice_plan_usage', 'messaging_usage','Bill_Cycle_Date ']
                if not all(col in mapped_columns for col in required_columns):
                    return Response({"message": "File is missing required columns"}, status=status.HTTP_200_OK)

                data_usage_column = mapped_columns.get('data_usage_gb')
                if not data_usage_column:
                    data_usage_column = mapped_columns.get('data_usage_kb') or mapped_columns.get('data_usage_mb')

                if not data_usage_column:
                    return Response({"message": "File is missing data usage columns"}, status=status.HTTP_200_OK)

                # Extract necessary data
                existing_data = Report_Billed_Data.objects.filter(
                    company=company,
                    organization=org,
                    vendor=vendor,
                    Account_Number=account_number,
                    Month=month,
                    Year=str(self.current_year)
                ).exists()

                if existing_data:
                    return Response(
                        {"message": "Report Billed Data with provided details already exists"},
                        status=status.HTTP_409_CONFLICT,
                    )
                for _, row in df.iterrows():
                    wireless_number = row.get(mapped_columns['wireless_number'], '')
                    user_name = row.get(mapped_columns['user_name'], '')
                    Bill_Cycle_Date = row.get(mapped_columns['Bill_Cycle_Date '], '')
                    voice_plan_usage = row.get(mapped_columns['voice_plan_usage'], '')
                    messaging_usage = row.get(mapped_columns['messaging_usage'], '')
                    data_usage_gb = row.get(mapped_columns.get('data_usage_gb'), None)
                    data_usage_kb = row.get(mapped_columns.get('data_usage_kb'), None)
                    data_usage_mb = row.get(mapped_columns.get('data_usage_mb'), None)

                    if data_usage_gb is None and data_usage_kb is not None:
                        data_usage_gb = data_usage_kb / 1024 / 1024  # Convert KB to GB
                    elif data_usage_gb is None and data_usage_mb is not None:
                        data_usage_gb = data_usage_mb / 1024  # Convert MB to GB

                    # if data_usage_gb is None:
                        # data_usage_gb = 0  # Handle cases where data usage columns are missing or empty

                    if wireless_number and user_name:
                        Report_Billed_Data.objects.create(
                            company=company,
                            organization=org,
                            vendor=vendor,
                            Account_Number=account_number,
                            Wireless_Number=wireless_number,
                            User_Name=user_name,
                            Report_Type=report_type,
                            Month=month,
                            Voice_Plan_Usage=voice_plan_usage,
                            Messaging_Usage=messaging_usage,
                            Data_Usage_GB=data_usage_gb,
                            File=file,
                            Bill_Cycle_Date = Bill_Cycle_Date,
                            Year=str(self.current_year)  
                        )
                saveuserlog(
                        request.user, 
                        f"Billed Data Usage Report uploaded of file {file.name} and Vendor: {vendor.name}"
                    )
                return Response({
                    "message": "Billed Data Usage Report uploaded successfully",
                }, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        elif report_type == 'Unbilled_Data_Usage':
            try:
                date = data['date']
                try:
                    date = datetime.strptime(date, '%Y-%m-%d')
                    month_start = datetime(date.year, date.month, 1)
                    week_number = (date - month_start).days // 7 + 1
                except ValueError:
                    return Response({"message": "Invalid date format. Expected format is YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)
                try:
                    if file.name.endswith('.csv'):
                        df = pd.read_csv(file)
                        file_format = 'csv'
                    elif file.name.endswith('.xlsx'):
                        df = pd.read_excel(file)
                        file_format = 'xlsx'
                    elif file.name.endswith('.txt'):
                        df = pd.read_csv(file, delimiter='\t')
                        file_format = 'txt'
                    else:
                        return Response({"message": "Unsupported file format"}, status=status.HTTP_400_BAD_REQUEST)
                except Exception as e:
                    return Response({"message": str(e)}, status=status.HTTP_400_BAD_REQUEST)
                for _, row in df.iterrows():
                    user_name = row.get('User name', '')
                    wireless_number = row.get('Wireless number', '')
                    data_usage_gb = row.get('Data Usage (GB)', None)

                    if data_usage_gb is None:
                        data_usage_kb = row.get('Data Usage (KB)', None)
                        if data_usage_kb is not None:
                            data_usage_gb = data_usage_kb / 1024 / 1024  # Convert KB to GB

                    if user_name and wireless_number and data_usage_gb is not None:
                        # Retrieve device and upgrade eligibility date from unique_pdf_data_table
                        unique_data = UniquePdfDataTable.objects.filter(wireless_number=wireless_number).first()
                        if unique_data:
                            device = unique_data.mobile_device
                            upgrade_eligibility_date = unique_data.upgrade_eligible_date
                        else:
                            device = 'NA'
                            upgrade_eligibility_date = 'NA'

                        # Check if the record already exists
                        record_exists = Report_Unbilled_Data.objects.filter(
                            company=company,
                            organization=org,
                            vendor=vendor,
                            Account_Number=account_number,
                            Wireless_Number=wireless_number,
                            User_Name=user_name,
                            Report_Type=report_type,
                            Month=month,
                            Date=date,
                        ).exists()

                        if record_exists:
                            return Response({"message": "Report Unbilled Data with provided details already exists"}, status=status.HTTP_409_CONFLICT)

                        # Create the new record
                        Report_Unbilled_Data.objects.create(
                            company=company,
                            organization=org,
                            vendor=vendor,
                            Account_Number=account_number,
                            Wireless_Number=wireless_number,
                            User_Name=user_name,
                            Report_Type=report_type,
                            Month=month,
                            Date=date,
                            Week=week_number,
                            Usage=data_usage_gb,
                            Device=device,
                            Upgrade_Eligibilty_Date=upgrade_eligibility_date,
                            File=file,
                            File_Format=file, 
                            Year=str(self.current_year)
                        )
                saveuserlog(
                        request.user, 
                        f"Unbilled Data Usage Report uploaded of file  {file.name} and Vendor: {vendor.name}"
                    )
                return Response({
                    "message": "Unbilled Data Usage Report uploaded successfully",
                }, status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"message": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({
                "message": "Invalid report type"
            }, status=status.HTTP_400_BAD_REQUEST)
        
    def put(self, request, pk, *args, **kwargs):
        pass
    def delete(self, request, report_type, pk, *args, **kwargs):
        self.com = None
        self.org = None
        self.vendor = None
        self.account_number = None
        self.month = None
        self.year = None
        self.report_type = report_type
        try:
            if report_type == 'Billed_Data_Usage':
                obj1 = Report_Billed_Data.objects.filter(id=pk)
                self.com = obj1[0].company.Company_name
                self.org = obj1[0].organization.Organization_name
                self.vendor = obj1[0].vendor.name
                self.account_number = obj1[0].Account_Number
                self.month = obj1[0].Month
                self.year = obj1[0].Year
                allobjs = Report_Billed_Data.objects.filter(
                    company=obj1[0].company,
                    organization=obj1[0].organization,
                    vendor=obj1[0].vendor,
                    Report_Type = obj1[0].Report_Type,
                    Month = obj1[0].Month,
                    Year=obj1[0].Year,
                    Account_Number=obj1[0].Account_Number,
                )
                try:
                    allobjs.delete()
                except Exception as e:
                    print(e)
                    return Response({"message": "Error while deleting related records"}, status=status.HTTP_400_BAD_REQUEST)
            elif report_type == 'Unbilled_Data_Usage':
                obj1 = Report_Unbilled_Data.objects.filter(id=pk)
                self.com = obj1[0].company.Company_name
                self.org = obj1[0].organization.Organization_name
                self.vendor = obj1[0].vendor.name
                self.account_number = obj1[0].Account_Number
                self.month = obj1[0].Month
                self.year = obj1[0].Year
                self.report_type = obj1[0].Report_Type
                allobjs = Report_Unbilled_Data.objects.filter(
                    company=obj1[0].company,
                    organization=obj1[0].organization,
                    vendor=obj1[0].vendor,
                    Report_Type = obj1[0].Report_Type,
                    Account_Number=obj1[0].Account_Number,
                    Month=obj1[0].Month,
                    Year=obj1[0].Year,
                )
                try:
                    allobjs.delete()
                except Exception as e:
                    print(e)
                    return Response({"message": "Error while deleting related records"}, status=status.HTTP_400_BAD_REQUEST)
            else:
                return Response({"message": "Invalid report type"}, status=status.HTTP_400_BAD_REQUEST)
            saveuserlog(
                request.user, 
                f"{self.report_type} report attributions company:{self.com}, organization:{self.org}, vendor:{self.vendor}, month:{self.month}, year:{self.year}. account number:{self.account_number}, month:{self.month}, and year:{self.year}"
            )
            return Response({"message": f"{report_type} Report deleted successfully"}, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"message": str(e)}, status=status.HTTP_400_BAD_REQUEST)

from django.db.models.functions import Cast
from django.db.models import FloatField

import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles import Font, Alignment, PatternFill
from django.db.models import Q
from django.core.files.base import ContentFile
from openpyxl import Workbook

from .models import Downloaded_reports
from collections import defaultdict
class ViewReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def __init__(self, **kwargs):
        self.current_year = datetime.now().year
        self.data = None
        self.unique_dates = None
        super().__init__(**kwargs)
    def get(self, request, report_type=None, sub_type=None, *args, **kwargs):
        if request.user.designation.name == 'Admin':
            orgs = OrganizationShowSerializer(Organizations.objects.filter(company=request.user.company), many=True)
            vendors = VendorShowSerializer(Vendors.objects.all(), many=True)
            bans = showBanSerializer(BaseDataTable.objects.filter(company=request.user.company, viewuploaded=None), many=True)
        else:
            orgs = OrganizationShowSerializer(Organizations.objects.all(), many=True)
            vendors = VendorShowSerializer(Vendors.objects.all(), many=True)
            bans = showBanSerializer(BaseDataTable.objects.filter(viewuploaded=None), many=True)

        if report_type and report_type == 'Billed_Data_Usage':
            if not sub_type:
                self.data = showBilledReport(Report_Billed_Data.objects.all(), many=True)
            elif sub_type == "Top10BilledUsers":
                company = request.GET.get('company')
                sub_company = request.GET.get('sub_company')
                report_type = request.GET.get('report_type')
                month = request.GET.get('month')
                vendor = request.GET.get('vendor')
                year = request.GET.get('year')
                filter_kwargs = {}
                if company:
                    filter_kwargs['company'] = Company.objects.filter(Company_name=company)[0]
                if sub_company:
                    filter_kwargs['organization'] = Organizations.objects.filter(Organization_name=sub_company)[0]
                if report_type:
                    filter_kwargs['Report_Type'] = report_type
                if month:
                    filter_kwargs['Month'] = month
                if year:
                    filter_kwargs['Year'] = year
                if vendor:
                    filter_kwargs['vendor'] = Vendors.objects.filter(name=vendor)[0]
                
                try:
                    filtered_data = Report_Billed_Data.objects.filter(**filter_kwargs).annotate(
                        usage_float=Cast('Data_Usage_GB', FloatField())
                    ).order_by('-usage_float')[:10]
                    self.data = showBilledReport(filtered_data, many=True)
                except Exception as e:
                    return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)   
            elif sub_type == "getZeroUsageReportbilledData":
                company = request.GET.get('company')
                sub_company = request.GET.get('sub_company')
                report_type = request.GET.get('report_type')
                month = request.GET.get('month')
                vendor = request.GET.get('vendor')
                year = request.GET.get('year')
                print(company, sub_company, report_type, vendor, year, month)
                filter_kwargs = {}
                if company:
                    filter_kwargs['company'] = Company.objects.filter(Company_name=company)[0]
                if sub_company:
                    filter_kwargs['organization'] = Organizations.objects.filter(Organization_name=sub_company)[0]
                if report_type:
                    filter_kwargs['Report_Type'] = report_type
                if month:
                    filter_kwargs['Month'] = month
                if year:
                    filter_kwargs['Year'] = year
                if vendor:
                    filter_kwargs['vendor'] = Vendors.objects.filter(name=vendor)[0]
                print(filter_kwargs)
                try:
                    filtered_data = Report_Billed_Data.objects.filter(
                        Q(Data_Usage_GB="0") | Q(Data_Usage_GB="0.0"),
                        **filter_kwargs
                    )

                    if not filtered_data.exists():
                        return Response({"message": "No data found for the given filters."}, status=status.HTTP_404_NOT_FOUND)

                    self.data = showBilledReport(filtered_data, many=True)

                except Exception as e:
                    return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        elif report_type and report_type == 'Unbilled_Data_Usage':
            if not sub_type:
                filter_kwargs = {}
                company = Company.objects.get(Company_name=request.user.company.Company_name) if request.user.company else None
                if company:
                    filter_kwargs['company'] = company
                if report_type:
                    filter_kwargs['Report_Type'] = report_type
                filtered_data = Report_Unbilled_Data.objects.filter(**filter_kwargs)

                report_data = []
                for record in filtered_data:
                    existing_report = next((r for r in report_data if r['Wireless_Number'] == record.Wireless_Number), None)
                    if not existing_report:
                        report_data.append({
                            "Account_Number": record.Account_Number,
                            'Wireless_Number': record.Wireless_Number,
                            'user_Name': record.User_Name,
                            'vendor': record.vendor,
                            record.Date: record.Usage,
                        })
                    else:
                        existing_report[record.Date] = record.Usage

                self.unique_dates = sorted(
                    list(set(
                        datetime.strptime(record.Date, '%Y-%m-%d %H:%M:%S').date()
                        for record in filtered_data
                    )),
                    key=lambda date: date
                )
                print(self.unique_dates)
                self.data = showUnbilledReport(Report_Unbilled_Data.objects.all(), many=True)

            elif sub_type == "getTop10byWeek":

                company = request.GET.get('company')
                sub_company = request.GET.get('sub_company')
                report_type = request.GET.get('report_type')
                month = request.GET.get('month')
                vendor = request.GET.get('vendor')
                year = request.GET.get('year')
                week = request.GET.get('week')
                print("week", week)
                filter_kwargs = {}
                if company:
                    filter_kwargs['company'] = Company.objects.filter(Company_name=company)[0]
                if sub_company:
                    filter_kwargs['organization'] = Organizations.objects.filter(Organization_name=sub_company)[0]
                if report_type:
                    filter_kwargs['Report_Type'] = report_type
                if month:
                    filter_kwargs['Month'] = month
                if year:
                    filter_kwargs['Year'] = year
                if vendor:
                    filter_kwargs['vendor'] = Vendors.objects.filter(name=vendor)[0]
                if week:
                    filter_kwargs['Week'] = week
                
                
                try:
                    filtered_data = Report_Unbilled_Data.objects.filter(**filter_kwargs)

                    friday_date = None
                    thursday_date = None
                    monday_date = None

                    for record in filtered_data:
                        date_obj = datetime.strptime(record.Date, '%Y-%m-%d %H:%M:%S').date()
                        if date_obj.weekday() == 4:  # Friday
                            friday_date = date_obj
                        elif date_obj.weekday() == 3 and not thursday_date:  # Thursday
                            thursday_date = date_obj
                        elif date_obj.weekday() == 0 and not monday_date:  # Monday
                            monday_date = date_obj

                    selected_date = friday_date or thursday_date or monday_date
                    if not selected_date:
                        return Response({"message": "No valid dates found for the specified week"}, status=status.HTTP_404_NOT_FOUND)

                    top_data = filtered_data.filter(Date=selected_date).annotate(
                        usage_float=Cast('Usage', FloatField())
                    ).order_by('-usage_float')[:10]

                    self.data = showUnbilledReport(top_data, many=True)

                except Exception as e:
                    print(e)
                    return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            elif sub_type == "getZeroUsageReportUnbilledData":
                company = request.GET.get('company')
                sub_company = request.GET.get('sub_company')
                month = request.GET.get('month')
                filter_kwargs = {}
                if company:
                    filter_kwargs['company'] = Company.objects.filter(Company_name=company)[0]
                if sub_company:
                    filter_kwargs['organization'] = Organizations.objects.filter(Organization_name=sub_company)[0]
                if report_type:
                    filter_kwargs['Report_Type'] = report_type
                if month:
                    filter_kwargs['Month'] = month

                try:
                    filtered_data = Report_Unbilled_Data.objects.filter(**filter_kwargs)

                    weeks = defaultdict(set)  
                    for record in filtered_data:
                        week = record.Week
                        date = record.Date
                        weeks[week].add(date)

            
                    self.data = showUnbilledReport(filtered_data, many=True)

                
                except Exception as e:
                    return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            

                
        return Response(
            {"orgs": orgs.data, "vendors": vendors.data, "bans":bans.data, "data":self.data.data if self.data else None, "unique_dates":self.unique_dates if report_type == "Unbilled_Data_Usage" else None},
            status=status.HTTP_200_OK,
        )
    
    def change_kwargs(self,kwargs):
        if 'organization' in kwargs and str(kwargs['organization']) != str:
            kwargs['organization'] = kwargs['organization'].Organization_name
        if 'company' in kwargs and str(kwargs['company']) != str:
            kwargs['company'] = kwargs['company'].Company_name
        if 'vendor' in kwargs and str(kwargs['vendor']) != str:
            kwargs['vendor'] = kwargs['vendor'].name

        return kwargs
    def post(self, request, *args, **kwargs):
        filter_kwargs = {}
        action = request.data.get('action')
        type = request.data.get('type')
        company = request.data.get('company')
        sub_company = request.data.get('sub_company')
        month = request.data.get('month')
        report_type = request.data.get('report_type')
        year = request.data.get('year')
        vendor = request.data.get('vendor')

        if company:
            filter_kwargs['company'] = Company.objects.filter(Company_name=company)[0]
        if sub_company:
            filter_kwargs['organization'] = Organizations.objects.filter(Organization_name=sub_company)[0]
        if report_type:
            filter_kwargs['Report_Type'] = report_type
        if month:
            filter_kwargs['Month'] = month
        if year:
            filter_kwargs['Year'] = year
        if vendor:
            filter_kwargs['vendor'] = Vendors.objects.filter(name=vendor)[0]
        print(action)
        print(type)
        if action == 'download-excel':
            if type == 'viewReportBilledExcel':
                try:
                    # Filter the data
                    filtered_data = Report_Billed_Data.objects.filter(**filter_kwargs)
                    print(filter_kwargs)
                    if not filtered_data.exists():
                        return Response({"message": "No data found for the given filters."}, status=200)

                    # Create a list of dictionaries to store the data in the required format
                    report_data = []
                    for record in filtered_data:
                        report_data.append({
                            "accountNumber": record.Account_Number,
                            'wirelessNumber': record.Wireless_Number,
                            'userName': record.User_Name,
                            'vendor': record.vendor.name,
                            'voicePlanUsage': float(record.Voice_Plan_Usage),  # New column
                            'messagingUsage': float(record.Messaging_Usage),  # New column
                            'usage': float(record.Data_Usage_GB),
                        })

                    # Convert the data to a pandas DataFrame
                    df = pd.DataFrame(report_data)

                    # Categorize data usage into specified ranges
                    bins = [0, 0.001, 1, 5, 7.5, 12.5, 22, float('inf')]
                    labels = ['Equal to 0', 'Less than 1 GB', '1-5 GB', '5-7.5 GB', '7.5-12.5 GB', '12.5-22 GB', '22 GB plus']
                    df['usage_category'] = pd.cut(df['usage'], bins=bins, labels=labels, right=False)

                    # Reorder columns and include new fields, removing 'date'
                    df = df[['accountNumber', 'wirelessNumber', 'userName', 'vendor', 'voicePlanUsage', 'messagingUsage', 'usage', 'usage_category']]

                    # Generate a bar graph for data usage categories
                    if vendor:
                        df_vendor = df[df['vendor'] == vendor]
                    else:
                        df_vendor = df

                    usage_counts = df_vendor['usage_category'].value_counts().reindex(labels, fill_value=0)
                    plt.figure(figsize=(12, 6))
                    usage_counts.plot(kind='bar')
                    plt.title(f'Data Usage Frequency for Vendor: {vendor}')
                    plt.xlabel('Data Usage Category')
                    plt.ylabel('Frequency')
                    plt.xticks(rotation=45, ha='right')
                    plt.tight_layout()  # Adjust the padding to fit the labels

                    # Save the plot to a BytesIO object
                    plot_output = BytesIO()
                    plt.savefig(plot_output, format='png')
                    plot_output.seek(0)

                    # Create an Excel writer object and write the DataFrame to it
                    output = BytesIO()
                    writer = pd.ExcelWriter(output, engine='openpyxl')
                    df.to_excel(writer, index=False, sheet_name='Billed Data Report')
                    writer.close()
                    output.seek(0)

                    # Load the workbook and insert the image
                    workbook = load_workbook(output)
                    worksheet = workbook['Billed Data Report']
                    img = Image(plot_output)
                    worksheet.add_image(img, 'K2')  # Adjust the position as needed

                    # Insert a dynamic heading
                    heading = f'{month} Usage Report'
                    worksheet.insert_rows(1)  # Shift everything down to make space for the header
                    worksheet.merge_cells('A1:H1')  # Adjust the merge to fit the new columns
                    heading_cell = worksheet['A1']
                    heading_cell.value = heading
                    heading_cell.font = Font(bold=True, color='FFFFFF', size=14)
                    heading_cell.fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')
                    heading_cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Define the background fill colors for 'usage' column based on conditions
                    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    light_yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
                    dark_navy_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")

                    # Define light blue fill for Voice_Plan_Usage, Messaging_Usage, and usage_category columns
                    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

                    # Apply specific formatting for the 'usage' column header (set text color to black)
                    usage_header_cell = worksheet["G2"]  # Assuming 'G2' is the header for the usage column
                    usage_header_cell.font = Font(bold=True, color="000000")  # Set usage header text to black

                    # Adjust column widths based on the content and header
                    for col_num, column_cells in enumerate(worksheet.columns, 1):
                        max_length = 0
                        column = get_column_letter(col_num)

                        for cell in column_cells:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))

                                # Apply background color and text formatting to specific columns
                                if column == 'G':  # Assuming 'G' is usage column
                                    usage_value = cell.value
                                    if isinstance(usage_value, float) or isinstance(usage_value, int):
                                        # Set content text color to white
                                        cell.font = Font(color="FFFFFF")
                                        # Apply conditional formatting based on usage values
                                        if usage_value == 0:
                                            cell.fill = red_fill
                                        elif 0 < usage_value <= 1:
                                            cell.fill = light_yellow_fill
                                        elif usage_value > 1:
                                            cell.fill = dark_navy_fill
                                elif column in ['E', 'F', 'H']:  # Voice_Plan_Usage, Messaging_Usage, usage_category columns
                                    cell.fill = light_blue_fill
                            except:
                                pass

                        adjusted_width = (max_length + 2) if max_length < 30 else 30
                        worksheet.column_dimensions[column].width = adjusted_width

                    # Save the modified workbook to the output
                    output_with_image = BytesIO()
                    workbook.save(output_with_image)
                    output_with_image.seek(0)
                    
                    report_file = Downloaded_reports(
                        report_type=report_type,
                        kwargs=self.change_kwargs(filter_kwargs)
                    )
                    report_file.file.save(f'billed_data_report_{month}_{year}.xlsx', ContentFile(output_with_image.getvalue()))
                    return Response({"data":report_file.file.url, "message" : "Excel file generated sucessfully!"}, status=status.HTTP_200_OK)

                except Exception as e:
                    print(e)
                    return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            elif type == "viewReportUnbilledExcel":
                try:
                    filtered_data = Report_Unbilled_Data.objects.filter(**filter_kwargs)

                    # Create a dictionary to store the data in the required format
                    report_data = []
                    week_dates = {}

                    for record in filtered_data:
                        week = int(record.Week)
                        date = record.Date
                        # Add dates for each week
                        if week not in week_dates:
                            week_dates[week] = set()
                        week_dates[week].add(date)

                        existing_report = next((r for r in report_data if r['wirelessNumber'] == record.Wireless_Number), None)
                        if not existing_report:
                            report_data.append({
                                "accountNumber": record.Account_Number,
                                'wirelessNumber': record.Wireless_Number,
                                'userName': record.User_Name,
                                'vendor': record.vendor.name,
                                record.Date: record.Usage,
                            })
                        else:
                            existing_report[record.Date] = record.Usage

                    # Sort the weeks and dates within each week
                    sorted_weeks = sorted(week_dates.keys())
                    sorted_week_dates = {week: sorted(dates) for week, dates in week_dates.items()}

                    # Flatten the sorted dates into a list of columns for the DataFrame
                    all_weeks_flat = []
                    for week in sorted_weeks:
                        all_weeks_flat.extend(sorted_week_dates[week])

                    # Convert the data to a pandas DataFrame
                    df = pd.DataFrame(report_data)
                    if not df.empty:
                        df = df[['accountNumber', 'wirelessNumber', 'userName', 'vendor'] + all_weeks_flat]

                    # Create an Excel writer object and write the DataFrame to it
                    output = BytesIO()
                    writer = pd.ExcelWriter(output, engine='openpyxl')
                    df.to_excel(writer, index=False, startrow=1, sheet_name='Unbilled Data Report')  # Start writing from row 2

                    # Get the workbook and worksheet objects
                    workbook = writer.book
                    worksheet = writer.sheets['Unbilled Data Report']

                    # Add Week 1, Week 2 headers in the first row (for merged cells)
                    current_col = 5  # Start after the first 4 columns (accountNumber, wirelessNumber, etc.)
                    for week_num in sorted_weeks:
                        num_dates_in_week = len(sorted_week_dates[week_num])
                        worksheet.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + num_dates_in_week - 1)
                        worksheet.cell(1, current_col).value = f"Week {week_num}"
                        worksheet.cell(1, current_col).alignment = Alignment(horizontal="center", vertical="center")
                        
                        current_col += num_dates_in_week  # Move to the next week's set of dates

                    # Adjust column widths based on the max length of each column (content, not headers)
                    for col_num, column in enumerate(df.columns, 1):
                        max_length = max(df[column].astype(str).map(len).max(), len(column))  # Find max content length
                        worksheet.column_dimensions[get_column_letter(col_num)].width = max_length + 2  # Add some padding

                    # Define specific colors for each week
                    week_colors = {
                        1: "FFFF99",  # Light Yellow for Week 1
                        2: "CCFFCC",  # Light Green for Week 2
                        3: "FFCCCC",  # Light Pink for Week 3
                        4: "CCCCFF",  # Light Blue for Week 4
                        5: "FFCC99",  # Light Orange for Week 5
                    }

                    # Apply the background color for each week uniformly across all columns for that week
                    current_col = 5
                    for week_num in sorted_weeks:
                        num_dates_in_week = len(sorted_week_dates[week_num])
                        fill_color = PatternFill(start_color=week_colors[week_num], end_color=week_colors[week_num], fill_type="solid")
                        
                        # Apply color to Week headers
                        worksheet.cell(1, current_col).fill = fill_color

                        for col_offset in range(num_dates_in_week):
                            for row in range(2, len(df) + 2):  # Start from row 2 (Actual data headers)
                                worksheet.cell(row=row, column=current_col + col_offset).fill = fill_color
                        
                        current_col += num_dates_in_week  # Move to the next week's set of dates

                    # Add blue background with white text to headers (second row)
                    header_fill = PatternFill(start_color="FF000066", end_color="FF000066", fill_type="solid")  # Blue
                    header_font = Font(color="FFFFFF")  # White text
                    for col_num in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(2, col_num)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    writer.close()
                    output.seek(0)
                    report_file = Downloaded_reports(
                        report_type=report_type,
                        kwargs=self.change_kwargs(filter_kwargs)
                    )
                    report_file.file.save(f'unbilled_data_report_{month}.xlsx', ContentFile(output.getvalue()))
                    return Response({"data":report_file.file.url, "message" : "Excel file generated sucessfully!"}, status=status.HTTP_200_OK)

                except Exception as e:
                    print(e)
                    return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            elif type == "getZeroUsageReportUnbilledExcel":
                try:
                    filtered_data = Report_Unbilled_Data.objects.filter(**filter_kwargs)
                    # Group data by week
                    week_data = defaultdict(list)
                    unique_vendors = set()
                    for record in filtered_data:
                        week_data[record.Week].append(record)
                        unique_vendors.add(record.vendor.name)  # Collect unique vendors

                    # Create an Excel workbook and worksheet
                    output = BytesIO()
                    workbook = Workbook()

                    # Apply formatting for titles and headers
                    title_font = Font(bold=True, size=12)
                    header_font = Font(bold=True, color="FFFFFF", size=14)
                    medium_orange_fill = PatternFill(start_color="FFFF8503", end_color="FFFF8503", fill_type="solid")  # Medium orange header
                    light_yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Light yellow for odd rows
                    light_grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light grey for even rows
                    main_header_fill = PatternFill(start_color="FF000066", end_color="FF000066", fill_type="solid")  # Dark blue bg for the main header

                    worksheet = workbook.active
                    worksheet.title = 'Zero Usage Report (Weekly)'

                    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)  # Merge across several columns
                    worksheet.cell(row=1, column=1).value = f"{month} Zero Usage Report (Weekly)"
                    worksheet.cell(row=1, column=1).font = header_font
                    worksheet.cell(row=1, column=1).fill = main_header_fill
                    worksheet.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

                    def get_valid_week_date(week_records):
                        week_dates = {'Monday': None, 'Wednesday': None, 'Friday': None}
                        for rec in week_records:
                            date_obj = datetime.strptime(rec.Date, '%Y-%m-%d')
                            weekday = date_obj.strftime('%A')
                            if weekday in week_dates and not week_dates[weekday]:
                                week_dates[weekday] = rec.Date
                        return week_dates['Monday'] or week_dates['Wednesday'] or week_dates['Friday']

                    start_column = 1
                    row_offset = 3  # Start from the 3rd row (after the main header)

                    for week, records in week_data.items():
                        selected_date = get_valid_week_date(records)

                        report_data = [
                            {'Wireless_Number': rec.Wireless_Number, 'User_Name': rec.User_Name, 'Device': rec.Device, 'Upgrade_Eligibility_Date': rec.Upgrade_Eligibilty_Date}
                            for rec in records
                        ]
                        df = pd.DataFrame(report_data)

                        # Write the title, week, and date
                        worksheet.merge_cells(start_row=row_offset, start_column=start_column, end_row=row_offset, end_column=start_column + 2)
                        worksheet.cell(row=row_offset, column=start_column).value = f"Week {week} - {selected_date}"
                        worksheet.cell(row=row_offset, column=start_column).font = title_font

                        # Show unique vendor in the next column to the date
                        worksheet.cell(row=row_offset, column=start_column + 3).value = ', '.join(unique_vendors)
                        worksheet.cell(row=row_offset, column=start_column + 3).font = title_font

                        # Write the headers
                        headers = ['Wireless_Number', 'User_Name', 'Device', 'Upgrade_Eligibility_Date']
                        for col_num, header in enumerate(headers, start=start_column):
                            cell = worksheet.cell(row=row_offset + 1, column=col_num)
                            cell.value = header
                            cell.fill = medium_orange_fill
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.alignment = Alignment(horizontal="center")

                        # Fill the data into the table with alternating row colors
                        for row_num, row_data in enumerate(df.itertuples(index=False), row_offset + 2):
                            fill_color = light_yellow_fill if row_num % 2 == 0 else light_grey_fill  # Alternate row colors
                            for col_num, value in enumerate(row_data, start=start_column):
                                cell = worksheet.cell(row=row_num, column=col_num)
                                cell.value = value
                                cell.fill = fill_color

                        # Adjust column widths dynamically based on content
                        for col_num in range(start_column, start_column + len(headers)):
                            max_len = max(len(str(cell.value)) for cell in worksheet[get_column_letter(col_num)])
                            worksheet.column_dimensions[get_column_letter(col_num)].width = max_len + 5

                        # Move start_column to the next set of columns for the next week's table
                        start_column += len(headers) + 3  # Shift the starting column by enough space to place tables side by side

                    # Save the modified workbook to the output stream
                    workbook.save(output)
                    output.seek(0)

                    report_file = Downloaded_reports(
                        report_type=report_type,
                        kwargs=self.change_kwargs(filter_kwargs)
                    )
                    report_file.file.save(f'zero_usage_report_{month}.xlsx', ContentFile(output.getvalue()))
                    return Response({"data": report_file.file.url, "message" : "Excel file generated sucessfully!"}, status=status.HTTP_200_OK)

                    # Create the HTTP response with the Excel file
                    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = f'attachment; filename=zero_usage_report_{month}.xlsx'

                    return response

                except Exception as e:
                    return Response({
                        'message': str(e),
                    }, status=status.HTTP_400_BAD_REQUEST)
            elif type == "getZeroUsageReportbilledExcel":
                try:
                    filtered_data = Report_Billed_Data.objects.filter(
                        Q(Data_Usage_GB="0") | Q(Data_Usage_GB="0.0"),
                        **filter_kwargs
                    )
                    print(filtered_data)
                    if not filtered_data.exists():
                        return Response({"message": "No data found for the given filters."}, status=status.HTTP_404_NOT_FOUND)

                    # Extract the unique vendor name from the filtered data
                    unique_vendor = filtered_data.values_list('vendor', flat=True).distinct()
                    if unique_vendor:
                        vendor_name = unique_vendor[0]  # Get the first unique vendor name
                    else:
                        vendor_name = "Unknown Vendor"  # Fallback in case there's no vendor

                    # Create a list of dictionaries to store the data in the required format
                    report_data = []
                    for record in filtered_data:
                        report_data.append({
                            "accountNumber": record.Account_Number,
                            'wirelessNumber': record.Wireless_Number,
                            'userName': record.User_Name,
                            'dataUsageGB': record.Data_Usage_GB,
                        })

                    # Convert the data to a pandas DataFrame
                    df = pd.DataFrame(report_data)

                    # Create an Excel writer object and write the DataFrame to it
                    output = BytesIO()
                    writer = pd.ExcelWriter(output, engine='openpyxl')
                    df.to_excel(writer, index=False, sheet_name='Zero Usage Billed Data Report')
                    writer.close()
                    output.seek(0)

                    # Load the workbook to apply formatting
                    workbook = load_workbook(output)
                    worksheet = workbook['Zero Usage Billed Data Report']

                    # Add the main heading for the report (dynamic based on month and vendor)
                    main_heading = f'{month} Zero Usage Report ({vendor_name})'
                    worksheet.insert_rows(1)
                    worksheet.merge_cells('A1:D1')  # Adjust the merge to cover the required columns
                    heading_cell = worksheet['A1']
                    heading_cell.value = main_heading
                    heading_cell.font = Font(bold=True, color="FFFFFF", size=14)
                    heading_cell.alignment = Alignment(horizontal="center", vertical="center")
                    heading_cell.fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')  # Dark navy blue background

                    # Define the background fill for 'dataUsageGB' column
                    light_yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

                    # Get the index of the 'dataUsageGB' column (4th column in this case)
                    data_usage_column_index = df.columns.get_loc('dataUsageGB') + 1  # Pandas index is 0-based, Excel is 1-based
                    
                    # Apply background color to 'dataUsageGB' column (4th column)
                    for row in worksheet.iter_rows(min_row=3, min_col=data_usage_column_index, max_col=data_usage_column_index):
                        for cell in row:
                            cell.fill = light_yellow_fill

                    # Adjust column widths based on content and header
                    for col_num, column_cells in enumerate(worksheet.columns, 1):
                        max_length = 0
                        column = get_column_letter(col_num)
                        
                        for cell in column_cells:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass

                        adjusted_width = (max_length + 2) if max_length < 30 else 30  # Adjust maximum column width to 30 characters
                        worksheet.column_dimensions[column].width = adjusted_width
                    output_with_formatting = BytesIO()
                    workbook.save(output_with_formatting)
                    output_with_formatting.seek(0)

                    report_file = Downloaded_reports(
                        report_type=report_type,
                        kwargs=self.change_kwargs(filter_kwargs)
                    )
                    report_file.file.save(f'zero_usage_billed_data_report_{month}_{year}.xlsx', ContentFile(output_with_formatting.getvalue()))
                    return Response({"data": report_file.file.url, "message" : "Excel file generated sucessfully!"}, status=status.HTTP_200_OK)

                    # response = HttpResponse(output_with_formatting, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    # response['Content-Disposition'] = f'attachment; filename=zero_usage_billed_data_report_{month}_{year}.xlsx'

                    # return response

                except Exception as e:
                    print(e)
                    return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            elif type == "getTop10byWeekExcel":
                try:
                    filtered_data = Report_Unbilled_Data.objects.filter(**filter_kwargs)

                    week_data = {}
                    for record in filtered_data:
                        week = record.Week
                        if week not in week_data:
                            week_data[week] = []
                        week_data[week].append(record)

                    # Create an Excel workbook and worksheet
                    output = BytesIO()
                    workbook = Workbook()

                    # Apply formatting for titles and headers
                    title_font = Font(bold=True, size=12)
                    header_font = Font(bold=True, color="FFFFFF", size=14)  # White text for the main header
                    light_yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
                    light_blue_fill = PatternFill(start_color="FF000066", end_color="FF000066", fill_type="solid")  # Blue header
                    main_header_fill = PatternFill(start_color="FF000066", end_color="FF000066", fill_type="solid")  # Dark blue bg for the main header

                    # Create the worksheet and set column offset for side-by-side tables
                    worksheet = workbook.active
                    worksheet.title = 'Top 10 Usage Report'
                    start_column = 1  # Initial starting column for the first table
                    row_offset = 3  # Start from the 3rd row (after the main header)

                    # Add Main Header for the report (e.g., September Usage Report)
                    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)  # Merge across several columns
                    worksheet.cell(row=1, column=1).value = f"{month} Usage Report (Weekly)"
                    worksheet.cell(row=1, column=1).font = header_font
                    worksheet.cell(row=1, column=1).fill = main_header_fill
                    worksheet.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

                    # Function to get the date with data for Monday, Wednesday, or Friday of a specific week
                    def get_valid_week_date(week_records):
                        week_dates = {'Monday': None, 'Wednesday': None, 'Friday': None}

                        for rec in week_records:
                            date_obj = datetime.strptime(rec.Date, '%Y-%m-%d')
                            weekday = date_obj.strftime('%A')
                            if weekday in week_dates and not week_dates[weekday]:
                                week_dates[weekday] = rec.Date
                        
                        # Return the first available date in order: Monday, Wednesday, Friday
                        return week_dates['Monday'] or week_dates['Wednesday'] or week_dates['Friday']

                    for week, records in week_data.items():
                        # Sort and get top 10 usage records for the week
                        records_sorted = sorted(records, key=lambda x: float(x.Usage), reverse=True)[:10]

                        # Get the valid date for the week (Monday, Wednesday, or Friday)
                        selected_date = get_valid_week_date(records_sorted)

                        if not selected_date:
                            continue  # If no valid date, skip this week

                        # Prepare data for the table
                        report_data = [
                            {'Wireless_Number': rec.Wireless_Number, 'User_Name': rec.User_Name, 'Usage': rec.Usage}
                            for rec in records_sorted
                        ]
                        df = pd.DataFrame(report_data)

                        # Write the title, week date, and vendor
                        worksheet.merge_cells(start_row=row_offset, start_column=start_column, end_row=row_offset, end_column=start_column)
                        worksheet.cell(row=row_offset, column=start_column).value = f"Top 10 Users"
                        worksheet.cell(row=row_offset, column=start_column).font = title_font

                        worksheet.cell(row=row_offset, column=start_column + 1).value = f"Week {week} - {selected_date}"
                        worksheet.cell(row=row_offset, column=start_column + 1).font = title_font
                        worksheet.cell(row=row_offset, column=start_column + 1).alignment = Alignment(horizontal="center")

                        # Write the vendor in the third column
                        worksheet.cell(row=row_offset, column=start_column + 2).value = f"Vendor: {vendor}"
                        worksheet.cell(row=row_offset, column=start_column + 2).font = title_font
                        worksheet.cell(row=row_offset, column=start_column + 2).alignment = Alignment(horizontal="center")

                        # Write the headers
                        headers = ['Wireless_Number', 'User_Name', 'Usage']
                        for col_num, header in enumerate(headers, start=start_column):
                            cell = worksheet.cell(row=row_offset + 1, column=col_num)
                            cell.value = header
                            cell.fill = light_blue_fill
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.alignment = Alignment(horizontal="center")

                        # Fill the data into the table
                        for row_num, row_data in enumerate(df.itertuples(index=False), row_offset + 2):
                            for col_num, value in enumerate(row_data, start=start_column):
                                cell = worksheet.cell(row=row_num, column=col_num)
                                cell.value = value
                                cell.fill = light_yellow_fill

                        # Adjust column widths dynamically based on content
                        for col_num in range(start_column, start_column + len(headers)):
                            max_len = max(len(str(cell.value)) for cell in worksheet[get_column_letter(col_num)])
                            worksheet.column_dimensions[get_column_letter(col_num)].width = max_len + 5

                        # Move start_column to the next set of columns for the next week's table
                        start_column += 5  # Shift the starting column by 5 to leave space between tables

                    # Save the workbook to the output stream
                    workbook.save(output)
                    output.seek(0)

                    report_file = Downloaded_reports(
                        report_type = report_type,
                        kwargs = self.change_kwargs(filter_kwargs)
                    )
                    report_file.file.save(f'top_10_usage_report_{month}.xlsx', ContentFile(output.getvalue()))
                    return Response({"data": report_file.file.url, "message" : "Excel file generated successfully!"}, status=status.HTTP_200_OK)
                    # Create the HTTP response with the Excel file
                    # response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    # response['Content-Disposition'] = f'attachment; filename=top_10_usage_report_{month}.xlsx'

                    # return response
                
                except Exception as e:
                    return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            elif type == "getTop10BilledUsersExcel":
                try:
                    filtered_data = Report_Billed_Data.objects.filter(**filter_kwargs).annotate(
                        usage_float=Cast('Data_Usage_GB', FloatField())
                    ).order_by('-usage_float')[:10]

                    # Create a list of dictionaries to store the data in the required format
                    report_data = []
                    for record in filtered_data:
                        report_data.append({
                            "accountNumber": record.Account_Number,
                            'wirelessNumber': record.Wireless_Number,
                            'userName': record.User_Name,
                            'dataUsageGB': float(record.Data_Usage_GB),
                        })

                    # Convert the data to a pandas DataFrame
                    df = pd.DataFrame(report_data, columns=['accountNumber', 'wirelessNumber', 'userName', 'dataUsageGB'])

                    # Create a bar graph for the top 10 data usage
                    plt.figure(figsize=(10, 6))
                    df.plot(kind='bar', x='userName', y='dataUsageGB', legend=False)
                    plt.title(f'Top 10 Data Usage for Vendor: {vendor}')
                    plt.xlabel('User Name')
                    plt.ylabel('Data Usage (GB)')
                    plt.xticks(rotation=45, ha='right')
                    plt.tight_layout()

                    # Save the plot to a BytesIO object
                    plot_output = BytesIO()
                    plt.savefig(plot_output, format='png')
                    plot_output.seek(0)

                    # Create a new BytesIO output for the Excel file
                    output = BytesIO()

                    # Write the DataFrame to an Excel file using pandas
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Top 10 Usage Billed Data Report')

                    # Reopen the Excel file for additional formatting with openpyxl
                    output.seek(0)
                    workbook = load_workbook(output)
                    worksheet = workbook['Top 10 Usage Billed Data Report']

                    # Insert the plot image into the worksheet
                    img = Image(plot_output)
                    worksheet.add_image(img, 'F2')

                    # Insert a dynamic heading for the top of the report
                    main_heading = f'{month} Top 10 Data Usage ({vendor})'
                    worksheet.insert_rows(1)
                    worksheet.merge_cells('A1:D1')
                    heading_cell = worksheet['A1']
                    heading_cell.value = main_heading
                    heading_cell.font = Font(bold=True, color="FFFFFF", size=14)
                    heading_cell.alignment = Alignment(horizontal="center", vertical="center")
                    heading_cell.fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')

                    # Apply formatting to the 'dataUsageGB' column and adjust column widths
                    dark_navy_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
                    white_font = Font(color="FFFFFF")
                    for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=4, max_col=4):
                        for cell in row:
                            cell.fill = dark_navy_fill
                            cell.font = white_font

                    for col_num, column_cells in enumerate(worksheet.columns, 1):
                        max_length = 0
                        column = get_column_letter(col_num)
                        for cell in column_cells:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        adjusted_width = (max_length + 2) if max_length < 30 else 30
                        worksheet.column_dimensions[column].width = adjusted_width

                    # Save the workbook with formatting
                    output_with_formatting = BytesIO()
                    workbook.save(output_with_formatting)
                    output_with_formatting.seek(0)

                    report_file = Downloaded_reports(
                        report_type = report_type,
                        kwargs = self.change_kwargs(filter_kwargs)
                    )
                    report_file.file.save(f'top_10_usage_billed_data_report_{month}_{year}.xlsx', ContentFile(output_with_formatting.getvalue()))
                    return Response({"data": report_file.file.url, "message" : "Excel file generated successfully!"}, status=status.HTTP_200_OK)

                    # Return the response as a downloadable Excel file
                    # response = HttpResponse(output_with_formatting, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    # response['Content-Disposition'] = f'attachment; filename=top_10_usage_billed_data_report_{month}_{year}.xlsx'

                    # return response
                
                except Exception as e:
                    return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            else:
                return Response({"message": "Invalid request"}, status=status.HTTP_400_BAD_REQUEST)

    def put(self, request, pk, *args, **kwargs):
        return Response(
            {"message": "Not Implemented"},status=status.HTTP_400_BAD_REQUEST
        )
    def delete(self, request, pk, *args, **kwargs):
        pass