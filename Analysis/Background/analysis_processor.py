import sqlite3
import json, re
from pathlib import Path
import time, os, zipfile
import pandas as pd
from io import BytesIO, StringIO
import pdfplumber
from ..Scripts.verizon import PDFExtractor, First, Model1, Model2, Model3, Model4
from ..Scripts.atnt import first_page_extractor, Att,process_all
from ..Scripts.t_mobile_1 import get_all_dataframes_type_1
from ..Scripts.t_mobile_2 import extract_text_from_t_mobile_2
batch_count = '000'
from datetime import datetime

def check_tmobile_type(Billpath):
    print("def check_tmobile_type")
    Lines = []
    with pdfplumber.open(Billpath) as pdf:
        for i, page in enumerate(pdf.pages):
            if i == 0:
                page_text = page.extract_text()
                lines = page_text.split('\n')
                Lines.extend(lines)
            else:
                break

    if 'Bill period Account Invoice Page' in Lines[0]:
        print("Type1 : Bill period Account Invoice Page")
        return 2
    elif 'Your Statement' in Lines[0]:
        print("Type2 : Your Statement")
        return 1
    else:
        None

def extract_data_from_pdf(pdf_path, vendor_nm, pdf_filename):
    print("def extract_data_from_pdf")
    at_vendor_list = ['AT&T', 'AT n T', 'ATT', 'at&t', 'att', 'at n t']
    zippath = None
    t_mobile_list = ['T_Mobile', 't-mobile', 'T-mobile', 't_mobile']
    if 'mobile' in str(vendor_nm).lower():
        print("vendor name is mobile")
        print(vendor_nm)
        print("tmobile list")
        tmobile_type = check_tmobile_type(pdf_path)
        if tmobile_type == 1:
            base_data_df,a,b,c,d,datess,acc_nos,zpath = get_all_dataframes_type_1(pdf_path) # extra calls (a,b,c,d) to prevent error dont remove it
            zippath = zpath
        elif tmobile_type == 2:
            print("tmobile type 2")
            base_data_df,datess,acc_nos,zpath = extract_text_from_t_mobile_2(pdf_path)
            print("tmobile typ2 done")
            zippath = zpath
        else:
            return {'message': 'Invalid tmobile type', 'Error' : 1}
        # base_data_df,pdf_df,duplicate_df,in_valid_df,usage_df,datess,acc_nos = extract_text_from_t_mobile(pdf_path)
        data_dict = base_data_df.to_dict(orient='records')
        data_dict = data_dict[0]
        #  acc_info = data_dict['accountnumber']
        acc_info = acc_nos
        bill_date_info = datess
        data_dict['vendor'] = vendor_nm
        data_dict['pdf_path'] = pdf_path
        data_dict['pdf_filename'] = pdf_filename
    elif 'verizon' in str(vendor_nm).lower():
        print("vendor name is ver")
        lines_to_extract = [2, 3, 4, 5]
        extractor = PDFExtractor(pdf_path)
        extractor.extract_data()
        extractor.process_pdf(lines_to_extract)
        data = extractor.get_result_df()
        acc_info = extractor.get_accounts_info()
        bill_date_info = extractor.get_bill_date()
        data_dict = data.to_dict(orient='records')
        for entry in data_dict:
            entry['vendor'] = vendor_nm
            entry['pdf_path'] = pdf_path
            entry['pdf_filename'] = pdf_filename
    else:
        print("vendor name is at")
        print("at&t list")
        first_obj = first_page_extractor(pdf_path)
        data_dict = first_obj.first_page_data_func()
        acc_info = first_obj.get_acc_info()
        bill_date_info = first_obj.get_bill_date_info()
        data_dict['vendor'] = vendor_nm
        data_dict['pdf_path'] = pdf_path
        data_dict['pdf_filename'] = pdf_filename
    
    return data_dict, acc_info, bill_date_info, zippath

def extract_total_pdf_data(pdf_path, acc_info, company_nm, vendor_nm):
    print("def extract_total_pdf_data")
    at_vendor_list = ['AT&T', 'AT n T', 'ATT', 'at&t', 'att', 'at n t']
    t_mobile_list = ['T_Mobile', 't-mobile', 'T-mobile', 't_mobile']
    total_dict = None
    result_df = None
    if 'mobile' in str(vendor_nm).lower():
        print("vendor name is mobile")
        tmobile_type = check_tmobile_type(pdf_path)
        if tmobile_type == 1:
            pdf_df,duplicate_df,c,d,usage_df,a,b, zippath = get_all_dataframes_type_1(pdf_path) # extra calls (a,b,c,d) to prevent error dont remove it
            print("data got", pdf_df,duplicate_df,usage_df, zippath)
        elif tmobile_type == 2:
            print("tmobile type 2")
            pdf_df,duplicate_df,usage_df, zippath = extract_text_from_t_mobile_2(pdf_path)
            print("data got",pdf_df,duplicate_df,usage_df, zippath)
        else:
            return {'message': 'Invalid tmobile type', 'Error' : 1}
        total_dict = pdf_df.to_dict(orient='records')
        print("duplicate df=====", duplicate_df)
        print("usage df=====",usage_df)
        if 'invoicenumber' in duplicate_df.to_dict():
            duplicate_df = duplicate_df.rename(columns={'invoicenumber': 'Wireless Number'})
        elif 'wireless number' in duplicate_df.to_dict():
            duplicate_df = duplicate_df.rename(columns={'wireless number': 'Wireless Number'})
        if 'wireless number' in usage_df.to_dict():
            usage_df = usage_df.rename(columns={'wireless number': 'Wireless Number'})
        
        dup_df = pd.merge(duplicate_df,usage_df,on='Wireless Number',how='inner')
        result_df = dup_df
    elif 'verizon' in str(vendor_nm).lower():
        print("vendor name is ver")
        extractor = Model4(pdf_path, acc_info)
        result,tmp_df = extractor.process_pdf()
        temp_result_df = result
        df_unique = temp_result_df.drop_duplicates(subset=['Wireless_number'])
        df_unique_dict = df_unique.to_dict(orient='records')
        total_dict = df_unique_dict
        result_df = result
    else:
        print("vendor name is att")
        df_unique,final_df,usage_df,ini_df = process_all(pdf_path)
        are_equal = df_unique.equals(final_df)
        df_unique_dict = df_unique.to_dict(orient='records')
        if are_equal:
            print('goat-goat')
            final_df.rename(columns={'Monthly charges Plan':'Monthly_Charges'},inplace=True)
        else:
            final_df = pd.merge(final_df,usage_df,on='Wireless_number',how='left')
        total_dict = df_unique_dict
        result_df = final_df
        result_df.rename(columns={'Plans_y':'Plans','Voice_Plan_Usage ':'Voice_Plan_Usage'},inplace=True)
    for entry in total_dict:
        entry['company'] = company_nm
        entry['vendor'] = vendor_nm
    res_data_dict = result_df.to_dict(orient='records')
    return res_data_dict, total_dict

def generate_plan_xlsx(data,user_id):
    print('def generate_plan_xlsx')
    print('holllllllllllllllllllaaaaaaaaaaa')
    df = data
    print(df.columns)
    print('df_columns',df.columns)
    print('this is df',df)
    if 'Item_Description' not in df.columns:
        df['Item Description'] = None
    if 'Item_Category' not in df.columns:
        df['Item Category'] = None
    if 'Charges' not in df.columns:
        df['Charges'] = None
    df.columns = df.columns.str.replace('_', ' ').str.title()
    df_unique = df.drop_duplicates(subset=['Wireless Number'])
    print(df_unique.columns)
    print("++++++++++++++++++++++_______________________")
    table_1_data = df_unique
    table_Bills_Details = df_unique
    table_2_data = df_unique
    unique_plans = table_2_data['Plans'].unique()
    plan_discount_costs = {}
    for plan in unique_plans:
        try:
            discount_data = table_2_data[(table_2_data['Plans'] == plan) &
                                        table_2_data['Item Description'].str.contains('Discount', na=False) &
                                        (table_2_data['Charges'] < 0)]
        except:
            discount_data = table_2_data[(table_2_data['Plans'] == plan) &
                                        table_2_data['Item Description'].str.contains('Discount', na=False) &
                                        (table_2_data['Charges'].str.replace('$','').astype(float) < 0)]
        discount_costs = discount_data['Charges'].unique()

        plan_discount_costs[plan] = discount_costs

    unique_plans_df = pd.DataFrame(unique_plans, columns=['Unique Plans'])
    discount_totals = {}

    # Iterating over each unique plan
    for plan in unique_plans:
        # Filtering data for the specific plan and rows with 'Discount' in "Item Description"
        plan_data = table_2_data[(table_2_data['Plans'] == plan) & table_2_data['Item Description'].str.contains('Discount', na=False)]

        # Calculating the total discount for the plan
        total_discount = plan_data['Charges'].astype(float).sum()

        # Storing the total discount in the dictionary
        discount_totals[plan] = total_discount

    # Converting the dictionary to a DataFrame for better visualization
    discount_totals_df = pd.DataFrame(list(discount_totals.items()), columns=['Plans', 'Total Discount'])
    discount_data = table_2_data[table_2_data['Item Description'].str.contains('Discount', na=False)]

    # Grouping the records by the Plans column and identifying if any line of service has a discount
    grouped_discount_data = discount_data.groupby('Plans').apply(lambda group: group['Charges'].astype(float).sum())

    # Getting the unique plans with discounts
    plans_with_discounts = grouped_discount_data.index.tolist()

    # For each plan with discounts, checking if all lines with that plan have the discount applied
    missing_discounts = []
    for plan in plans_with_discounts:
        total_lines_with_plan = table_2_data[table_2_data['Plans'] == plan]
        total_lines_with_discount = discount_data[discount_data['Plans'] == plan]

        if len(total_lines_with_plan) != len(total_lines_with_discount):
            missing_lines = total_lines_with_plan[~total_lines_with_plan['Wireless Number'].isin(total_lines_with_discount['Wireless Number'])]
            missing_discounts.extend(missing_lines[['User Name', 'Wireless Number', 'Plans']].to_dict('records'))

    # Convert the results into a DataFrame
    missing_discounts_df = pd.DataFrame(missing_discounts)

    # Print the results
    missing_discounts_df = missing_discounts_df.drop_duplicates(subset=['Wireless Number', 'Plans'], ignore_index=True)
    discount_data = table_2_data[table_2_data['Item Description'].str.contains('Discount', na=False)]

    # Grouping the records by the Plans column and identifying if any line of service has a discount
    grouped_discount_data = discount_data.groupby('Plans').apply(lambda group: group['Charges'].astype(float).sum())

    # Getting the unique plans with discounts
    plans_with_discounts = grouped_discount_data.index.tolist()

    # For each plan with discounts, checking if all lines with that plan have the discount applied
    missing_discounts = []
    for plan in plans_with_discounts:
        total_lines_with_plan = table_2_data[table_2_data['Plans'] == plan]
        total_lines_with_discount = discount_data[discount_data['Plans'] == plan]

        if len(total_lines_with_plan) != len(total_lines_with_discount):
            missing_lines = total_lines_with_plan[~total_lines_with_plan['Wireless Number'].isin(total_lines_with_discount['Wireless Number'])]
            missing_discounts.extend(missing_lines[['User Name', 'Wireless Number', 'Plans']].to_dict('records'))

    # Calculate the potential missed discount for each missing wireless number
    for record in missing_discounts:
        plan = record['Plans']
        if plan in plans_with_discounts:
            average_discount = discount_data[discount_data['Plans'] == plan]['Charges'].astype(float).mean()
            record['Missed Discount'] = average_discount

    # Convert the results into a DataFrame
    missing_discounts_df = pd.DataFrame(missing_discounts)
    missing_discounts_df = missing_discounts_df.drop_duplicates(subset=['Wireless Number', 'Plans'], ignore_index=True)
    data_device_criteria = table_2_data['Plans'].str.contains('Data Device|MIFI|Tablet|Data Dvc', case=False, na=False)
    data_devices = table_2_data[data_device_criteria]

    # Filtering basic based on plan names
    Basic_criteria = table_2_data['Plans'].str.contains('Basic', case=False, na=False)
    Basic = table_2_data[Basic_criteria]

    # Filtering smartphones based on plan names
    smartphone_criteria = table_2_data['Plans'].str.contains('Smartphone|Smartphn|Smrtphn|Phone', case=False, na=False)
    smartphones = table_2_data[smartphone_criteria]
    print('vavava',table_1_data.columns)
    print('datatata',data_devices.columns)
    data_device_usage = table_1_data[table_1_data['Wireless Number'].isin(data_devices['Wireless Number'])][['Wireless Number', 'Data Usage']]
    Basic_usage = table_1_data[table_1_data['Wireless Number'].isin(Basic['Wireless Number'])][['Wireless Number', 'Data Usage']]
    smartphone_usage = table_1_data[table_1_data['Wireless Number'].isin(smartphones['Wireless Number'])][['Wireless Number', 'Data Usage']]
    # Convert the 'Data Usage' column to numeric values by removing the 'GB' suffix
    try:
        data_device_usage['Data Usage'] = data_device_usage['Data Usage'].str.replace('GB', '', regex=False).replace('NA', 0).astype(float)
        Basic_usage['Data Usage'] = Basic_usage['Data Usage'].str.replace('GB', '', regex=False).replace('NA', 0).astype(float)
        smartphone_usage['Data Usage'] = smartphone_usage['Data Usage'].str.replace('GB', '', regex=False).replace('NA', 0).astype(float)
    except:
        data_device_usage['Data Usage'] = data_device_usage['Data Usage'].astype(float)
        Basic_usage['Data Usage'] = Basic_usage['Data Usage'].astype(float)
        smartphone_usage['Data Usage'] = smartphone_usage['Data Usage'].astype(float)

    # Recalculate the average data usage for each line of service
    data_device_usage['Average Data Usage'] = data_device_usage['Data Usage'].mean()
    Basic_usage['Average Data Usage'] = Basic_usage['Data Usage'].mean()
    smartphone_usage['Average Data Usage'] = smartphone_usage['Data Usage'].mean()

    data_device_usage_sorted = data_device_usage[data_device_usage['Data Usage'] != 0].sort_values(by=['Data Usage'], ascending=False)

    # For smartphones: First, we need to retrieve voice and text usage columns from the original data
    smartphone_usage['Voice Usage'] = table_1_data['Voice Plan Usage']
    smartphone_usage['Text Usage'] = table_1_data['Messaging Usage']

    # Filter out lines with 0 data usage and custom sort using Data usage, Voice, and Text columns
    # smartphone_usage_sorted = smartphone_usage[smartphone_usage['Data Usage'] != 0].sort_values(by=['Data Usage', 'Voice Usage', 'Text Usage'], ascending=False)
    # print(temp_smart)
    print('__________________))))))))))))))))))(((((((((((((((((((((((((1)))))))))))))))))))))))))')
    try:
        smartphone_usage_sorted = smartphone_usage.reset_index()[smartphone_usage['Data Usage'] != 0].sort_values(by=['Data Usage', 'Voice Usage', 'Text Usage'], ascending=False)
    except:
        smartphone_usage_sorted = smartphone_usage[smartphone_usage['Data Usage'] != 0].sort_values(by=['Data Usage', 'Voice Usage', 'Text Usage'], ascending=False)
        data_device_zero_usage = data_device_usage[data_device_usage['Data Usage'] == 0]
        smartphone_zero_usage = smartphone_usage[smartphone_usage['Data Usage'] == 0]
        Basic_usage = Basic_usage[Basic_usage['Data Usage'] == 0]

    # Remove the 0 usage devices from the sorted datasets
    data_device_usage_sorted = data_device_usage_sorted[data_device_usage_sorted['Data Usage'] != 0]
    smartphone_usage_sorted = smartphone_usage_sorted[smartphone_usage_sorted['Data Usage'] != 0]
    data_device_usage_sorted = data_device_usage_sorted.sort_values(by='Average Data Usage', ascending=False)
    smartphone_usage_sorted = smartphone_usage_sorted.sort_values(by='Average Data Usage', ascending=False)
    plan_costs = []
    print('__________________))))))))))))))))))(((((((((((((((((((((((((2)))))))))))))))))))))))))')
    print(table_2_data['Plans'])
    unique_plan = table_2_data["Plans"].unique()
    for plan in unique_plan:
            # print(plan)
            # item_list = list(table_2_data['Item Category'])
            # print(item_list)
            # temp_mon_data = table_2_data['Item Category'] == 'Monthly Charges'
            # print(temp_mon_data)
            # temp_df = table_2_data[temp_mon_data]
            if plan != 'NA':
                for index, row in df.iterrows():
                    plan_costs.append({
                        'Plans': plan,
                        'Monthly Charge': row['Monthly Charges']
                    })
    print('__________________))))))))))))))))))(((((((((((((((((((((((((3)))))))))))))))))))))))))')
    # Convert the results to a DataFrame for display
    plan_costs_df = pd.DataFrame(plan_costs)
    print('__________________))))))))))))))))))(((((((((((((((((((((((((4)))))))))))))))))))))))))')
    print(plan_costs_df)
    print(plan_costs_df.columns)
    plan_costs_df = plan_costs_df[plan_costs_df['Monthly Charge'] != 'NA']
    plan_costs_df = plan_costs_df.drop_duplicates(subset=['Plans'], ignore_index=True)
    try:
        plan_costs_df['Monthly Charge'] = plan_costs_df['Monthly Charge'].str.replace('$','').astype(float)
    except Exception as e:
        print(e)    
    data_usage = table_1_data[['Wireless Number', 'Data Usage']]
    try:
        data_usage['Data Usage'] = data_usage['Data Usage'].str.replace('GB', '', regex=False).replace('NA', '0').astype(float)
    except:
        data_usage['Data Usage'] = data_usage['Data Usage'].astype(float)
    user_data_plan = pd.merge(data_usage, table_2_data[['Wireless Number', 'Plans']], on='Wireless Number', how='inner')
    user_data_plan = user_data_plan.drop_duplicates(subset=['Wireless Number', 'Plans'], ignore_index=True)
    user_data_plan = pd.merge(user_data_plan, plan_costs_df[['Plans', 'Monthly Charge']], on='Plans', how='left')
    user_data_plan['Data Usage'] = user_data_plan['Data Usage'].astype(float)
    try:

        user_data_plan['Monthly Charge'] = user_data_plan['Monthly Charge'].astype(float)
    except Exception as e:
        print(e)    
    def extract_data_usage(plan):
        print("def extract_data_usage")
        if isinstance(plan, str):  # Ensure the value is a string
            match = re.search(r'(\d+\.?\d*)(GB|MB)', plan)
            if match:
                value = float(match.group(1))
                if match.group(2) == "MB":
                    value = value / 1000  # Convert MB to GB
                return value
        return None

    plan_costs_df['Data Usage (GB)'] = plan_costs_df['Plans'].apply(extract_data_usage)

    def find_best_plan(row):
        print("def find_best_plan")
        if row['Data Usage'] == 0 or row['Plans'] == 'NA':
            return row['Plans']
        current_plan_data = plan_costs_df[plan_costs_df['Plans'] == row['Plans']].iloc[0]
        potential_plans = plan_costs_df[
            (plan_costs_df['Plans'].str.contains(re.split(r'\d+GB|\d+MB', current_plan_data['Plans'])[0])) &
            (plan_costs_df['Data Usage (GB)'] < current_plan_data['Data Usage (GB)'])
        ].sort_values(by='Data Usage (GB)', ascending=False)
        if potential_plans.empty:
            return row['Plans']
        for _, plan in potential_plans.iterrows():
            if plan['Data Usage (GB)'] >= row['Data Usage']:
                return plan['Plans']
        return row['Plans']

    user_data_plan['Best Plan'] = user_data_plan.apply(find_best_plan, axis=1)
    user_data_plan = pd.merge(user_data_plan, plan_costs_df[['Plans', 'Monthly Charge']], left_on='Best Plan', right_on='Plans', how='left')
    user_data_plan = user_data_plan.rename(columns={'Plans_x': 'Current Plan', 'Plans_y': 'Best Plan Name'})
    print(user_data_plan['Monthly Charge_x'], user_data_plan['Monthly Charge_y'])
    user_data_plan['Monthly Charge Difference'] = user_data_plan['Monthly Charge_x'] - user_data_plan['Monthly Charge_y']

    # Ensure the columns being subtracted contain only numeric values
    user_data_plan['Monthly Charge_x'] = pd.to_numeric(user_data_plan['Monthly Charge_x'], errors='coerce')
    user_data_plan['Monthly Charge_y'] = pd.to_numeric(user_data_plan['Monthly Charge_y'], errors='coerce')

    # Displaying the final dataframe
    user_data_plan = user_data_plan[['Wireless Number', 'Data Usage', 'Current Plan', 'Best Plan', 'Monthly Charge_x', 'Monthly Charge_y', 'Monthly Charge Difference']]
    user_data_plan
    def extract_data_allowance(plan_name):
        print("def extract_data_allowance")
        try:
            if "GB" in plan_name:
                return float(''.join(filter(str.isdigit, plan_name.split("GB")[0])))
            elif "MB" in plan_name:
                return round(float(''.join(filter(str.isdigit, plan_name.split("MB")[0]))) / 1024, 3)  # Convert MB to GB
            else:
                return "Unlimited"
        except:
            pass
    # Function to get total data usage for a plan
    def get_total_data_usage(plan):
        print("def get_total_data_usage")
        Wireless_Numbers = table_2_data[table_2_data['Plans'] == plan]['Wireless Number'].unique()
        data_usage = table_1_data[table_1_data['Wireless Number'].isin(Wireless_Numbers)]['Data Usage']
        try:
            total_usage = data_usage.str.extract(r'(.\d+|\d+.\d+)GB').dropna().astype(float).sum()
        except:
            total_usage = data_usage.dropna().astype(float).sum()
        return total_usage[0]
    
    def int_get_total_data_usage(plan):
        print("def int_get_total_data_usage")
        Wireless_Numbers = table_2_data[table_2_data['Plans'] == plan]['Wireless Number'].unique()
        data_usage = table_1_data[table_1_data['Wireless Number'].isin(Wireless_Numbers)]['Data Usage']
        try:
            total_usage = data_usage.str.extract(r'(.\d+|\d+.\d+)GB').dropna().astype(float).sum()
        except:
            total_usage = data_usage.dropna().astype(float).sum()
        return total_usage

    # Extract unique plan names excluding "NA"
    unique_plans = [plan for plan in table_2_data['Plans'].unique() if plan != "NA"]

    # Create the dataframe with the modifications
    plans_data_allowances = pd.DataFrame({
        'Plan Name': unique_plans,
        'Data Allowance': [extract_data_allowance(plan) for plan in unique_plans]
    })

    # Populate the dataframe with the required data
    plans_data_allowances['Qty of Wireless Numbers'] = plans_data_allowances['Plan Name'].apply(
        lambda plan: table_2_data[table_2_data['Plans'] == plan]['Wireless Number'].nunique())
    try:
        plans_data_allowances['Total Data Allowance'] = plans_data_allowances.apply(
            lambda row: row['Data Allowance'] * row['Qty of Wireless Numbers'] if row['Data Allowance'] != "Unlimited" else "Unlimited", axis=1)
    except:
        plans_data_allowances.dropna(subset=['Data Allowance'], inplace=True)
        plans_data_allowances['Total Data Allowance'] = plans_data_allowances.apply(
            lambda row: row['Data Allowance'] * row['Qty of Wireless Numbers'] if row['Data Allowance'] != "Unlimited" else "Unlimited", axis=1)
    try:
        plans_data_allowances['Total Data Usage'] = plans_data_allowances['Plan Name'].apply(get_total_data_usage)
    except:
        plans_data_allowances['Total Data Usage'] = plans_data_allowances['Plan Name'].apply(int_get_total_data_usage)

    # Create separate DataFrames based on the keywords
    def categorize_plan(plan_name):
        print("def categorize_plan")
        smartphone_keywords = ['Smartphone', 'Smartphn', 'Smrtphn']
        data_device_keywords = ['Data Device', 'Data Dvc']

        if any(keyword in plan_name for keyword in smartphone_keywords):
            return "smartphone"
        elif any(keyword in plan_name for keyword in data_device_keywords):
            return "data_device"
        else:
            return "other"

    # Apply categorize_plan function to categorize plans into smartphone, data_device, and other
    plans_data_allowances['Category'] = plans_data_allowances['Plan Name'].apply(categorize_plan)

    # Calculate the totals
    total_data_allowance = plans_data_allowances[plans_data_allowances['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
    total_data_usage = plans_data_allowances['Total Data Usage'].sum()

    # Append the totals row to the dataframe
    totals_row = pd.DataFrame({
        'Plan Name': ['Total'],
        'Data Allowance': ['NA'],
        'Qty of Wireless Numbers': ['NA'],
        'Total Data Allowance': [total_data_allowance],
        'Total Data Usage': [total_data_usage],
        'Category': ['NA']
    })

    plans_data_allowances = pd.concat([plans_data_allowances, totals_row], ignore_index=True)

    # Display the result
    plans_data_allowances
    smartphone_df = plans_data_allowances[plans_data_allowances['Category'] == "smartphone"].drop(columns=['Category'])

    # Calculate the totals
    total_data_allowance = smartphone_df[smartphone_df['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
    total_data_usage = smartphone_df['Total Data Usage'].sum()

    # Append the totals row to the dataframe
    totals_row = pd.DataFrame({
        'Plan Name': ['Total'],
        'Data Allowance': ['NA'],
        'Qty of Wireless Numbers': ['NA'],
        'Total Data Allowance': [total_data_allowance],
        'Total Data Usage': [total_data_usage],
    })

    # smartphone_df = smartphone_df.append(totals_row, ignore_index=True)
    smartphone_df = pd.concat([smartphone_df,totals_row], ignore_index=True)

    data_device_df = plans_data_allowances[plans_data_allowances['Category'] == "data_device"].drop(columns=['Category'])

    # Calculate the totals
    total_data_allowance = data_device_df[data_device_df['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
    total_data_usage = data_device_df['Total Data Usage'].sum()

    # Append the totals row to the dataframe
    totals_row = pd.DataFrame({
        'Plan Name': ['Total'],
        'Data Allowance': ['NA'],
        'Qty of Wireless Numbers': ['NA'],
        'Total Data Allowance': [total_data_allowance],
        'Total Data Usage': [total_data_usage],
    })

    # data_device_df = data_device_df.append(totals_row, ignore_index=True)
    data_device_df = pd.concat([data_device_df,totals_row], ignore_index=True)

    

    other_df = plans_data_allowances[plans_data_allowances['Category'] == "other"].drop(columns=['Category'])

    # Calculate the totals
    total_data_allowance = other_df[other_df['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
    total_data_usage = other_df['Total Data Usage'].sum()

    # Append the totals row to the dataframe
    totals_row = pd.DataFrame({
        'Plan Name': ['Total'],
        'Data Allowance': ['NA'],
        'Qty of Wireless Numbers': ['NA'],
        'Total Data Allowance': [total_data_allowance],
        'Total Data Usage': [total_data_usage],
    })

    # other_df = other_df.append(totals_row, ignore_index=True)
    other_df = pd.concat([other_df,totals_row], ignore_index=True)

    try:
        ptrt = table_1_data['Data Usage'].str.extract(r'(.\d+|\d+.\d+)GB').astype(float)[0] >= 15.0
    except:
        ptrt = table_1_data['Data Usage'].astype(float) >= 15.0
    try:
        high_data_usage = table_1_data[ptrt]
        # Merge with table_2_data to get the Plans column
        high_data_usage = high_data_usage.merge(table_2_data[['Wireless Number', 'Plans']], on='Wireless Number', how='left')
        # Reset the index
        high_data_usage = high_data_usage.reset_index(drop=True)
    except:
        empty_dict = {}
        high_data_usage = pd.DataFrame(empty_dict)
    # Select the desired columns
    if not high_data_usage.empty:
        result_high_data_usage = high_data_usage[['Wireless Number', 'User Name', 'Monthly Charges', 'Plans_y', 'Data Usage']]
        result_high_data_usage = result_high_data_usage.drop_duplicates(subset=['Wireless Number'], ignore_index=True)
    try:
        zero_data_usage = table_1_data[table_1_data['Data Usage'].str.extract(r'(NA)').replace('NA', '0').astype(float)[0] >= 0.0]
    except:
        # Corrected code
        zero_data_usage = table_1_data[table_1_data['Data Usage'].astype(float) == 0.0]
    zero_data_usage = zero_data_usage.merge(table_2_data[['Wireless Number', 'Plans']], on='Wireless Number', how='left')
    zero_data_usage = zero_data_usage.reset_index(drop=True)
    result_zero_data_usage = zero_data_usage[['Wireless Number', 'User Name', 'Monthly Charges', 'Plans_y', 'Data Usage']]
    result_zero_data_usage = result_zero_data_usage.drop_duplicates(subset=['Wireless Number'], ignore_index=True)
    # Modify the original table_1_data to exclude wireless numbers with data usage >= 15.00 GB
    try:
        lower_data_usage = table_1_data[~table_1_data['Wireless Number'].isin(high_data_usage['Wireless Number'])]
    except:
        if high_data_usage.empty:
            lower_data_usage = table_1_data[table_1_data['Wireless Number']]
    lower_data_usage = lower_data_usage[~lower_data_usage['Wireless Number'].isin(zero_data_usage['Wireless Number'])]

    # Merge with table_2_data to get the Plans column
    lower_data_usage = lower_data_usage.merge(table_2_data[['Wireless Number', 'Plans']], on='Wireless Number', how='left')

    # Reset the index
    lower_data_usage = lower_data_usage.reset_index(drop=True)

    # Select the desired columns
    result_lower_data_usage = lower_data_usage[['Wireless Number', 'User Name', 'Monthly Charges', 'Plans_y', 'Data Usage']]
    result_lower_data_usage = result_lower_data_usage.drop_duplicates(subset=['Wireless Number'], ignore_index=True)
    filtered_table_1_data = table_1_data[~table_1_data['Wireless Number'].isin(high_data_usage['Wireless Number'])]
    filtered_table_1_data = filtered_table_1_data[~filtered_table_1_data['Wireless Number'].isin(zero_data_usage['Wireless Number'])]
    filtered_table_1_data = filtered_table_1_data.reset_index(drop=True)
    filtered_table_2_data = table_2_data[~table_2_data['Wireless Number'].isin(high_data_usage['Wireless Number'])]
    filtered_table_2_data = filtered_table_2_data[~filtered_table_2_data['Wireless Number'].isin(zero_data_usage['Wireless Number'])]
    filtered_table_2_data = filtered_table_2_data.reset_index(drop=True)
    # Function to extract data allowance from plan name
    def extract_data_allowance(plan_name):
        print("def extract_data_allowance")
        if "GB" in plan_name:
            return float(''.join(filter(str.isdigit, plan_name.split("GB")[0])))
        elif "MB" in plan_name:
            return round(float(''.join(filter(str.isdigit, plan_name.split("MB")[0]))) / 1024, 3)  # Convert MB to GB
        else:
            return "Unlimited"

    # Function to get total data usage for a plan
    def get_total_data_usage(plan):
        print("def get_total_data_usage")
        Wireless_Numbers = filtered_table_2_data[filtered_table_2_data['Plans'] == plan]['Wireless Number'].unique()
        data_usage = filtered_table_1_data[filtered_table_1_data['Wireless Number'].isin(Wireless_Numbers)]['Data Usage']
        try:
            total_usage = data_usage.str.extract(r'(.\d+|\d+.\d+)GB').dropna().astype(float).sum()
        except:
            total_usage = data_usage.astype(float).sum()
        return total_usage[0]

    def int_get_total_data_usaged(plan):
        print("def int_get_total_data_usaged")
        Wireless_Numbers = filtered_table_2_data[filtered_table_2_data['Plans'] == plan]['Wireless Number'].unique()
        data_usage = filtered_table_1_data[filtered_table_1_data['Wireless Number'].isin(Wireless_Numbers)]['Data Usage']
        try:
            total_usage = data_usage.str.extract(r'(.\d+|\d+.\d+)GB').dropna().astype(float).sum()
        except:
            total_usage = data_usage.astype(float).sum()
        return total_usage
        
    # Extract unique plan names excluding "NA"
    unique_plans = [plan for plan in filtered_table_2_data['Plans'].unique() if plan != "NA"]

    # Create the dataframe with the modifications
    plans_data_allowances1 = pd.DataFrame({
        'Plan Name': unique_plans,
        'Data Allowance': [extract_data_allowance(plan) for plan in unique_plans]
    })

    # Populate the dataframe with the required data
    plans_data_allowances1['Qty of Wireless Numbers'] = plans_data_allowances1['Plan Name'].apply(
        lambda plan: filtered_table_2_data[filtered_table_2_data['Plans'] == plan]['Wireless Number'].nunique())
    plans_data_allowances1['Total Data Allowance'] = plans_data_allowances1.apply(
        lambda row: row['Data Allowance'] * row['Qty of Wireless Numbers'] if row['Data Allowance'] != "Unlimited" else "Unlimited", axis=1)
    try:
        plans_data_allowances1['Total Data Usage'] = plans_data_allowances1['Plan Name'].apply(get_total_data_usage)
    except:
        plans_data_allowances1['Total Data Usage'] = plans_data_allowances1['Plan Name'].apply(int_get_total_data_usaged)

    # Create separate DataFrames based on the keywords
    def categorize_plan(plan_name):
        print("def categorize_plan")
        smartphone_keywords = ['Smartphone', 'Smartphn', 'Smrtphn']
        data_device_keywords = ['Data Device', 'Data Dvc']

        if any(keyword in plan_name for keyword in smartphone_keywords):
            return "smartphone"
        elif any(keyword in plan_name for keyword in data_device_keywords):
            return "data_device"
        else:
            return "other"

    # Apply categorize_plan function to categorize plans into smartphone, data_device, and other
    plans_data_allowances1['Category'] = plans_data_allowances1['Plan Name'].apply(categorize_plan)

    # Calculate the totals
    total_data_allowance1 = plans_data_allowances1[plans_data_allowances1['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
    total_data_usage1 = plans_data_allowances1['Total Data Usage'].sum()

    # Append the totals row to the dataframe
    totals_row1 = pd.DataFrame({
        'Plan Name': ['Total'],
        'Data Allowance': ['NA'],
        'Qty of Wireless Numbers': ['NA'],
        'Total Data Allowance': [total_data_allowance1],
        'Total Data Usage': [total_data_usage1],
        'Category': ['NA']
    })

    # plans_data_allowances1 = plans_data_allowances1.append(totals_row1, ignore_index=True)
    plans_data_allowances1 = pd.concat([plans_data_allowances1, totals_row1], ignore_index=True)


    smartphone_df1 = plans_data_allowances1[plans_data_allowances1['Category'] == "smartphone"].drop(columns=['Category'])

    # Calculate the totals
    total_data_allowance1 = smartphone_df1[smartphone_df1['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
    total_data_usage1 = smartphone_df1['Total Data Usage'].sum()

    # Append the totals row to the dataframe
    totals_row = pd.DataFrame({
        'Plan Name': ['Total'],
        'Data Allowance': ['NA'],
        'Qty of Wireless Numbers': ['NA'],
        'Total Data Allowance': [total_data_allowance1],
        'Total Data Usage': [total_data_usage1],
    })

    # smartphone_df1 = smartphone_df1.append(totals_row, ignore_index=True)
    smartphone_df1 = pd.concat([smartphone_df1, totals_row], ignore_index=True)

    
    data_device_df1 = plans_data_allowances1[plans_data_allowances1['Category'] == "data_device"].drop(columns=['Category'])

    # Calculate the totals
    total_data_allowance1 = data_device_df1[data_device_df1['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
    total_data_usage1 = data_device_df1['Total Data Usage'].sum()

    # Append the totals row to the dataframe
    totals_row = pd.DataFrame({
        'Plan Name': ['Total'],
        'Data Allowance': ['NA'],
        'Qty of Wireless Numbers': ['NA'],
        'Total Data Allowance': [total_data_allowance1],
        'Total Data Usage': [total_data_usage1],
    })

    # data_device_df1 = data_device_df1.append(totals_row, ignore_index=True)
    data_device_df1 = pd.concat([data_device_df1, totals_row], ignore_index=True)

    other_df1 = plans_data_allowances1[plans_data_allowances1['Category'] == "other"].drop(columns=['Category'])

    # Calculate the totals
    total_data_allowance1 = other_df1[other_df1['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
    total_data_usage1 = other_df1['Total Data Usage'].sum()

    # Append the totals row to the dataframe
    totals_row = pd.DataFrame({
        'Plan Name': ['Total'],
        'Data Allowance': ['NA'],
        'Qty of Wireless Numbers': ['NA'],
        'Total Data Allowance': [total_data_allowance1],
        'Total Data Usage': [total_data_usage1],
    })

    # other_df1 = other_df1.append(totals_row, ignore_index=True)
    other_df1 = pd.concat([other_df1, totals_row], ignore_index=True)

    plan_costs = []
    unique_plan = filtered_table_2_data["Plans"].unique()

    for plan in unique_plans:
            # Filter rows where the 'Item Category' is 'Monthly Charges'
            if plan != 'NA':
                for index, row in df.iterrows():
                    plan_costs.append({
                        'Plans': plan,
                        'Monthly Charge': row['Charges']
                    })

    # Convert the results to a DataFrame for display
    print(plan_costs)
    plan_costs_df1 = pd.DataFrame(plan_costs)
    plan_costs_df1 = plan_costs_df1.drop_duplicates(subset=['Plans'], ignore_index=True)
    try:
        plan_costs_df1['Monthly Charge'] = plan_costs_df1['Monthly Charge'].str.replace('$','').astype(float)
    except:
        pass
    data_usage = filtered_table_1_data[['Wireless Number', 'Data Usage']]
    try:
        data_usage.loc[:, 'Data Usage'] = data_usage['Data Usage'].str.replace('GB', '', regex=False).replace('NA', '0').astype(float)
    except:
        data_usage.loc[:, 'Data Usage'] = data_usage['Data Usage'].astype(float)        
    user_data_plan1 = pd.merge(data_usage, filtered_table_2_data[['Wireless Number', 'Plans']], on='Wireless Number', how='inner')
    user_data_plan1 = user_data_plan1.drop_duplicates(subset=['Wireless Number', 'Plans'], ignore_index=True)
    user_data_plan1 = pd.merge(user_data_plan1, plan_costs_df1[['Plans', 'Monthly Charge']], on='Plans', how='left')
    user_data_plan1['Data Usage'] = user_data_plan1['Data Usage'].astype(float)
    user_data_plan1['Monthly Charge'] = user_data_plan1['Monthly Charge'].astype(float)
    def extract_data_usage(plan):
        print("def extract_data_usage")
        match = re.search(r'(\d+\.?\d*)(GB|MB)', plan)
        if match:
            value = float(match.group(1))
            if match.group(2) == "MB":
                value = value / 1000
            return value
        return None

    plan_costs_df1['Data Usage (GB)'] = plan_costs_df1['Plans'].apply(extract_data_usage)

    def find_best_plan(row):
        print("def find_best_plan")
        if row['Data Usage'] == 0 or row['Plans'] == 'NA':
            return row['Plans']
        current_plan_data = plan_costs_df1[plan_costs_df1['Plans'] == row['Plans']].iloc[0]
        potential_plans = plan_costs_df1[
            (plan_costs_df1['Plans'].str.contains(re.split(r'\d+GB|\d+MB', current_plan_data['Plans'])[0])) &
            (plan_costs_df1['Data Usage (GB)'] < current_plan_data['Data Usage (GB)'])
        ].sort_values(by='Data Usage (GB)', ascending=False)
        if potential_plans.empty:
            return row['Plans']
        for _, plan in potential_plans.iterrows():
            if plan['Data Usage (GB)'] >= row['Data Usage']:
                return plan['Plans']
        return row['Plans']

    user_data_plan1['Best Plan'] = user_data_plan1.apply(find_best_plan, axis=1)
    user_data_plan1 = pd.merge(user_data_plan1, plan_costs_df1[['Plans', 'Monthly Charge']], left_on='Best Plan', right_on='Plans', how='left')
    user_data_plan1 = user_data_plan1.rename(columns={'Plans_x': 'Current Plan', 'Plans_y': 'Best Plan Name'})
    user_data_plan1['Monthly Charge Difference'] = user_data_plan1['Monthly Charge_x'] - user_data_plan1['Monthly Charge_y']

    # Ensure the columns being subtracted contain only numeric values
    user_data_plan1['Monthly Charge_x'] = pd.to_numeric(user_data_plan1['Monthly Charge_x'], errors='coerce')
    user_data_plan1['Monthly Charge_y'] = pd.to_numeric(user_data_plan1['Monthly Charge_y'], errors='coerce')

    # Displaying the final dataframe
    user_data_plan1 = user_data_plan1[['Wireless Number', 'Data Usage', 'Current Plan', 'Best Plan', 'Monthly Charge_x', 'Monthly Charge_y', 'Monthly Charge Difference']]
    def extract_data_allowance(plan_name):
        print("def extract_data_allowance")
        if "GB" in plan_name:
            return float(''.join(filter(str.isdigit, plan_name.split("GB")[0])))
        elif "MB" in plan_name:
            return round(float(''.join(filter(str.isdigit, plan_name.split("MB")[0]))) / 1024, 3)  # Convert MB to GB
        else:
            return "Unlimited"

    # Function to get total data usage for a plan
    def get_total_data_usage(plan):
        print("def get_total_data_usage")
        Wireless_Numbers = user_data_plan1[user_data_plan1['Best Plan'] == plan]['Wireless Number'].unique()
        data_usage = user_data_plan1[user_data_plan1['Wireless Number'].isin(Wireless_Numbers)]['Data Usage']
        data_usage = data_usage.astype(str)
        total_usage = data_usage.str.extract(r'(.\d+|\d+.\d+)').dropna().astype(float).sum()
        return total_usage[0]

    # Extract unique plan names excluding "NA"
    unique_plans = [plan for plan in user_data_plan1['Best Plan'].unique() if plan != "NA"]

    # Create the dataframe with the modifications
    plans_data_allowances2 = pd.DataFrame({
        'Plan Name': unique_plans,
        'Data Allowance': [extract_data_allowance(plan) for plan in unique_plans]
    })

    # Populate the dataframe with the required data
    plans_data_allowances2['Qty of Wireless Numbers'] = plans_data_allowances2['Plan Name'].apply(
        lambda plan: user_data_plan1[user_data_plan1['Best Plan'] == plan]['Wireless Number'].nunique())
    plans_data_allowances2['Total Data Allowance'] = plans_data_allowances2.apply(
        lambda row: row['Data Allowance'] * row['Qty of Wireless Numbers'] if row['Data Allowance'] != "Unlimited" else "Unlimited", axis=1)
    plans_data_allowances2['Total Data Usage'] = plans_data_allowances2['Plan Name'].apply(get_total_data_usage)

    def categorize_plan(plan_name):
        print("def categorize_plan")
        smartphone_keywords = ['Smartphone', 'Smartphn', 'Smrtphn']
        data_device_keywords = ['Data Device', 'Data Dvc']

        if any(keyword in plan_name for keyword in smartphone_keywords):
            return "smartphone"
        elif any(keyword in plan_name for keyword in data_device_keywords):
            return "data_device"
        else:
            return "other"

    plans_data_allowances2['Category'] = plans_data_allowances2['Plan Name'].apply(categorize_plan)

    total_data_allowance2 = plans_data_allowances2[plans_data_allowances2['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
    total_data_usage2 = plans_data_allowances2['Total Data Usage'].sum()

    totals_row1 = pd.DataFrame({
        'Plan Name': ['Total'],
        'Data Allowance': ['NA'],
        'Qty of Wireless Numbers': ['NA'],
        'Total Data Allowance': [total_data_allowance2],
        'Total Data Usage': [total_data_usage2],
        'Category': ['NA']
    })

    plans_data_allowances2 = pd.concat([plans_data_allowances2, totals_row1], ignore_index=True)
    smartphone_df2 = plans_data_allowances2[plans_data_allowances2['Category'] == "smartphone"].drop(columns=['Category'])


    total_data_allowance2 = smartphone_df2[smartphone_df2['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
    total_data_usage2 = smartphone_df2['Total Data Usage'].sum()

    totals_row = pd.DataFrame({
        'Plan Name': ['Total'],
        'Data Allowance': ['NA'],
        'Qty of Wireless Numbers': ['NA'],
        'Total Data Allowance': [total_data_allowance2],
        'Total Data Usage': [total_data_usage2],
    })

    smartphone_df2 = pd.concat([smartphone_df2, totals_row], ignore_index=True)
    data_device_df2 = plans_data_allowances2[plans_data_allowances2['Category'] == "data_device"].drop(columns=['Category'])

    total_data_allowance2 = data_device_df2[data_device_df2['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
    total_data_usage2 = data_device_df2['Total Data Usage'].sum()

    totals_row = pd.DataFrame({
        'Plan Name': ['Total'],
        'Data Allowance': ['NA'],
        'Qty of Wireless Numbers': ['NA'],
        'Total Data Allowance': [total_data_allowance2],
        'Total Data Usage': [total_data_usage2],
    })

    data_device_df2 = pd.concat([data_device_df2, totals_row], ignore_index=True)
    other_df2 = plans_data_allowances2[plans_data_allowances2['Category'] == "other"].drop(columns=['Category'])

    total_data_allowance2 = other_df2[other_df2['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
    total_data_usage2 = other_df2['Total Data Usage'].sum()

    totals_row = pd.DataFrame({
        'Plan Name': ['Total'],
        'Data Allowance': ['NA'],
        'Qty of Wireless Numbers': ['NA'],
        'Total Data Allowance': [total_data_allowance2],
        'Total Data Usage': [total_data_usage2],
    })

    other_df2 = pd.concat([other_df2, totals_row], ignore_index=True)
    total_cost_current = user_data_plan1['Monthly Charge_x'].sum()
    total_cost_best = user_data_plan1['Monthly Charge_y'].sum()
    total_cost_save = user_data_plan1['Monthly Charge Difference'].sum()

    totals_row = pd.DataFrame({
        'Wireless Number': ['Total Charges'],
        'Data Usage': ['NA'],
        'Current Plan': ['NA'],
        'Best Plan': ['NA'],
        'Monthly Charge_x': [total_cost_current],
        'Monthly Charge_y': [total_cost_best],
        'Monthly Charge Difference': [total_cost_save]
    })

    user_data_plan2 = pd.concat([user_data_plan1, totals_row], ignore_index=True)
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows


    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    ws1 = wb.create_sheet(title='15GB+ Data Usage')
    ws2 = wb.create_sheet(title='0 Data Usage')
    ws3 = wb.create_sheet(title='Except 15GB+ && 0 Data Usage')
    ws4 = wb.create_sheet(title='Recomendations using user data')
    ws5 = wb.create_sheet(title='All in One Bucket Analysis')
    ws6 = wb.create_sheet(title='Smartphone Bucket Analysis')
    ws7 = wb.create_sheet(title='Data Device Bucket Analysis')
    result_zero_data_usage['Data Usage'] = result_zero_data_usage['Data Usage'].replace('NA',0).astype(float)
    dataframes = {
        '15GB+ Data Usage': result_high_data_usage if 'result_high_data_usage' in locals() else 'NA',
        '0 Data Usage': result_zero_data_usage,
        'Except 15GB+ && 0 Data Usage': result_lower_data_usage,
        'Recomendations using user data':user_data_plan2,
        'All in One Bucket Analysis': plans_data_allowances2,
        'Smartphone Bucket Analysis': smartphone_df2,
        'Data Device Bucket Analysis': data_device_df2
    }
    # conn = psycopg2.connect(database="mobilityprod", user="postgres", password="GetOutOfJail@2023", host="localhost", port="5432")
    # cursor = conn.cursor()
    # if 'result_high_data_usage' in locals():
    #     result_high_data_usage['user_id'] = user_id
    #     # result_high_data_usage.to_sql('tech_fifteen_plus_gb', con=conn, if_exists='append', index=False, method='multi',chunksize=1000)
    #     high_dict = result_high_data_usage.to_dict(orient='records')
    #     for item in high_dict:
    #         keys = ', '.join(item.keys())
    #         values = ', '.join([f"'{value}'" for value in item.values()])
    #         cursor.execute(f"INSERT INTO tech_fifteen_plus_gb ({keys}) VALUES ({values})")
    # # result_zero_data_usage.to_sql('tech_zero_data_usage', con=conn, if_exists='append', index=False, method='multi',chunksize=1000)
    # result_zero_data_usage.columns = result_zero_data_usage.columns.str.replace(' ','_')
    # zero_dict = result_zero_data_usage.to_dict(orient='records')
    # for item in zero_dict:
    #     keys = ', '.join(item.keys())
    #     values = ', '.join([f"'{value}'" for value in item.values()])
    #     cursor.execute(f"INSERT INTO tech_zero_data_usage ({keys}) VALUES ({values})")
    # result_lower_data_usage.columns = result_lower_data_usage.columns.str.replace(' ','_')
    # lower_dict = result_lower_data_usage.to_dict(orient='records')
    # for item in lower_dict:
    #     keys = ', '.join(item.keys())
    #     values = ', '.join([f"'{value}'" for value in item.values()])
    #     cursor.execute(f"INSERT INTO tech_fifteen_plus_gb_and_zero_data_usage ({keys}) VALUES ({values})")
    # plans_data_allowances2.columns = plans_data_allowances2.columns.str.replace(' ','_')
    # allowance_dict = plans_data_allowances2.to_dict(orient='records')
    # for item in allowance_dict:
    #     keys = ', '.join(item.keys())
    #     values = ', '.join([f"'{value}'" for value in item.values()])
    #     cursor.execute(f"INSERT INTO tech_all_in_one_bucket ({keys}) VALUES ({values})")
    # smartphone_df2.columns = smartphone_df2.columns.str.replace(' ','_')
    # smart_dict = smartphone_df2.to_dict(orient='records')
    # for item in smart_dict:
    #     keys = ', '.join(item.keys())
    #     values = ', '.join([f"'{value}'" for value in item.values()])
    #     cursor.execute(f"INSERT INTO tech_smartphone_bucket ({keys}) VALUES ({values})")
    # data_device_df2.columns = data_device_df2.columns.str.replace(' ','_')
    # device_dict = data_device_df2.to_dict(orient='records')
    # for item in device_dict:
    #     keys = ', '.join(item.keys())
    #     values = ', '.join([f"'{value}'" for value in item.values()])
    #     cursor.execute(f"INSERT INTO tech_device_bucket_analysis ({keys}) VALUES ({values})")   
    # conn.commit()
    total_sheets = []
    dataframes = {key: value for key, value in dataframes.items() if not isinstance(value, str) or value != 'NA'}
    for sheet_name, df in dataframes.items():
        sheet = wb[sheet_name]
        total_sheets.append(sheet)
        for row in dataframe_to_rows(df, index=False, header=True):
            sheet.append(row)

    import os
    directory = 'op_test_dir'
    if not os.path.exists(directory):
        os.makedirs(directory)

    excel_files = total_sheets[:-3]
    recommendations_worksheet = excel_files[-1]
    recommendations_df = pd.DataFrame(recommendations_worksheet.values)
    recommendations_header = recommendations_df.iloc[0]
    recommendations_df = recommendations_df[1:]
    recommendations_df.columns = recommendations_header
    recommendations_df.columns = recommendations_df.columns.str.replace(' ','_')
    recommendations_df['user_id'] = user_id
    merged_df = recommendations_df.copy()

    other_worksheets = excel_files[:-1]
    plan_dict = {
        "Business Unlimited Smartphone":1000,
        "Flex Business Data Device 2GB":2,
        "Business UNL Plus Smartphone":1000,
        "Business UNL Tablet Start":1000,
        "BUS UNL Plus 2.0 Smartphone":1000,
        "Business Unlimited Data":1000,
        "The new Verizon Plan BUS 25GB":25,
        "SMB UNL TLK&TXT 200GB":200,
        "More Evr SMB UNL TLK&TXT 400GB":400,
        "Bus. Unlim Plus with Private WiFi Multi Line":1000,
        "DataConnect 10GB for Mobile Hotspot and Laptop":10,
        "Bus Unl Data 10 GB HS Mobile Internet Gigabytes U n l i m ited -":10,
        "BUSENUNADVVVM4GI":1000,
        "BUSENHUNLADVTAB4GI":1000,
        "BUSENUNADVVVM5GI":1000,
        "MOBSELPOOL1GBIPAD5G":1,
        "MOBSELPOOL5GBIPAD5G":5,
        "MBLSEL10GBDCLTEUIO":10
    }
    pricing_dictionary = {
        "Business Unlimited Smartphone": 45,
        "Flex Business Data Device 2GB": 31.5,
        "Business UNL Plus Smartphone": 40,
        "Business UNL Tablet Start": 30,
        "BUS UNL Plus 2.0 Smartphone": 35,
        "Business Unlimited Data": 45,
        "The new Verizon Plan BUS 25GB": 136,
        "SMB UNL TLK&TXT 200GB": 1140,
        "More Evr SMB UNL TLK&TXT 400GB": 1170,
        "Bus. Unlim Plus with Private WiFi Multi Line":119,
        "DataConnect 10GB for Mobile Hotspot and Laptop":50,
        "Bus Unl Data 10 GB HS Mobile Internet Gigabytes U n l i m ited -":25,
        "BUSENUNADVVVM4GI":45,
        "BUSENHUNLADVTAB4GI":30,
        "BUSENUNADVVVM5GI":45,
        "MOBSELPOOL1GBIPAD5G":20,
        "MOBSELPOOL5GBIPAD5G":50,
        "MBLSEL10GBDCLTEUIO":75
    }

    for worksheet in other_worksheets:
        df = pd.DataFrame(worksheet.values)
        header = df.iloc[0]
        df = df[1:]
        df.columns = header
        try:
            for index, row in df.iterrows():
                wireless_number = row['Wireless_Number']
                
                matching_row = recommendations_df[recommendations_df['Wireless_Number'] == wireless_number]
                
                if not matching_row.empty:
                    for col in df.columns:
                        merged_df.at[matching_row.index[0], col] = row[col]
                else:
                    # merged_df = merged_df.append(row, ignore_index=True)
                    merged_df = pd.concat([merged_df,row], ignore_index=True)

        except:
            for index, row in df.iterrows():
                wireless_number = row['Wireless Number']
                
                matching_row = recommendations_df[recommendations_df['Wireless_Number'] == wireless_number]
                print(matching_row)
                if not matching_row.empty:
                    for col in df.columns:
                        try:
                            merged_df.at[matching_row.index[0], col] = row[col]
                        except:
                            pass
                else:
                    try:
                        # merged_df = merged_df.append(row, ignore_index=True)
                        merged_df = pd.concat([merged_df,row], ignore_index=True)

                    except:
                        merged_df = merged_df.loc[:, ~merged_df.columns.duplicated()]
                        print('qwertyuiopiuytrew')
                        print(merged_df)
                        merged_df = merged_df.reset_index(drop=True)
                        row = row.reset_index(drop=True)
                        # merged_df = merged_df.append(row, ignore_index=True)
                        merged_df = pd.concat([merged_df,row], ignore_index=True)


    merged_df['Current_Plan'] = merged_df.apply(lambda row: row['Plans_y'] if pd.isna(row['Current_Plan']) else row['Current_Plan'], axis=1)
    def best_recommended_plan(data_usage):
        print("def best_recommended_plan")
        if data_usage == 'NA':
            return 'NA'
        data_usage = str(data_usage)
        print(type(data_usage))
        data_usage = data_usage.strip()
        if 'GB' in data_usage:
            data_usage = data_usage.replace('GB', '')
        
        try:
            data_usage = float(data_usage)
        except:
            data_usage = 0

        best_plan = None
        best_value = float('inf')
        
        data_usage = float(data_usage)
        
        for plan, value in plan_dict.items():
            if value > data_usage and value < best_value:
                best_plan = plan
                best_value = value
        
        return best_plan

    merged_df['Best_Plan'] = merged_df['Data_Usage'].apply(best_recommended_plan)
    plan_value_dict = {
        "Current_Plan":"Monthly_Charge_x",
        "Best_Plan":"Monthly_Charge_y"
    }
    for k in plan_value_dict:
        for index, row in merged_df.iterrows():
            current_plan = row[k]
            if current_plan in pricing_dictionary:
                merged_df.at[index, plan_value_dict[k]] = pricing_dictionary[current_plan]
    merged_df['Recommendations'] = ""

    for index, row in merged_df.iterrows():
        current_plan = row['Current_Plan']
        best_plan = row['Best_Plan']

        if best_plan == 'NA':
            merged_df.at[index, 'Recommendations'] = "Verify 0 usage-Cancel"
        elif current_plan == best_plan:
            merged_df.at[index, 'Recommendations'] = ""
        else:
            merged_df.at[index, 'Recommendations'] = "Change Plan"

    merged_df = merged_df.rename(columns={'Wireless_Number': 'Mobile_no',
                                        'Best_Plan': 'Recommended_plan',
                                        'Plans_y': 'Plans',
                                        'Monthly_Charge_x': 'Current_price',
                                        'Monthly_Charge_y': 'Recommended_price'})
    try:
        merged_df = merged_df[['Recommendations', 'User_Name', 'Mobile_no', 'Current_Plan', 'Current_price',
                            'Recommended_plan', 'Recommended_price', 'Plans', 'Data_Usage']]
    except:
        merged_df = merged_df[['Recommendations', 'User Name', 'Mobile_no', 'Current_Plan', 'Current_price',
                            'Recommended_plan', 'Recommended_price', 'Plans', 'Data_Usage']]        
    merged_df['Charge_Difference'] = merged_df['Current_price'] - merged_df['Recommended_price']
    # merged_df_dict = merged_df.to_dict(orient='records')
    # for item in merged_df_dict:
    #     keys = ', '.join(item.keys())
    #     values = ', '.join([f"'{value}'" for value in item.values()])
    #     cursor.execute(f"INSERT INTO tech_Recommendations_table ({keys}) VALUES ({values})")   
    # conn.commit()
    # conn.close()
    w8 = wb.create_sheet(title='Recomendations As Per User Data')
    print('>>>>>>>>>>>><<<<<<<<<<<<')
    print(merged_df.columns)
    merged_df['Data_Usage'] = merged_df['Data_Usage'].replace('NA',0)
    merged_df['User Name'] = 'NA'
    print('hakakakaka')
    print(merged_df)
    dataframess = {'Recomendations As Per User Data':merged_df}
    dataframes = {key: value for key, value in dataframess.items() if not isinstance(value, str) or value != 'NA'}
    for sheet_name, df in dataframes.items():
        sheet = wb[sheet_name]
        for row in dataframe_to_rows(df, index=False, header=True):
            sheet.append(row)
    wb.remove(ws4)
    wb.move_sheet(w8, offset=-3)
    return wb

import io
def workbook_to_bytes(wb):
    print("def workbook_to_bytes")
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

from ..models import Analysis
import shutil
from django.core.files.base import File
def process_analysis_pdf_from_buffer(instance, buffer_data):
    print("Processing analysis PDF from buffer")

    # Extract data from buffer
    pdf_path = buffer_data.get('pdf_path')
    vendor_name = buffer_data.get('vendor_name')
    organization = buffer_data.get('organization')
    remark = buffer_data.get('remark')
    first_name = buffer_data.get('first_name')
    last_name = buffer_data.get('last_name')
    store = buffer_data.get('store')
    pdf_filename = buffer_data.get('pdf_filename')
    user_id = buffer_data.get('user_id')
    company_name = "xyz.ltd"
    acc_info = '1234567'

    # Extract PDF Data
    xlsx_first_dict, acc_info, bill_date_info, zpath = extract_data_from_pdf(pdf_path, vendor_name, pdf_filename)
    pdf_data, unique_pdf_data = extract_total_pdf_data(pdf_path, acc_info, company_name, vendor_name)

    # Convert Data to DataFrame
    anlys_df = pd.DataFrame(pdf_data)
    print(anlys_df.columns)

    # Rename columns for T-Mobile vendor
        
    print(user_id, bill_date_info,organization, remark, first_name, last_name, store, pdf_path, pdf_filename, vendor_name)

    # If bill_date_info is a list, take the first element
    if isinstance(bill_date_info, list):
        bill_date_info = bill_date_info[0]
    
    # Define temporary output directory
    obj = Analysis.objects.get(id=instance.id)
    obj.bill_date_info = bill_date_info
    obj.save()
    if 'mobile' not in str(vendor_name).lower():
        out_directory = 'output_test_dir'
        os.makedirs(out_directory, exist_ok=True)

        # Save DataFrame as Excel file
        file_path1 = os.path.join(out_directory, "output_dataframes.xlsx")
        anlys_df.to_excel(file_path1, index=False)

        # Ensure the file is properly closed before reading it again
        workbook_name = str(instance.uploadBill.name).replace('.pdf', '.xlsx')

        # Save the Excel file to the Django model
        obj = Analysis.objects.get(id=instance.id)
        with open(file_path1, 'rb') as f:
            obj.excel.save(workbook_name, File(f))  # Using save() method ensures proper handling
        obj.is_processed = True
        obj.save()

        shutil.rmtree(out_directory)

        print("File saved successfully and temp directory removed.")
    else:
        anlys_df['User Name'] = 'NA'
        anlys_df.rename(columns={
            'wireless number': 'Wireless_number',
            'item category': 'Item_Category',
            'item description': 'Item_Description',
            'charges': 'Charges',
            'Data Usage': 'Data_Usage',
            'Account Number': 'Account_number',
            'Data Plan': 'Plans',
            'Voice Roaming': 'Voice Plan Usage',
            'Messaging Roaming': 'Messaging Usage',
            'Recurring Charges': 'Monthly_Charges'
        }, inplace=True)
        if zpath:
            print(zpath)
            with open(zpath, 'rb') as  f:
                obj.excel.save(str(zpath), File(f)) 
            obj.is_processed = True
            obj.save()
            os.remove(zpath)
        instance.is_processed = True
        instance.save()
        
    

    